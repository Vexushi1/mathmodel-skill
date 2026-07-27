"""定位项目根目录、校验三轴工作簿契约并写入每问两类中文 Excel。"""
from __future__ import annotations

import math
import os
import re
from pathlib import Path
from typing import Any, Mapping, Sequence

import pandas as pd
import yaml
from openpyxl import load_workbook

INVALID_SHEET_CHARS = set('[]:*?/\\')
PROBLEM_PATTERN = re.compile(r"问题[一二三四五六七八九十百]+")
VALID_WORKBOOK_KINDS = {"solution", "robustness"}

_FALLBACK_SCHEMA: dict[str, Any] = {
    "global_rules": {"empty_worksheet_allowed": False},
    "capability_contract": {
        "allowed": [
            "has_explicit_constraints", "requires_feasibility_check",
            "requires_equilibrium_residual", "requires_conservation_residual",
            "requires_discretization_check", "requires_convergence_diagnostic",
            "requires_out_of_sample_validation", "requires_uncertainty_quantification",
            "requires_leakage_check", "requires_calibration_check",
            "requires_identifiability_check",
        ],
        "required_sheets": {
            "has_explicit_constraints": ["约束违反检查"],
            "requires_feasibility_check": ["约束违反检查"],
            "requires_equilibrium_residual": ["均衡残差"],
            "requires_conservation_residual": ["守恒残差"],
            "requires_discretization_check": ["离散精度"],
            "requires_convergence_diagnostic": ["收敛诊断"],
            "requires_out_of_sample_validation": ["外样本验证"],
            "requires_uncertainty_quantification": ["不确定性区间"],
            "requires_leakage_check": ["泄漏检查"],
            "requires_calibration_check": ["校准结果"],
            "requires_identifiability_check": ["可识别性检查"],
        },
        "legacy_problem_type_fallback": {
            "problem_types": ["optimization", "scheduling"],
            "required_sheets": ["约束违反检查"],
        },
    },
    "solution_workbook": {
        "common_required_sheets": {
            "核心指标": {"required_columns": ["指标", "数值"]},
            "数据审计": {"required_columns": ["等级", "检查项", "信息", "处理方式"]},
        },
        "common_recommended_sheets": {
            "推荐方案": {"required_columns": ["方案"]},
            "明细结果": {"required_columns": ["记录键"]},
            "多算法对比": {"required_columns": ["算法", "目标值", "可行性"]},
            "约束违反检查": {"required_columns": ["约束编号", "约束含义", "违反量", "容差", "是否满足"]},
            "均衡残差": {"required_columns": ["主体或均衡", "残差", "容差", "是否满足"]},
            "守恒残差": {"required_columns": ["守恒量", "残差", "容差", "是否满足"]},
            "离散精度": {"required_columns": ["离散参数", "取值", "目标指标", "相对变化"]},
            "收敛诊断": {"required_columns": ["迭代或样本数", "指标", "数值", "判定"]},
            "外样本验证": {"required_columns": ["划分或窗口", "指标", "数值"]},
            "不确定性区间": {"required_columns": ["指标", "下界", "上界"]},
            "泄漏检查": {"required_columns": ["检查项", "是否通过", "证据"]},
            "校准结果": {"required_columns": ["分组或方法", "预测概率", "实际频率"]},
            "可识别性检查": {"required_columns": ["参数或效应", "检查方法", "结论"]},
            "预测明细": {"required_columns": ["记录键", "真实值", "预测值"]},
            "误差指标": {"required_columns": ["指标", "数值"]},
            "综合评分": {"required_columns": ["对象", "综合评分"]},
            "排序结果": {"required_columns": ["对象", "排名"]},
            "模型指标": {"required_columns": ["指标", "数值"]},
            "预测或分类结果": {"required_columns": ["记录键", "真实标签", "预测标签"]},
            "空间诊断": {"required_columns": ["诊断项", "数值"]},
            "参数估计": {"required_columns": ["参数", "估计值"]},
            "节点结果": {"required_columns": ["节点", "数值"]},
            "边结果": {"required_columns": ["起点", "终点", "数值"]},
            "路径或流结果": {"required_columns": ["路径或流", "数值"]},
        },
        "objective_profiles": {
            "prediction": {"required_any": ["预测明细", "误差指标", "外样本验证"]},
            "evaluation": {"required_any": ["综合评分", "排序结果"]},
            "inference": {"required_any": ["模型指标", "参数估计", "预测或分类结果"]},
            "optimization": {"required_any": ["推荐方案", "明细结果", "核心指标"]},
        },
        "structure_profiles": {
            "spatial": {"required_any": ["空间诊断", "参数估计"]},
            "network": {"required_any": ["节点结果", "边结果", "路径或流结果"]},
        },
        "task_profiles": {
            "prediction": {"required_any": ["预测明细", "误差指标", "外样本验证"]},
            "evaluation": {"required_any": ["综合评分", "排序结果"]},
            "statistics_ml": {"required_any": ["模型指标", "预测或分类结果"]},
            "spatial": {"required_any": ["空间诊断", "参数估计"]},
            "graph_network": {"required_any": ["节点结果", "边结果", "路径或流结果"]},
        },
    },
    "sensitivity_robustness_workbook": {
        "required_any_sheets": ["参数敏感性", "鲁棒性区间", "扰动明细", "算法稳定性", "适用性说明"],
        "sheet_schemas": {
            "参数敏感性": {"required_columns": ["参数", "基准值", "扰动值", "目标指标"]},
            "鲁棒性区间": {"required_columns": ["指标", "下界", "上界"]},
            "扰动明细": {"required_columns": ["扰动编号", "扰动对象", "扰动值", "结果指标"]},
            "算法稳定性": {"required_columns": ["算法", "重复编号", "目标值", "是否可行"]},
            "适用性说明": {"required_columns": ["分析类型", "不适用原因", "替代检验"]},
        },
    },
}


def find_project_root(start: Path) -> Path:
    start = Path(start).resolve()
    current = start.parent if start.is_file() else start
    for candidate in (current, *current.parents):
        if (candidate / "结果数据表").is_dir():
            return candidate
        if candidate.name == "结果数据表":
            return candidate.parent
        if candidate.parent.name == "结果数据表" and PROBLEM_PATTERN.fullmatch(candidate.name):
            return candidate.parent.parent
    return start.parent if start.is_file() else current


def result_data_dir(project_root: Path, problem_name: str) -> Path:
    if not PROBLEM_PATTERN.fullmatch(problem_name):
        raise ValueError("problem_name 应为问题一、问题二等中文名称")
    path = Path(project_root) / "结果数据表" / problem_name
    path.mkdir(parents=True, exist_ok=True)
    return path


def figure_dir(project_root: Path, problem_name: str) -> Path:
    path = result_data_dir(project_root, problem_name) / "图表"
    path.mkdir(parents=True, exist_ok=True)
    return path


def workbook_paths(project_root: Path, problem_name: str) -> tuple[Path, Path]:
    base = result_data_dir(project_root, problem_name)
    return base / f"{problem_name}求解结果.xlsx", base / f"{problem_name}敏感性与鲁棒性结果.xlsx"


def matlab_script_path(project_root: Path, problem_name: str, question_number: int) -> Path:
    if question_number < 1:
        raise ValueError("question_number 必须为正整数")
    return result_data_dir(project_root, problem_name) / f"q{question_number}_plot.m"


def not_applicable_table(
    reason: str,
    analysis_type: str = "敏感性与鲁棒性分析",
    alternative_test: str = "边界条件、有效性或一致性检查",
    evidence_location: str = "",
) -> pd.DataFrame:
    values = [str(item).strip() for item in (analysis_type, reason, alternative_test)]
    if not all(values):
        raise ValueError("分析类型、不适用原因和替代检验均不能为空")
    data = {"分析类型": [values[0]], "不适用原因": [values[1]], "替代检验": [values[2]]}
    if str(evidence_location).strip():
        data["证据位置"] = [str(evidence_location).strip()]
    return pd.DataFrame(data)


def _sheet_name(name: str) -> str:
    safe = "".join("_" if ch in INVALID_SHEET_CHARS else ch for ch in str(name)).strip()
    if not safe:
        raise ValueError("工作表名称不能为空")
    return safe[:31]


def _to_frame(value: Any) -> pd.DataFrame:
    if isinstance(value, pd.DataFrame):
        frame = value.copy()
    elif isinstance(value, Mapping):
        frame = pd.DataFrame([dict(value)])
    elif isinstance(value, (list, tuple)):
        frame = pd.DataFrame(value)
    else:
        frame = pd.DataFrame({"数值": [value]})
    frame = frame.dropna(how="all").reset_index(drop=True)
    if frame.empty:
        raise ValueError("禁止写入空工作表；不适用时请使用 not_applicable_table() 说明原因")
    columns = [str(column).strip() for column in frame.columns]
    if not columns or any(not column for column in columns):
        raise ValueError("工作表至少需要一个非空字段")
    if len(columns) != len(set(columns)):
        raise ValueError("工作表包含重复字段")
    frame.columns = columns
    return frame


def _schema_candidates(explicit: Path | None) -> list[Path]:
    candidates: list[Path] = []
    if explicit is not None:
        candidates.append(explicit)
    if os.getenv("HSK_WORKBOOK_SCHEMA"):
        candidates.append(Path(os.environ["HSK_WORKBOOK_SCHEMA"]))
    for parent in Path(__file__).resolve().parents:
        candidates.append(parent / "core" / "workbook_schema.yaml")
    return candidates


def load_workbook_schema(schema_path: Path | None = None) -> dict[str, Any]:
    for candidate in _schema_candidates(schema_path):
        if candidate.is_file():
            payload = yaml.safe_load(candidate.read_text(encoding="utf-8")) or {}
            if "solution_workbook" in payload and "sensitivity_robustness_workbook" in payload:
                return payload
    return _FALLBACK_SCHEMA


def _required_sheet_schemas(schema: Mapping[str, Any], kind: str) -> tuple[set[str], dict[str, Mapping[str, Any]]]:
    if kind == "solution":
        section = schema["solution_workbook"]
        required = dict(section.get("common_required_sheets", {}))
        all_schemas = {**required, **dict(section.get("common_recommended_sheets", {}))}
        return set(required), all_schemas
    section = schema["sensitivity_robustness_workbook"]
    return set(), dict(section.get("sheet_schemas", {}))


def _conditional_required_sheets(
    schema: Mapping[str, Any], problem_types: Sequence[str], capabilities: Mapping[str, bool] | None,
) -> set[str]:
    contract = schema.get("capability_contract", {})
    required: set[str] = set()
    if capabilities is not None:
        allowed = set(contract.get("allowed", []))
        unknown = sorted(set(capabilities) - allowed)
        if unknown:
            raise ValueError(f"未知验证能力标志: {unknown}")
        for capability, enabled in capabilities.items():
            if enabled:
                required.update(contract.get("required_sheets", {}).get(capability, []))
        return required
    fallback = contract.get("legacy_problem_type_fallback", {})
    if set(problem_types).intersection(fallback.get("problem_types", [])):
        required.update(fallback.get("required_sheets", []))
    return required


def _profile_requirements(
    schema: Mapping[str, Any], objective: str | None,
    structures: Sequence[str], problem_types: Sequence[str],
) -> list[tuple[str, set[str]]]:
    section = schema.get("solution_workbook", {})
    requirements: list[tuple[str, set[str]]] = []
    if objective:
        required = set(section.get("objective_profiles", {}).get(objective, {}).get("required_any", []))
        if required:
            requirements.append((f"objective:{objective}", required))
    for structure in structures:
        required = set(section.get("structure_profiles", {}).get(structure, {}).get("required_any", []))
        if required:
            requirements.append((f"structure:{structure}", required))
    if not objective and not structures:
        profiles = section.get("task_profiles", {})
        for problem_type in problem_types:
            spec = profiles.get(problem_type, {})
            if isinstance(spec, Mapping):
                required = set(spec.get("required_any", []))
                if required:
                    requirements.append((f"legacy:{problem_type}", required))
    return requirements


def _check_required_columns(sheet: str, frame: pd.DataFrame, spec: Mapping[str, Any]) -> None:
    missing = [str(column) for column in spec.get("required_columns", []) if str(column) not in frame.columns]
    if missing:
        raise ValueError(f"工作表“{sheet}”缺少必需字段: {missing}")


def _check_record_keys(sheet: str, frame: pd.DataFrame) -> None:
    for key in ("记录键", "样本键"):
        if key not in frame.columns:
            continue
        series = frame[key]
        if series.isna().any() or series.astype(str).str.strip().eq("").any():
            raise ValueError(f"工作表“{sheet}”的主键字段“{key}”存在空值")
        if series.duplicated().any():
            duplicate = series[series.duplicated()].iloc[0]
            raise ValueError(f"工作表“{sheet}”的主键字段“{key}”存在重复值: {duplicate}")
        break


def _check_finite_numbers(sheet: str, frame: pd.DataFrame) -> None:
    for column in frame.select_dtypes(include="number").columns:
        values = frame[column].dropna()
        if not values.map(lambda value: math.isfinite(float(value))).all():
            raise ValueError(f"工作表“{sheet}”的数值字段“{column}”包含 NaN 以外的非有限值")


def _as_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "是", "满足", "通过"}:
        return True
    if text in {"false", "0", "no", "否", "不满足", "未通过"}:
        return False
    return None


def _check_residual_consistency(sheet: str, frame: pd.DataFrame) -> None:
    value_column = "违反量" if "违反量" in frame.columns else "残差" if "残差" in frame.columns else None
    if value_column is None or "容差" not in frame.columns or "是否满足" not in frame.columns:
        return
    for index, row in frame.iterrows():
        if pd.isna(row[value_column]) or pd.isna(row["容差"]):
            continue
        expected = abs(float(row[value_column])) <= float(row["容差"])
        actual = _as_bool(row["是否满足"])
        if actual is None or actual != expected:
            raise ValueError(f"工作表“{sheet}”第 {index + 2} 行的是否满足与残差/容差不一致")


def _check_missing_value_audit(prepared: Sequence[tuple[str, pd.DataFrame]]) -> None:
    affected = [name for name, frame in prepared if name != "数据审计" and frame.isna().any().any()]
    if not affected:
        return
    audit = next((frame for name, frame in prepared if name == "数据审计"), None)
    if audit is None:
        raise ValueError(f"工作表 {affected} 包含缺失值，但缺少数据审计说明")
    text = " ".join(audit.astype(str).fillna("").to_numpy().ravel()).lower()
    if "缺失" not in text and "missing" not in text:
        raise ValueError(f"工作表 {affected} 包含缺失值，但数据审计未说明缺失处理")


def validate_workbook_tables(
    tables: Mapping[str, Any],
    workbook_kind: str,
    problem_types: Sequence[str] = (),
    capabilities: Mapping[str, bool] | None = None,
    schema_path: Path | None = None,
    *,
    objective: str | None = None,
    structures: Sequence[str] = (),
) -> list[tuple[str, pd.DataFrame]]:
    if workbook_kind not in VALID_WORKBOOK_KINDS:
        raise ValueError(f"workbook_kind 必须为 {sorted(VALID_WORKBOOK_KINDS)}")
    if not tables:
        raise ValueError("没有可写入的结果表")
    prepared: list[tuple[str, pd.DataFrame]] = []
    used: set[str] = set()
    for raw_name, value in tables.items():
        name = _sheet_name(raw_name)
        if name in used:
            raise ValueError(f"工作表名称截断后重复: {name}")
        used.add(name)
        prepared.append((name, _to_frame(value)))
    schema = load_workbook_schema(schema_path)
    required_sheets, sheet_schemas = _required_sheet_schemas(schema, workbook_kind)
    names = {name for name, _ in prepared}
    if workbook_kind == "solution":
        required_sheets.update(_conditional_required_sheets(schema, problem_types, capabilities))
        missing = sorted(required_sheets - names)
        if missing:
            raise ValueError(f"求解工作簿缺少必需工作表: {missing}")
        for profile, required_any in _profile_requirements(schema, objective, structures, problem_types):
            if required_any and not names.intersection(required_any):
                raise ValueError(f"分类剖面“{profile}”至少需要一个专项工作表: {sorted(required_any)}")
    else:
        allowed_any = set(schema["sensitivity_robustness_workbook"].get("required_any_sheets", []))
        if not names.intersection(allowed_any):
            raise ValueError(f"敏感性与鲁棒性工作簿至少需要一个工作表: {sorted(allowed_any)}")
    for name, frame in prepared:
        if name in sheet_schemas:
            _check_required_columns(name, frame, sheet_schemas[name])
        _check_record_keys(name, frame)
        _check_finite_numbers(name, frame)
        _check_residual_consistency(name, frame)
    _check_missing_value_audit(prepared)
    return prepared


def read_workbook_tables(path: Path) -> dict[str, pd.DataFrame]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    tables: dict[str, pd.DataFrame] = {}
    try:
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                raise ValueError(f"工作表“{worksheet.title}”为空")
            headers = ["" if value is None else str(value).strip() for value in rows[0]]
            if any(not header for header in headers):
                raise ValueError(f"工作表“{worksheet.title}”存在空字段名")
            if len(headers) != len(set(headers)):
                raise ValueError(f"工作表“{worksheet.title}”包含重复字段")
            tables[worksheet.title] = pd.DataFrame(rows[1:], columns=headers)
    finally:
        workbook.close()
    return tables


def validate_workbook_file(
    path: Path,
    workbook_kind: str,
    problem_types: Sequence[str] = (),
    capabilities: Mapping[str, bool] | None = None,
    schema_path: Path | None = None,
    *,
    objective: str | None = None,
    structures: Sequence[str] = (),
) -> list[tuple[str, pd.DataFrame]]:
    if not Path(path).is_file():
        raise FileNotFoundError(path)
    return validate_workbook_tables(
        read_workbook_tables(Path(path)), workbook_kind,
        problem_types=problem_types, capabilities=capabilities,
        schema_path=schema_path, objective=objective, structures=structures,
    )


def _infer_workbook_kind(path: Path) -> str | None:
    if "敏感性与鲁棒性" in path.name:
        return "robustness"
    if "求解结果" in path.name:
        return "solution"
    return None


def write_workbook(
    path: Path,
    tables: Mapping[str, Any],
    *,
    workbook_kind: str | None = None,
    problem_types: Sequence[str] = (),
    capabilities: Mapping[str, bool] | None = None,
    schema_path: Path | None = None,
    objective: str | None = None,
    structures: Sequence[str] = (),
) -> Path:
    path = Path(path)
    kind = workbook_kind or _infer_workbook_kind(path)
    if kind is None:
        prepared = [(_sheet_name(name), _to_frame(value)) for name, value in tables.items()]
        for name, frame in prepared:
            _check_record_keys(name, frame)
            _check_finite_numbers(name, frame)
    else:
        prepared = validate_workbook_tables(
            tables, workbook_kind=kind, problem_types=problem_types,
            capabilities=capabilities, schema_path=schema_path,
            objective=objective, structures=structures,
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
        for name, frame in prepared:
            frame.to_excel(writer, sheet_name=name, index=False)
    return path
