"""Shared workbook contract validation for writers, artifact checks and project sync."""
from __future__ import annotations

import math
from pathlib import Path
from typing import Any, Callable, Mapping, Sequence

import pandas as pd
from openpyxl import load_workbook

VALID_WORKBOOK_KINDS = {"solution", "result_analysis"}
WORKBOOK_KIND_ALIASES = {"robustness": "result_analysis"}


def _normalize_kind(kind: str) -> str:
    normalized = WORKBOOK_KIND_ALIASES.get(kind, kind)
    if normalized not in VALID_WORKBOOK_KINDS:
        allowed = sorted(VALID_WORKBOOK_KINDS | set(WORKBOOK_KIND_ALIASES))
        raise ValueError(f"workbook_kind 必须为 {allowed}")
    return normalized


def _as_frame(value: Any) -> pd.DataFrame:
    if isinstance(value, pd.DataFrame):
        frame = value.copy()
    elif isinstance(value, Mapping):
        frame = pd.DataFrame([dict(value)])
    elif isinstance(value, (list, tuple)):
        frame = pd.DataFrame(value)
    else:
        frame = pd.DataFrame({"数值": [value]})
    frame = frame.dropna(how="all").reset_index(drop=True)
    if frame.empty:
        raise ValueError("禁止写入空工作表")
    columns = [str(column).strip() for column in frame.columns]
    if not columns or any(not column for column in columns):
        raise ValueError("工作表至少需要一个非空字段")
    if len(columns) != len(set(columns)):
        raise ValueError("工作表包含重复字段")
    frame.columns = columns
    return frame


def prepare_tables(
    tables: Mapping[str, Any],
    *,
    name_normalizer: Callable[[str], str] | None = None,
) -> list[tuple[str, pd.DataFrame]]:
    if not tables:
        raise ValueError("没有可写入的结果表")
    prepared: list[tuple[str, pd.DataFrame]] = []
    used: set[str] = set()
    for raw_name, value in tables.items():
        name = name_normalizer(str(raw_name)) if name_normalizer else str(raw_name).strip()
        if not name:
            raise ValueError("工作表名称不能为空")
        if len(name) > 31:
            raise ValueError(f"工作表名称超过31字符: {name}")
        if name in used:
            raise ValueError(f"工作表名称重复: {name}")
        used.add(name)
        prepared.append((name, _as_frame(value)))
    return prepared


def read_workbook_tables(path: Path) -> dict[str, pd.DataFrame]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    tables: dict[str, pd.DataFrame] = {}
    try:
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                raise ValueError(f"工作表“{worksheet.title}”为空")
            headers = ["" if value is None else str(value).strip() for value in rows[0]]
            while headers and not headers[-1]:
                headers.pop()
            if not headers or any(not header for header in headers):
                raise ValueError(f"工作表“{worksheet.title}”存在空字段名")
            if len(headers) != len(set(headers)):
                raise ValueError(f"工作表“{worksheet.title}”包含重复字段")
            data = [
                tuple(row[: len(headers)])
                for row in rows[1:]
                if any(value is not None for value in row[: len(headers)])
            ]
            if not data:
                raise ValueError(f"工作表“{worksheet.title}”为空")
            tables[worksheet.title] = pd.DataFrame(data, columns=headers)
    finally:
        workbook.close()
    return tables


def _required_sheet_schemas(
    schema: Mapping[str, Any], kind: str,
) -> tuple[set[str], dict[str, Mapping[str, Any]]]:
    if kind == "solution":
        section = schema["solution_workbook"]
        required = dict(section.get("common_required_sheets", {}))
        all_schemas = {**required, **dict(section.get("common_recommended_sheets", {}))}
        return set(required), all_schemas
    section = schema["result_analysis_workbook"]
    required = dict(section.get("common_required_sheets", {}))
    all_schemas = {**required, **dict(section.get("sheet_schemas", {}))}
    return set(required), all_schemas


def _conditional_required_sheets(
    schema: Mapping[str, Any],
    problem_types: Sequence[str],
    capabilities: Mapping[str, bool] | None,
) -> set[str]:
    contract = schema.get("capability_contract", {})
    required: set[str] = set()
    if capabilities is not None:
        allowed = set(contract.get("allowed", []))
        unknown = sorted(set(capabilities) - allowed)
        if unknown:
            raise ValueError(f"未知验证能力标志: {unknown}")
        for capability, enabled in capabilities.items():
            if enabled:
                required.update(contract.get("required_sheets", {}).get(capability, []))
        return required
    fallback = contract.get("legacy_problem_type_fallback", {})
    if set(problem_types).intersection(fallback.get("problem_types", [])):
        required.update(fallback.get("required_sheets", []))
    return required


def _profile_requirements(
    schema: Mapping[str, Any],
    objective: str | None,
    structures: Sequence[str],
    problem_types: Sequence[str],
) -> list[tuple[str, set[str]]]:
    section = schema.get("solution_workbook", {})
    requirements: list[tuple[str, set[str]]] = []
    if objective:
        required = set(section.get("objective_profiles", {}).get(objective, {}).get("required_any", []))
        if required:
            requirements.append((f"objective:{objective}", required))
    for structure in structures:
        required = set(section.get("structure_profiles", {}).get(structure, {}).get("required_any", []))
        if required:
            requirements.append((f"structure:{structure}", required))
    if not objective and not structures:
        profiles = section.get("task_profiles", {})
        for problem_type in problem_types:
            required = set((profiles.get(problem_type, {}) or {}).get("required_any", []))
            if required:
                requirements.append((f"legacy:{problem_type}", required))
    return requirements


def _check_required_columns(sheet: str, frame: pd.DataFrame, spec: Mapping[str, Any]) -> None:
    missing = [str(column) for column in spec.get("required_columns", []) if str(column) not in frame.columns]
    if missing:
        raise ValueError(f"工作表“{sheet}”缺少必需字段: {missing}")


def _check_record_keys(sheet: str, frame: pd.DataFrame) -> None:
    for key in ("记录键", "样本键"):
        if key not in frame.columns:
            continue
        series = frame[key]
        if series.isna().any() or series.astype(str).str.strip().eq("").any():
            raise ValueError(f"工作表“{sheet}”的主键字段“{key}”存在空值")
        if series.duplicated().any():
            duplicate = series[series.duplicated()].iloc[0]
            raise ValueError(f"工作表“{sheet}”的主键字段“{key}”存在重复值: {duplicate}")
        break


def _check_finite_numbers(sheet: str, frame: pd.DataFrame) -> None:
    for column in frame.select_dtypes(include="number").columns:
        values = frame[column].dropna()
        if not values.map(lambda value: math.isfinite(float(value))).all():
            raise ValueError(f"工作表“{sheet}”的数值字段“{column}”包含 NaN 以外的非有限值")


def _as_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "是", "满足", "通过"}:
        return True
    if text in {"false", "0", "no", "否", "不满足", "未通过"}:
        return False
    return None


def _check_residual_consistency(sheet: str, frame: pd.DataFrame) -> None:
    value_column = "违反量" if "违反量" in frame.columns else "残差" if "残差" in frame.columns else None
    if value_column is None or "容差" not in frame.columns or "是否满足" not in frame.columns:
        return
    for index, row in frame.iterrows():
        if pd.isna(row[value_column]) or pd.isna(row["容差"]):
            continue
        expected = abs(float(row[value_column])) <= float(row["容差"])
        actual = _as_bool(row["是否满足"])
        if actual is None or actual != expected:
            raise ValueError(f"工作表“{sheet}”第 {index + 2} 行的是否满足与残差/容差不一致")


def _check_quality_gate(prepared: Sequence[tuple[str, pd.DataFrame]]) -> None:
    frame = next((item for name, item in prepared if name == "主结果质量门"), None)
    if frame is None:
        return
    failed = []
    for _, row in frame.iterrows():
        if _as_bool(row["是否通过"]) is not True:
            failed.append(str(row["检查项"]))
    if failed:
        raise ValueError(f"主结果质量门存在未通过项: {failed}")


def _check_missing_value_audit(prepared: Sequence[tuple[str, pd.DataFrame]], kind: str) -> None:
    affected = [name for name, frame in prepared if name != "数据审计" and frame.isna().any().any()]
    if not affected or kind != "solution":
        return
    audit = next((frame for name, frame in prepared if name == "数据审计"), None)
    if audit is None:
        raise ValueError(f"工作表 {affected} 包含缺失值，但缺少数据审计说明")
    text = " ".join(audit.astype(str).fillna("").to_numpy().ravel()).lower()
    if "缺失" not in text and "missing" not in text:
        raise ValueError(f"工作表 {affected} 包含缺失值，但数据审计未说明缺失处理")


def validate_tables(
    tables: Mapping[str, Any],
    workbook_kind: str,
    *,
    schema: Mapping[str, Any],
    problem_types: Sequence[str] = (),
    capabilities: Mapping[str, bool] | None = None,
    objective: str | None = None,
    structures: Sequence[str] = (),
    name_normalizer: Callable[[str], str] | None = None,
    require_quality_passed: bool = True,
) -> list[tuple[str, pd.DataFrame]]:
    kind = _normalize_kind(workbook_kind)
    prepared = prepare_tables(tables, name_normalizer=name_normalizer)
    required_sheets, sheet_schemas = _required_sheet_schemas(schema, kind)
    names = {name for name, _ in prepared}
    missing = sorted(required_sheets - names)
    if missing:
        label = "求解" if kind == "solution" else "结果深化分析"
        raise ValueError(f"{label}工作簿缺少必需工作表: {missing}")
    if kind == "solution":
        required = _conditional_required_sheets(schema, problem_types, capabilities)
        missing_conditional = sorted(required - names)
        if missing_conditional:
            active = sorted(name for name, enabled in (capabilities or {}).items() if enabled)
            note = f"；启用的capabilities: {active}" if active else ""
            raise ValueError(f"求解工作簿缺少必需工作表: {missing_conditional}{note}")
        for profile, required_any in _profile_requirements(schema, objective, structures, problem_types):
            if required_any and not names.intersection(required_any):
                raise ValueError(f"分类剖面“{profile}”至少需要一个专项工作表: {sorted(required_any)}")
    else:
        section = schema["result_analysis_workbook"]
        allowed_any = set(section.get("required_any_sheets", []))
        allowed_all = set(sheet_schemas)
        unknown_sheets = sorted(names - allowed_all)
        if unknown_sheets:
            raise ValueError(f"结果深化分析工作簿包含未登记工作表: {unknown_sheets}")
        if not names.intersection(allowed_any):
            raise ValueError(f"结果深化分析工作簿至少需要一个实质分析表: {sorted(allowed_any)}")
    for name, frame in prepared:
        if name in sheet_schemas:
            _check_required_columns(name, frame, sheet_schemas[name])
        _check_record_keys(name, frame)
        _check_finite_numbers(name, frame)
        _check_residual_consistency(name, frame)
    if kind == "solution" and require_quality_passed:
        _check_quality_gate(prepared)
    _check_missing_value_audit(prepared, kind)
    return prepared


def validate_workbook_file(
    path: Path,
    workbook_kind: str,
    *,
    schema: Mapping[str, Any],
    problem_types: Sequence[str] = (),
    capabilities: Mapping[str, bool] | None = None,
    objective: str | None = None,
    structures: Sequence[str] = (),
    require_quality_passed: bool = True,
) -> list[tuple[str, pd.DataFrame]]:
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)
    return validate_tables(
        read_workbook_tables(path), workbook_kind, schema=schema,
        problem_types=problem_types, capabilities=capabilities,
        objective=objective, structures=structures,
        require_quality_passed=require_quality_passed,
    )


def validate_pair(
    solution_path: Path | None,
    result_analysis_path: Path | None,
    *,
    schema: Mapping[str, Any],
    require_solution: bool,
    require_result_analysis: bool = False,
    require_robustness: bool | None = None,
    problem_types: Sequence[str] = (),
    capabilities: Mapping[str, bool] | None = None,
    objective: str | None = None,
    structures: Sequence[str] = (),
    require_quality_passed: bool = True,
) -> list[str]:
    if require_robustness is not None:
        require_result_analysis = require_robustness
    issues: list[str] = []
    if require_solution and (solution_path is None or not Path(solution_path).is_file()):
        issues.append("缺少标准求解结果工作簿")
    if require_result_analysis and (
        result_analysis_path is None or not Path(result_analysis_path).is_file()
    ):
        issues.append("缺少标准结果深化分析工作簿")
    for path, kind in ((solution_path, "solution"), (result_analysis_path, "result_analysis")):
        if path is None or not Path(path).is_file():
            continue
        try:
            validate_workbook_file(
                Path(path), kind, schema=schema, problem_types=problem_types,
                capabilities=capabilities, objective=objective, structures=structures,
                require_quality_passed=require_quality_passed if kind == "solution" else True,
            )
        except Exception as exc:  # noqa: BLE001
            issues.append(f"{Path(path).name}: {exc}")
    return issues
