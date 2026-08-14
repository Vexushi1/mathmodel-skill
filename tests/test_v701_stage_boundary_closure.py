from __future__ import annotations

import hashlib
import importlib.util
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[1]
FALSE_FLAGS = (
    "allow_reduced_data", "allow_coarser_grid", "allow_shorter_horizon",
    "allow_fewer_repetitions", "allow_relaxed_tolerance", "allow_silent_solver_fallback",
)


def load_receipt():
    spec = importlib.util.spec_from_file_location(
        "validate_user_execution_v701", ROOT / "scripts/validate_user_execution.py"
    )
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def config(stage: str, workbook: str, problem: str = "问题一") -> dict:
    return {
        "execution_owner": "user",
        "execution_profile": "full_fidelity",
        "stage": stage,
        "problem_name": problem,
        "data_paths": ["data.csv"],
        "data_sha256": "a" * 64,
        "solver": "test",
        "solver_version": "1",
        "random_seed": 2026,
        "tolerance": 1e-8,
        "iteration_or_time_limit": "full",
        "expected_workbook": workbook,
        **{flag: False for flag in FALSE_FLAGS},
    }


def write_code(path: Path, cfg: dict, marker: int = 0) -> None:
    path.write_text(
        "FULL_FIDELITY_CONFIG = " + repr(cfg)
        + f"\n\ndef main():\n    return {marker}\n\nif __name__ == '__main__':\n    raise SystemExit(main())\n",
        encoding="utf-8",
    )


def state_payload(primary_hash: str, analysis_hash: str | None = None) -> dict:
    entry = {
        "status": "solved",
        "selected_model": "test",
        "capabilities": {},
        "result_quality_status": "passed",
        "result_analysis_status": "pending",
        "framework_section": "Q1",
        "result_summary_status": "current",
        "code": "问题一求解/问题一求解.py",
        "primary_code_sha256": primary_hash,
        "primary_execution_status": "accepted",
        "analysis_execution_status": "pending",
        "data_hash": "a" * 64,
        "artifacts_stale": False,
        "stale_layers": [],
    }
    if analysis_hash:
        entry["result_analysis_code"] = "问题一求解/问题一结果深化分析.py"
        entry["analysis_code_sha256"] = analysis_hash
    return {
        "project": {"competition": "test", "problem": "A", "current_phase": "result_analysis", "version": "7.2.1"},
        "data": {"active_source_mode": "raw"},
        "preprocessing": {
            "decision": "not_needed", "level": "none", "status": "not_applicable",
            "evidence": ["legacy layout fixture uses raw data directly"], "operations": [],
            "forbidden_operations": [], "downstream_data_source": "raw",
            "quality_status": "not_applicable",
        },
        "subproblems": {"Q1": entry},
    }


def runtime_items(stage: str, problem: str, code_hash: str) -> dict:
    return {
        "execution_owner": "user",
        "execution_profile": "full_fidelity",
        "stage": stage,
        "problem_name": problem,
        "code_sha256": code_hash,
        "data_sha256": "a" * 64,
        "solver": "test",
        "solver_version": "1",
        "tolerance": 1e-8,
        "iteration_or_time_limit": "full",
        "actual_stop_reason": "optimal",
        "random_seed": 2026,
        "repetitions_or_scenarios": 10,
        "grid_or_time_range": "full",
        "fallback_used": False,
        "platform": "test",
        **{flag: False for flag in FALSE_FLAGS},
    }


def write_workbook(path: Path, stage: str, problem: str, code_hash: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    book = openpyxl.Workbook()
    runtime = book.active
    runtime.title = "运行配置"
    runtime.append(["项目", "值"])
    for key, value in runtime_items(stage, problem, code_hash).items():
        runtime.append([key, value])
    if stage == "primary":
        quality = book.create_sheet("主结果质量门")
        quality.append(["检查项", "是否通过", "证据"])
        quality.append(["完整运行", True, "ok"])
    else:
        design = book.create_sheet("分析设计")
        design.append(["风险来源", "分析问题", "方法", "指标", "通过标准"])
        design.append(["参数", "稳定性", "敏感性", "目标值", "保持"])
        summary = book.create_sheet("结论稳定性汇总")
        summary.append(["核心结论", "分析方法", "稳定范围", "是否保持"])
        summary.append(["A", "敏感性", "full", True])
    book.save(path)


class TestV701StageBoundaryClosure(unittest.TestCase):
    def test_unscoped_code_delivery_preserves_accepted_primary_and_delivers_analysis(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            folder = root / "问题一求解"
            folder.mkdir(parents=True)
            primary = folder / "问题一求解.py"
            analysis = folder / "问题一结果深化分析.py"
            write_code(primary, config("primary", "问题一求解结果.xlsx"))
            write_code(analysis, config("analysis", "问题一结果深化分析.xlsx"))
            primary_hash = hashlib.sha256(primary.read_bytes()).hexdigest()
            (root / "state").mkdir()
            state_path = root / "state/project_state.yaml"
            state_path.write_text(
                yaml.safe_dump(state_payload(primary_hash), allow_unicode=True, sort_keys=False),
                encoding="utf-8",
            )

            result = subprocess.run(
                [sys.executable, str(ROOT / "scripts/validate_code_delivery.py"), str(root), "--write", "--strict"],
                text=True,
                capture_output=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
            state = yaml.safe_load(state_path.read_text(encoding="utf-8"))
            entry = state["subproblems"]["Q1"]
            self.assertEqual(entry["primary_execution_status"], "accepted")
            self.assertEqual(entry["primary_code_sha256"], primary_hash)
            self.assertEqual(entry["analysis_execution_status"], "awaiting_user_execution")
            self.assertEqual(
                entry["analysis_code_sha256"], hashlib.sha256(analysis.read_bytes()).hexdigest()
            )

    def test_wrong_stage_in_standard_filename_is_rejected_before_state_write(self):
        receipt = load_receipt()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            folder = root / "问题一求解"
            folder.mkdir()
            primary = folder / "问题一求解.py"
            analysis = folder / "问题一结果深化分析.py"
            write_code(primary, config("primary", "问题一求解结果.xlsx"))
            write_code(analysis, config("analysis", "问题一结果深化分析.xlsx"))
            state = state_payload(
                hashlib.sha256(primary.read_bytes()).hexdigest(),
                hashlib.sha256(analysis.read_bytes()).hexdigest(),
            )
            workbook = folder / "问题一求解结果.xlsx"
            write_workbook(workbook, "analysis", "问题一", state["subproblems"]["Q1"]["analysis_code_sha256"])
            before = yaml.safe_dump(state, allow_unicode=True, sort_keys=False)
            issues = receipt.validate_one(root, workbook, state, True)
            self.assertTrue(any("文件名对应primary阶段不一致" in item for item in issues), issues)
            self.assertEqual(yaml.safe_dump(state, allow_unicode=True, sort_keys=False), before)

    def test_wrong_problem_name_is_rejected_before_state_write(self):
        receipt = load_receipt()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            folder = root / "问题一求解"
            folder.mkdir()
            primary = folder / "问题一求解.py"
            write_code(primary, config("primary", "问题一求解结果.xlsx"))
            primary_hash = hashlib.sha256(primary.read_bytes()).hexdigest()
            state = state_payload(primary_hash)
            workbook = folder / "问题一求解结果.xlsx"
            write_workbook(workbook, "primary", "问题二", primary_hash)
            issues = receipt.validate_one(root, workbook, state, True)
            self.assertTrue(any("problem_name与工作簿目录/文件名不一致" in item for item in issues), issues)
            self.assertNotIn("solution_workbook", state["subproblems"]["Q1"])

    def test_nonstandard_workbook_directory_is_rejected(self):
        receipt = load_receipt()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            misc = root / "misc"
            primary_hash = "b" * 64
            state = state_payload(primary_hash)
            workbook = misc / "问题一求解结果.xlsx"
            write_workbook(workbook, "primary", "问题一", primary_hash)
            issues = receipt.validate_one(root, workbook, state, True)
            self.assertTrue(any("工作簿必须位于" in item for item in issues), issues)

    def test_legacy_result_directory_remains_read_compatible(self):
        receipt = load_receipt()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            primary_hash = "c" * 64
            state = state_payload(primary_hash)
            workbook = root / "结果数据表/问题一/问题一求解结果.xlsx"
            write_workbook(workbook, "primary", "问题一", primary_hash)
            issues = receipt.validate_one(root, workbook, state, True)
            self.assertEqual(issues, [])
            self.assertEqual(state["subproblems"]["Q1"]["primary_execution_status"], "accepted")

    def test_resolver_docstring_uses_current_release(self):
        text = (ROOT / "scripts/resolve_workflow.py").read_text(encoding="utf-8")
        self.assertNotIn("v6.6.0 execution plan", text)
        self.assertNotIn("v7.0.1 execution plan", text)
        self.assertNotIn("v7.1.0 execution plan", text)
        self.assertNotIn("v7.2.2 execution plan", text)
        self.assertIn("v7.3.0 execution plan", text)


if __name__ == "__main__":
    unittest.main()
