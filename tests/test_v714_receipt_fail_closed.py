from __future__ import annotations

import hashlib
import importlib.util
import tempfile
import unittest
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[1]
FALSE_FLAGS = (
    "allow_reduced_data", "allow_coarser_grid", "allow_shorter_horizon",
    "allow_fewer_repetitions", "allow_relaxed_tolerance",
    "allow_silent_solver_fallback",
)


def load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


RECEIPT = load_module("validate_user_execution_v714_fail_closed", ROOT / "scripts" / "validate_user_execution.py")


class V714ReceiptFailClosedTests(unittest.TestCase):
    def make_project(self, root: Path, *, protocol: str | None) -> tuple[Path, dict]:
        (root / "state").mkdir()
        folder = root / "问题一求解"
        folder.mkdir()
        code = folder / "问题一求解.py"
        delivered_config = {"stage": "primary", "problem_name": "问题一"}
        if protocol is not None:
            delivered_config["primary_quality_protocol_version"] = protocol
        code.write_text(
            "FULL_FIDELITY_CONFIG = " + repr(delivered_config)
            + "\n\ndef main():\n    return 0\n\nif __name__ == '__main__':\n    raise SystemExit(main())\n",
            encoding="utf-8",
        )
        code_hash = hashlib.sha256(code.read_bytes()).hexdigest()
        state = {
            "project": {"competition": "test", "problem": "A", "current_phase": "solve_validate"},
            "preprocessing": {"decision": "not_needed", "status": "not_applicable", "quality_status": "not_applicable"},
            "subproblems": {
                "Q1": {
                    "status": "designed",
                    "selected_model": "m",
                    "capabilities": {},
                    "framework_section": "Q1",
                    "result_summary_status": "pending",
                    "result_quality_status": "pending",
                    "result_analysis_status": "pending",
                    "code": "问题一求解/问题一求解.py",
                    "primary_code_sha256": code_hash,
                    "data_hash": "a" * 64,
                }
            },
        }
        return code, state

    def make_workbook(self, root: Path, code: Path, *, protocol: str | None, strict_trace: bool) -> Path:
        workbook = root / "问题一求解" / "问题一求解结果.xlsx"
        book = openpyxl.Workbook()
        runtime = book.active
        runtime.title = "运行配置"
        runtime.append(["项目", "值"])
        items = {
            "execution_owner": "user",
            "execution_profile": "full_fidelity",
            "stage": "primary",
            "problem_name": "问题一",
            "code_sha256": hashlib.sha256(code.read_bytes()).hexdigest(),
            "data_sha256": "a" * 64,
            "solver": "test",
            "solver_version": "1",
            "tolerance": 1e-8,
            "iteration_or_time_limit": "full",
            "actual_stop_reason": "optimal",
            "random_seed": 2026,
            "repetitions_or_scenarios": 1,
            "grid_or_time_range": "full",
            "fallback_used": False,
            "platform": "test",
            **{flag: False for flag in FALSE_FLAGS},
        }
        if protocol is not None:
            items["primary_quality_protocol_version"] = protocol
        for key, value in items.items():
            runtime.append([key, value])

        quality = book.create_sheet("主结果质量门")
        if strict_trace:
            quality.append([
                "Verification ID", "检查项", "是否通过", "证据", "判定关系",
                "阈值或容差", "实际值", "证据工作表", "阈值来源",
            ])
            quality.append([
                "PQ-Q1-01", "完整主计算", True, "底层证据", "bool_true",
                True, True, "基础数值证据", "solver_tolerance",
            ])
            evidence = book.create_sheet("基础数值证据")
            evidence.append(["检查项", "数值"])
            evidence.append(["full_fidelity", 1])
        else:
            quality.append(["检查项", "是否通过", "证据"])
            quality.append(["完整运行", True, "legacy"])
        book.save(workbook)
        return workbook

    def test_v714_delivered_code_cannot_downgrade_by_omitting_workbook_marker(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            code, state = self.make_project(root, protocol="1.0.0")
            workbook = self.make_workbook(root, code, protocol=None, strict_trace=False)
            issues = RECEIPT.validate_one(root, workbook, state, False)
            self.assertTrue(any("不得通过省略标记降级" in item for item in issues))
            self.assertTrue(any("Verification ID" in item for item in issues))

    def test_v714_matching_code_and_workbook_protocol_uses_strict_trace(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            code, state = self.make_project(root, protocol="1.0.0")
            workbook = self.make_workbook(root, code, protocol="1.0.0", strict_trace=True)
            issues = RECEIPT.validate_one(root, workbook, state, False)
            self.assertEqual(issues, [])

    def test_legacy_code_without_protocol_remains_read_compatible(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            code, state = self.make_project(root, protocol=None)
            workbook = self.make_workbook(root, code, protocol=None, strict_trace=False)
            issues = RECEIPT.validate_one(root, workbook, state, False)
            self.assertEqual(issues, [])


if __name__ == "__main__":
    unittest.main()
