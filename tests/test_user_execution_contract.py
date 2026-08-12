from __future__ import annotations

import hashlib
import importlib.util
import tempfile
import unittest
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[1]
FALSE_FLAGS = (
    "allow_reduced_data", "allow_coarser_grid", "allow_shorter_horizon",
    "allow_fewer_repetitions", "allow_relaxed_tolerance",
    "allow_silent_solver_fallback",
)


def load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


CODE = load_module("validate_code_delivery", ROOT / "scripts" / "validate_code_delivery.py")
RECEIPT = load_module("validate_user_execution", ROOT / "scripts" / "validate_user_execution.py")


class UserExecutionContractTests(unittest.TestCase):
    def make_project(self, root: Path) -> Path:
        (root / "state").mkdir()
        folder = root / "问题一求解"
        folder.mkdir()
        code = folder / "问题一求解.py"
        config = self.config("primary", "问题一求解结果.xlsx")
        self.write_code(code, config)
        state = {
            "project": {"competition": "test", "problem": "A", "current_phase": "solve_validate", "version": "7.2.1"},
            "data": {"active_source_mode": "raw"},
            "preprocessing": {
                "decision": "not_needed", "level": "none", "status": "not_applicable",
                "evidence": ["fixture raw data is directly usable"], "operations": [],
                "forbidden_operations": [], "downstream_data_source": "raw",
                "quality_status": "not_applicable",
            },
            "requirements": {"total": 0, "completed": [], "pending": []}, "decisions": {},
            "subproblems": {"Q1": {"status": "designed", "selected_model": "m", "capabilities": {},
                "result_quality_status": "pending", "result_analysis_status": "pending",
                "framework_section": "Q1", "result_summary_status": "pending"}},
            "variables": {"locked": [], "source": {}},
            "paper_framework": {"path": "模型论文框架.md", "version": "1", "sync_status": "stale",
                "last_sync_scope": "design", "proposition_limit": 4, "proposition_count": 0,
                "proposition_status": "not_assessed", "propositions": []},
            "artifacts": {"code": [], "results": [], "figures": [], "papers": []},
            "risks": [], "next_gate": {"module": "solve_validate", "condition": "code"},
        }
        self.write_state(root, state)
        return code

    def config(self, stage: str, workbook: str) -> dict:
        return {
            "execution_owner": "user", "execution_profile": "full_fidelity", "stage": stage,
            "problem_name": "问题一", "data_paths": ["data.csv"], "data_sha256": "a" * 64,
            "solver": "test", "solver_version": "1", "random_seed": 2026, "tolerance": 1e-8,
            "iteration_or_time_limit": "full", "expected_workbook": workbook,
            **{flag: False for flag in FALSE_FLAGS},
        }

    def write_code(self, path: Path, config: dict, marker: int = 0) -> None:
        path.write_text(
            "FULL_FIDELITY_CONFIG = " + repr(config)
            + f"\n\ndef main():\n    return {marker}\n\nif __name__ == \"__main__\":\n    raise SystemExit(main())\n",
            encoding="utf-8",
        )

    def write_state(self, root: Path, state: dict) -> None:
        (root / "state" / "project_state.yaml").write_text(
            yaml.safe_dump(state, allow_unicode=True, sort_keys=False), encoding="utf-8"
        )

    def read_state(self, root: Path) -> dict:
        return yaml.safe_load((root / "state" / "project_state.yaml").read_text(encoding="utf-8"))

    def make_primary_workbook(self, root: Path, code: Path, data_hash: str = "a" * 64) -> Path:
        workbook = root / "问题一求解" / "问题一求解结果.xlsx"
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "运行配置"
        sheet.append(["项目", "值"])
        items = self.runtime_items("primary", code, data_hash)
        for key, value in items.items():
            sheet.append([key, value])
        quality = book.create_sheet("主结果质量门")
        quality.append(["检查项", "是否通过", "证据"])
        quality.append(["完整运行", True, "ok"])
        book.save(workbook)
        return workbook

    def make_analysis_workbook(self, root: Path, code: Path) -> Path:
        workbook = root / "问题一求解" / "问题一结果深化分析.xlsx"
        book = openpyxl.Workbook()
        runtime = book.active
        runtime.title = "运行配置"
        runtime.append(["项目", "值"])
        for key, value in self.runtime_items("analysis", code, "a" * 64).items():
            runtime.append([key, value])
        design = book.create_sheet("分析设计")
        design.append(["风险来源", "分析问题", "方法", "指标", "通过标准"])
        design.append(["参数", "稳定性", "敏感性", "目标值", "结论保持"])
        sensitivity = book.create_sheet("参数敏感性")
        sensitivity.append(["参数", "基准值", "变化值", "结果指标"])
        sensitivity.append(["p", 1.0, 1.1, 2.0])
        summary = book.create_sheet("结论稳定性汇总")
        summary.append(["核心结论", "分析方法", "稳定范围", "是否保持"])
        summary.append(["方案A", "敏感性", "0.9--1.1", True])
        book.save(workbook)
        return workbook

    def runtime_items(self, stage: str, code: Path, data_hash: str) -> dict:
        return {
            "execution_owner": "user", "execution_profile": "full_fidelity", "stage": stage,
            "problem_name": "问题一", "code_sha256": hashlib.sha256(code.read_bytes()).hexdigest(),
            "data_sha256": data_hash, "solver": "test", "solver_version": "1", "tolerance": 1e-8,
            "iteration_or_time_limit": "full", "actual_stop_reason": "optimal", "random_seed": 2026,
            "repetitions_or_scenarios": 100, "grid_or_time_range": "full", "fallback_used": False,
            "platform": "test", **{flag: False for flag in FALSE_FLAGS},
        }

    def accept_primary(self, root: Path, code: Path) -> dict:
        _, config = CODE.validate_script(root, code, "primary")
        CODE.update_state(root, config, code)
        workbook = self.make_primary_workbook(root, code)
        state = self.read_state(root)
        issues = RECEIPT.validate_one(root, workbook, state, True)
        self.assertEqual(issues, [])
        self.write_state(root, state)
        return state

    def make_analysis_code(self, root: Path, marker: int = 0) -> Path:
        code = root / "问题一求解" / "问题一结果深化分析.py"
        self.write_code(code, self.config("analysis", "问题一结果深化分析.xlsx"), marker)
        return code

    def test_code_delivery_does_not_mark_solved(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            code = self.make_project(root)
            issues, config = CODE.validate_script(root, code, "primary")
            self.assertEqual(issues, [])
            CODE.update_state(root, config, code)
            state = self.read_state(root)
            self.assertEqual(state["subproblems"]["Q1"]["status"], "designed")
            self.assertEqual(state["subproblems"]["Q1"]["primary_execution_status"], "awaiting_user_execution")
            self.assertEqual(state["subproblems"]["Q1"]["data_hash"], "a" * 64)

    def test_reduced_flag_is_rejected(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            code = self.make_project(root)
            text = code.read_text(encoding="utf-8").replace("'allow_reduced_data': False", "'allow_reduced_data': True")
            code.write_text(text, encoding="utf-8")
            issues, _ = CODE.validate_script(root, code, "primary")
            self.assertTrue(any("allow_reduced_data" in item for item in issues))

    def test_primary_workbook_acceptance_marks_solved(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            code = self.make_project(root)
            state = self.accept_primary(root, code)
            self.assertEqual(state["subproblems"]["Q1"]["status"], "solved")
            self.assertEqual(state["subproblems"]["Q1"]["result_quality_status"], "passed")

    def test_data_hash_mismatch_is_rejected(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            code = self.make_project(root)
            _, config = CODE.validate_script(root, code, "primary")
            CODE.update_state(root, config, code)
            workbook = self.make_primary_workbook(root, code, data_hash="b" * 64)
            state = self.read_state(root)
            issues = RECEIPT.validate_one(root, workbook, state, True)
            self.assertTrue(any("data_sha256" in item for item in issues))

    def test_analysis_filename_and_stage_must_match(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            self.make_project(root)
            code = root / "问题一求解" / "问题一结果深化分析.py"
            self.write_code(code, self.config("primary", "问题一求解结果.xlsx"))
            issues, _ = CODE.validate_script(root, code)
            self.assertTrue(any("文件名对应analysis阶段" in item for item in issues))

    def test_analysis_requires_accepted_primary(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            self.make_project(root)
            analysis = self.make_analysis_code(root)
            issues, config = CODE.validate_script(root, analysis, "analysis")
            self.assertEqual(issues, [])
            with self.assertRaisesRegex(ValueError, "主工作簿未accepted"):
                CODE.update_state(root, config, analysis)

    def test_analysis_delivery_keeps_primary_path_and_hash_frozen(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            primary = self.make_project(root)
            state = self.accept_primary(root, primary)
            primary_hash = state["subproblems"]["Q1"]["primary_code_sha256"]
            analysis = self.make_analysis_code(root)
            issues, config = CODE.validate_script(root, analysis, "analysis")
            self.assertEqual(issues, [])
            CODE.update_state(root, config, analysis)
            state = self.read_state(root)
            entry = state["subproblems"]["Q1"]
            self.assertEqual(entry["code"], "问题一求解/问题一求解.py")
            self.assertEqual(entry["result_analysis_code"], "问题一求解/问题一结果深化分析.py")
            self.assertEqual(entry["primary_code_sha256"], primary_hash)
            self.assertEqual(entry["analysis_code_sha256"], hashlib.sha256(analysis.read_bytes()).hexdigest())

    def test_analysis_change_does_not_invalidate_primary_quality(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            primary = self.make_project(root)
            self.accept_primary(root, primary)
            analysis = self.make_analysis_code(root)
            _, config = CODE.validate_script(root, analysis, "analysis")
            CODE.update_state(root, config, analysis)
            state = self.read_state(root)
            state["subproblems"]["Q1"]["result_quality_status"] = "passed"
            self.write_state(root, state)
            self.write_code(analysis, self.config("analysis", "问题一结果深化分析.xlsx"), marker=1)
            _, config = CODE.validate_script(root, analysis, "analysis")
            CODE.update_state(root, config, analysis)
            entry = self.read_state(root)["subproblems"]["Q1"]
            self.assertEqual(entry["result_quality_status"], "passed")
            self.assertEqual(entry["result_analysis_status"], "pending")
            self.assertIn("result_analysis_workbook", entry["stale_layers"])
            self.assertNotIn("solution_workbook", entry["stale_layers"])

    def test_accepted_primary_is_frozen_outside_solve_validate(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            primary = self.make_project(root)
            self.accept_primary(root, primary)
            state = self.read_state(root)
            state["project"]["current_phase"] = "result_analysis"
            self.write_state(root, state)
            self.write_code(primary, self.config("primary", "问题一求解结果.xlsx"), marker=2)
            issues, config = CODE.validate_script(root, primary, "primary")
            self.assertEqual(issues, [])
            with self.assertRaisesRegex(ValueError, "已accepted并冻结"):
                CODE.update_state(root, config, primary)


if __name__ == "__main__":
    unittest.main()
