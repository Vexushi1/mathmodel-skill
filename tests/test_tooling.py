import importlib.util
import json
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

import yaml
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]


def load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    parent = str(path.parent)
    added = parent not in sys.path
    if added:
        sys.path.insert(0, parent)
    try:
        sys.modules[name] = module
        spec.loader.exec_module(module)
    finally:
        if added:
            sys.path.remove(parent)
    return module


class TestTooling(unittest.TestCase):
    def test_compile_profiles_are_complete_and_entry_driven(self):
        payload = yaml.safe_load((ROOT / "core/compile_profiles.yaml").read_text(encoding="utf-8"))
        for name in ("cumcm", "mcm_icm", "diangong"):
            profile = payload["profiles"][name]
            self.assertTrue(profile["sequence"])
            self.assertIn(profile["engine"], {"xelatex", "pdflatex", "lualatex"})
            self.assertTrue(profile["project_main"])
            template = ROOT / profile["template_directory"] / profile["template_main"]
            self.assertTrue(template.is_file(), template)
        self.assertIn("biber", payload["profiles"]["cumcm"]["sequence"])

    def test_competition_profiles_separate_stable_and_edition_rules(self):
        payload = yaml.safe_load((ROOT / "config/competition_profiles.yaml").read_text(encoding="utf-8"))
        required = set(payload["edition_rule_contract"]["required_fields"])
        for profile in payload["profiles"].values():
            self.assertIn("stable", profile)
            self.assertTrue(required.issubset(profile["edition_rules"]))
            self.assertIn(
                profile["edition_rules"]["verification_status"],
                payload["edition_rule_contract"]["verification_status"],
            )

    def test_packager_excludes_multisuffix_latex_files(self):
        module = load_module("hsk_pack_submission", ROOT / "scripts/hsk_pack_submission.py")
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            output = root / "package.zip"
            for name in ("main.synctex.gz", "main.run.xml", "main.bcf", "main.pdf"):
                (root / name).write_text("x", encoding="utf-8")
            self.assertTrue(module.should_exclude(root / "main.synctex.gz", root, output))
            self.assertTrue(module.should_exclude(root / "main.run.xml", root, output))
            self.assertTrue(module.should_exclude(root / "main.bcf", root, output))
            self.assertFalse(module.should_exclude(root / "main.pdf", root, output))

    def _append(self, book: Workbook, title: str, headers, row) -> None:
        if len(book.sheetnames) == 1 and book.active["A1"].value is None:
            sheet = book.active
            sheet.title = title
            sheet.delete_rows(1, 1)
        else:
            sheet = book.create_sheet(title)
        sheet.append(list(headers))
        sheet.append(list(row))

    def _write_standard_workbooks(self, solution: Path, analysis: Path) -> None:
        workbook = Workbook()
        self._append(workbook, "核心指标", ["指标", "数值"], ["目标值", 1.0])
        self._append(workbook, "数据审计", ["等级", "检查项", "信息", "处理方式"], ["Info", "字段", "通过", "无"])
        self._append(workbook, "主结果质量门", ["检查项", "是否通过", "证据"], ["收敛", True, "终止条件"])
        workbook.save(solution)

        workbook = Workbook()
        self._append(
            workbook,
            "分析设计",
            ["风险来源", "分析问题", "方法", "指标", "通过标准"],
            ["算法偶然性", "结论是否一致", "多算法", "目标差异", "小于1%"],
        )
        self._append(workbook, "算法一致性", ["算法", "重复编号", "目标值", "是否可行"], ["A", 1, 1.0, True])
        self._append(
            workbook,
            "结论稳定性汇总",
            ["核心结论", "分析方法", "稳定范围", "是否保持"],
            ["主结论", "多算法", "三种算法", True],
        )
        workbook.save(analysis)

    def test_artifact_checker_accepts_split_standard_workbooks(self):
        module = load_module("hsk_check_artifact", ROOT / "scripts/hsk_check_artifact.py")
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            solution, analysis = root / "solution.xlsx", root / "analysis.xlsx"
            self._write_standard_workbooks(solution, analysis)
            self.assertEqual(module.inspect_workbook(solution, "solution"), [])
            self.assertEqual(module.inspect_workbook(analysis, "result_analysis"), [])

    def test_artifact_checker_reuses_required_column_validation(self):
        module = load_module("hsk_check_artifact_columns", ROOT / "scripts/hsk_check_artifact.py")
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "solution.xlsx"
            workbook = Workbook()
            self._append(workbook, "核心指标", ["结果"], [1.0])
            self._append(workbook, "数据审计", ["等级", "检查项", "信息", "处理方式"], ["Info", "字段", "通过", "无"])
            self._append(workbook, "主结果质量门", ["检查项", "是否通过", "证据"], ["收敛", True, "通过"])
            workbook.save(path)
            issues = module.inspect_workbook(path, "solution")
            self.assertTrue(any("缺少必需字段" in issue for issue in issues), issues)

    def test_artifact_checker_enforces_conditional_constraint_sheet(self):
        module = load_module("hsk_check_artifact_constraints", ROOT / "scripts/hsk_check_artifact.py")
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "solution.xlsx"
            analysis = Path(temp) / "analysis.xlsx"
            self._write_standard_workbooks(path, analysis)
            issues = module.inspect_workbook(path, "solution", problem_types=("optimization",))
            self.assertTrue(any("约束违反检查" in issue for issue in issues), issues)

    def test_framework_validator_accepts_template_and_links_solved_summary(self):
        module = load_module("validate_model_paper_framework", ROOT / "scripts/validate_model_paper_framework.py")
        template = ROOT / "templates/model/model_paper_framework.md"
        self.assertEqual(module.validate_framework_file(template), [])
        text = template.read_text(encoding="utf-8")
        state = {
            "paper_framework": {"sync_status": "current"},
            "subproblems": {
                "Q1": {
                    "status": "solved",
                    "framework_section": "### Q1：__QUESTION_NAME__",
                    "result_summary_status": "current",
                    "result_summary_anchor": "#### 结果摘要",
                    "artifacts_stale": False,
                }
            },
        }
        self.assertEqual(module.validate_framework_text(text, state=state), [])

    def test_framework_validator_rejects_stale_solved_summary(self):
        module = load_module("validate_model_paper_framework_stale", ROOT / "scripts/validate_model_paper_framework.py")
        text = (ROOT / "templates/model/model_paper_framework.md").read_text(encoding="utf-8")
        state = {
            "paper_framework": {"sync_status": "stale"},
            "subproblems": {
                "Q1": {
                    "status": "solved",
                    "framework_section": "### Q1：__QUESTION_NAME__",
                    "result_summary_status": "stale",
                    "result_summary_anchor": "",
                    "artifacts_stale": True,
                }
            },
        }
        issues = module.validate_framework_text(text, state=state)
        self.assertTrue(any("sync_status" in issue for issue in issues), issues)
        self.assertTrue(any("result_summary_status" in issue for issue in issues), issues)

    def test_manifest_digest_normalizes_text_line_endings(self):
        module = load_module("generate_indexes", ROOT / "scripts/generate_indexes.py")
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "sample.txt"
            path.write_bytes(b"alpha\r\nbeta\r\n")
            crlf_digest = module.digest_file(path)
            path.write_bytes(b"alpha\nbeta\n")
            self.assertEqual(crlf_digest, module.digest_file(path))
        active = {path.as_posix() for path in module.iter_files()}
        self.assertIn("legacy/README.md", active)
        self.assertFalse(any(path.startswith("legacy/") and path != "legacy/README.md" for path in active))

    def test_render_paper_refuses_ambiguous_profile_and_uses_profile_main(self):
        module = load_module("render_paper", ROOT / "scripts/render_paper.py")
        profiles = module.load_profiles()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            main = root / "main.tex"
            main.write_text("\\documentclass{article}\n\\begin{document}\ntest\n\\end{document}\n", encoding="utf-8")
            with self.assertRaisesRegex(SystemExit, "cannot safely infer"):
                module.infer_profile(main, profiles)
            hsk = root / "hsk_main.tex"
            hsk.write_text(main.read_text(encoding="utf-8"), encoding="utf-8")
            self.assertEqual(module.resolve_main(root, None, profiles["cumcm"]), main)

    def test_review_scorer_uses_active_weights(self):
        module = load_module("score_submission", ROOT / "scripts/score_submission.py")
        config = json.loads((ROOT / "config/review_weights.json").read_text(encoding="utf-8"))
        report = {"scores": {name: 80 for name in config["dimensions"]}, "hard_fail": []}
        result = module.score_submission(config, report)
        self.assertEqual(result["total"], 80.0)
        report["hard_fail"] = ["latex_compile_failure"]
        self.assertEqual(module.score_submission(config, report)["status"], "reject_or_major_rework")

    def test_cumcm_font_patch_is_narrow_and_idempotent(self):
        module = load_module("prepare_cumcm_class", ROOT / "scripts/prepare_cumcm_class.py")
        source = ROOT / "templates/latex/cumcm/cumcmthesis/cumcmthesis.cls"
        original = source.read_text(encoding="utf-8")
        self.assertEqual(original.count(module.ORIGINAL_FONT_BLOCK), 1)
        suffix = original.split(module.ORIGINAL_FONT_BLOCK, 1)[1]
        with tempfile.TemporaryDirectory() as temp:
            target = Path(temp) / "cumcmthesis.cls"
            shutil.copyfile(source, target)
            self.assertTrue(module.patch_cumcm_class(target))
            patched = target.read_text(encoding="utf-8")
            self.assertIn(module.FALLBACK_FONT_BLOCK, patched)
            self.assertTrue(patched.endswith(suffix))
            self.assertFalse(module.patch_cumcm_class(target))

    def test_matlab_templates_use_real_headers_fixed_columns_and_titles(self):
        plotting = (ROOT / "templates/matlab/q1_plot.m").read_text(encoding="utf-8")
        style = (ROOT / "templates/matlab/hsk_apply_scientific_style.m").read_text(encoding="utf-8")
        reader = (ROOT / "templates/matlab/hsk_read_result_workbooks.m").read_text(encoding="utf-8")
        self.assertNotIn("hsk_find_project_root", plotting)
        self.assertIn('fullfile(resultDir, "问题一求解结果.xlsx")', plotting)
        self.assertIn('fullfile(resultDir, "图表")', plotting)
        self.assertIn("信息效率", plotting)
        self.assertIn("readcell", plotting)
        self.assertIn("xColumn = NaN", plotting)
        self.assertIn("actualXHeader == xHeader", plotting)
        self.assertNotIn("readtable(", plotting)
        self.assertIn('figureTitle = "__ACTUAL_FIGURE_TITLE__"', plotting)
        self.assertIn("title(ax, figureTitle", plotting)
        self.assertIn("FontWeight", plotting)
        self.assertIn("listfonts", style)
        self.assertIn("Noto Sans CJK SC", style)
        self.assertIn("ax.Title", style)
        self.assertIn("结果深化分析.xlsx", reader)
        self.assertIn("books.analysis", reader)
        self.assertIn("readcell", reader)
        self.assertIn("fixedColumns", reader)
        self.assertIn("expectedHeaders", reader)
        self.assertNotIn("missingColumns", reader)

    def test_matlab_handoff_requires_title_caption_and_framework_registry(self):
        module = load_module("matlab_handoff", ROOT / "templates/code/hsk_pipeline/matlab_handoff.py")
        for field in ("matlab_title", "paper_caption", "framework_registry"):
            self.assertIn(field, module.REQUIRED_FIELDS)


if __name__ == "__main__":
    unittest.main()
