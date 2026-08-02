import importlib.util
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]


def load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


SYNC = load_module("sync_v632", ROOT / "scripts/sync_project.py")
VALIDATOR = load_module("validator_v632", ROOT / "templates/code/hsk_pipeline/workbook_validation.py")
STATE_VALIDATOR = load_module("state_validator_v632", ROOT / "scripts/validate_project_state.py")
SCHEMA = yaml.safe_load((ROOT / "core/workbook_schema.yaml").read_text(encoding="utf-8"))


def append_sheet(book: Workbook, title: str, headers, row):
    if len(book.sheetnames) == 1 and book.active["A1"].value is None:
        sheet = book.active
        sheet.title = title
        sheet.delete_rows(1, 1)
    else:
        sheet = book.create_sheet(title)
    sheet.append(headers)
    sheet.append(row)


def solution_book(path: Path, specialized="推荐方案"):
    book = Workbook()
    append_sheet(book, "运行配置", ["项目", "值"], ["stage", "primary"])
    append_sheet(book, "核心指标", ["指标", "数值"], ["目标值", 1.0])
    append_sheet(book, "数据审计", ["等级", "检查项", "信息", "处理方式"], ["Info", "字段", "通过", "无"])
    append_sheet(book, "主结果质量门", ["检查项", "是否通过", "证据"], ["收敛", True, "通过"])
    if specialized == "推荐方案":
        append_sheet(book, "推荐方案", ["方案"], ["A"])
    elif specialized == "机理分析":
        append_sheet(book, "机理分析", ["对象或机制", "关系或结论"], ["对象", "结论"])
    elif specialized == "仿真明细":
        append_sheet(book, "仿真明细", ["记录键", "场景", "时刻", "数值"], ["r1", "s1", 0, 1.0])
    book.save(path)


def analysis_book(path: Path):
    book = Workbook()
    append_sheet(book, "运行配置", ["项目", "值"], ["stage", "analysis"])
    append_sheet(
        book,
        "分析设计",
        ["风险来源", "分析问题", "方法", "指标", "通过标准"],
        ["算法", "一致性", "多算法", "差异", "小于1%"],
    )
    append_sheet(book, "算法一致性", ["算法", "重复编号", "目标值", "是否可行"], ["A", 1, 1.0, True])
    append_sheet(
        book,
        "结论稳定性汇总",
        ["核心结论", "分析方法", "稳定范围", "是否保持"],
        ["结论", "多算法", "三种算法", True],
    )
    book.save(path)


def caps():
    return {name: False for name in SCHEMA["capability_contract"]["allowed"]}


def project_state(root: Path, *, status="designed", phase="model_design", objective="optimization"):
    (root / "state").mkdir(parents=True, exist_ok=True)
    solved = status in {"solved", "analyzed", "validated", "written", "completed"}
    analyzed = status in {"analyzed", "validated", "written", "completed"}
    payload = {
        "project": {"competition": "测试", "problem": "A", "current_phase": phase},
        "requirements": {"total": 0, "completed": [], "pending": []},
        "decisions": {},
        "subproblems": {
            "Q1": {
                "status": status,
                "selected_model": "测试模型",
                "classification": {"objective": objective, "structures": []},
                "capabilities": caps(),
                "result_quality_status": "passed" if solved else "pending",
                "result_analysis_status": "passed" if analyzed else "pending",
                "validation_status": "passed" if status in {"validated", "written", "completed"} else "pending",
                "framework_section": "### Q1",
                "result_summary_status": "current" if solved else "pending",
                "result_summary_anchor": "### Q1" if solved else "",
                "artifacts_stale": False,
                "stale_layers": [],
                "analysis_methods": ["算法一致性"] if analyzed else [],
                "evidence": ["evidence"] if status in {"validated", "written", "completed"} else [],
                "artifact_hashes": {},
                "validated_artifact_hashes": {},
                "optimality_claim": "none",
            }
        },
        "variables": {"locked": [], "source": {}},
        "paper_framework": {
            "path": "模型论文框架.md", "version": "1", "mode": "compact",
            "sync_status": "current", "last_sync_scope": "design",
            "proposition_limit": 4, "proposition_count": 0,
            "proposition_status": "not_assessed", "propositions": [],
        },
        "artifacts": {"code": [], "results": [], "figures": [], "papers": []},
        "risks": [],
        "next_gate": {"module": "solve_validate", "condition": "test"},
    }
    path = root / "state/project_state.yaml"
    path.write_text(yaml.safe_dump(payload, allow_unicode=True, sort_keys=False), encoding="utf-8")
    return path


def framework_text() -> str:
    return (
        "# 模型论文框架\n只保留当前有效版本\n"
        "## 当前有效口径\n## 各问模型与结果\n### Q1\n"
        "## 图表证据链\n## 待办与缺口\n"
    )


class TestV632DeliveryGateClosure(unittest.TestCase):
    def base_tables(self):
        return {
            "运行配置": pd.DataFrame({"项目": ["stage"], "值": ["primary"]}),
            "核心指标": pd.DataFrame({"指标": ["x"], "数值": [1.0]}),
            "数据审计": pd.DataFrame({"等级": ["Info"], "检查项": ["字段"], "信息": ["通过"], "处理方式": ["无"]}),
            "主结果质量门": pd.DataFrame({"检查项": ["收敛"], "是否通过": [True], "证据": ["通过"]}),
        }

    def test_objective_profiles_require_specialized_evidence(self):
        for objective in ("explanation", "optimization", "simulation"):
            with self.subTest(objective=objective):
                with self.assertRaisesRegex(ValueError, objective):
                    VALIDATOR.validate_tables(
                        self.base_tables(), "solution", schema=SCHEMA,
                        objective=objective, capabilities=caps(),
                    )

    def test_design_scope_requires_framework(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            project_state(root)
            report = SYNC.synchronize(root, write=False, delivery_scope="design")
            self.assertTrue(any("模型论文框架" in issue for issue in report["issues"]))

    def test_latex_scope_requires_concrete_artifacts_without_results(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            project_state(root)
            (root / "模型论文框架.md").write_text(framework_text(), encoding="utf-8")
            report = SYNC.synchronize(root, write=False, delivery_scope="latex")
            joined = "\n".join(report["issues"])
            self.assertNotIn("求解结果工作簿", joined)
            self.assertNotIn("结果深化分析工作簿", joined)
            self.assertIn("final_latex/main.tex", joined)
            self.assertIn("final_latex/main.pdf", joined)
            self.assertIn("compile_report", joined)

    def test_submission_zip_content_is_checked(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            package = root / "submission" / "submission.zip"
            package.parent.mkdir(parents=True)
            with zipfile.ZipFile(package, "w") as archive:
                archive.writestr("main.pdf", b"pdf")
            issues = SYNC._submission_zip_issues(package, require_matlab=True)
            self.assertTrue(any("Python" in issue for issue in issues))
            self.assertTrue(any("结果工作簿" in issue for issue in issues))
            self.assertTrue(any("MATLAB" in issue for issue in issues))

    def test_data_hash_uses_declared_sources(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            (root / "data.csv").write_text("x\n1\n", encoding="utf-8")
            (root / "paper.pdf").write_bytes(b"old")
            state = {"data": {"sources": [{"name": "data", "path": "data.csv", "role": "input"}]}}
            files, mode, issues, _ = SYNC.data_source_files(root, state)
            first = SYNC.combined_hash(files, root)
            (root / "paper.pdf").write_bytes(b"new")
            second = SYNC.combined_hash(files, root)
            self.assertEqual(mode, "declared_sources")
            self.assertFalse(issues)
            self.assertEqual(first, second)

    def test_empty_matlab_reference_is_rejected(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            result = root / "结果数据表" / "问题一"
            (result / "图表").mkdir(parents=True)
            solution_book(result / "问题一求解结果.xlsx")
            analysis_book(result / "问题一结果深化分析.xlsx")
            (result / "q1_plot.m").write_text('title(gca, "结果");', encoding="utf-8")
            (result / "图表/a.png").write_bytes(b"x")
            project_state(root, status="analyzed", phase="figure_evidence")
            (root / "模型论文框架.md").write_text(framework_text(), encoding="utf-8")
            report = SYNC.synchronize(root, write=False, delivery_scope="figures")
            self.assertTrue(any("未发现标准工作簿引用" in issue for issue in report["issues"]))

    def test_framework_hash_is_per_question_section(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "模型论文框架.md"
            path.write_text("# 头部\n### Q1\nA\n### Q2\nB\n", encoding="utf-8")
            first = SYNC.framework_section_hash(path, "### Q1")
            path.write_text("# 改过头部\n### Q1\nA\n### Q2\nC\n", encoding="utf-8")
            second = SYNC.framework_section_hash(path, "### Q1")
            self.assertEqual(first, second)

    def test_hash_mismatch_forces_stale_even_when_solved(self):
        state = {
            "artifact_hashes": {"model": "a" * 64},
            "validated_artifact_hashes": {"model": "b" * 64},
            "artifacts_stale": False,
            "stale_layers": [],
        }
        issues = STATE_VALIDATOR._validate_hashes("Q1", state, "solved")
        self.assertTrue(any("artifacts_stale must be true" in issue for issue in issues))


if __name__ == "__main__":
    unittest.main()
