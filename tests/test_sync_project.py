import hashlib
import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path

import yaml
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]


def load_syncer():
    spec = importlib.util.spec_from_file_location(
        "sync_project", ROOT / "scripts/sync_project.py"
    )
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def append_sheet(book: Workbook, title: str, headers, row):
    if (
        len(book.sheetnames) == 1
        and book.active.max_row == 1
        and book.active["A1"].value is None
    ):
        ws = book.active
        ws.title = title
        ws.delete_rows(1, 1)
    else:
        ws = book.create_sheet(title)
    ws.append(list(headers))
    ws.append(list(row))


def write_solution(
    path: Path, *, constraint=False, out_of_sample=False, objective="optimization"
):
    book = Workbook()
    append_sheet(book, "核心指标", ["指标", "数值"], ["目标值", 1.0])
    append_sheet(
        book,
        "数据审计",
        ["等级", "检查项", "信息", "处理方式"],
        ["Info", "完整性", "通过", "无需处理"],
    )
    append_sheet(
        book,
        "主结果质量门",
        ["检查项", "是否通过", "证据"],
        ["收敛", True, "达到终止条件"],
    )
    if objective == "optimization":
        append_sheet(book, "推荐方案", ["方案"], ["A"])
    if constraint:
        append_sheet(
            book,
            "约束违反检查",
            ["约束编号", "约束含义", "违反量", "容差", "是否满足"],
            ["C1", "容量", 0.0, 1e-8, "是"],
        )
    if out_of_sample:
        append_sheet(
            book,
            "外样本验证",
            ["划分或窗口", "指标", "数值"],
            ["测试集", "MAE", 0.1],
        )
    book.save(path)


def write_analysis(path: Path):
    book = Workbook()
    append_sheet(book, "运行配置", ["项目", "值"], ["stage", "analysis"])
    append_sheet(
        book,
        "分析设计",
        ["风险来源", "分析问题", "方法", "指标", "通过标准"],
        ["局部最优", "算法是否一致", "多算法", "目标差异", "小于1%"],
    )
    append_sheet(
        book,
        "算法一致性",
        ["算法", "重复编号", "目标值", "是否可行"],
        ["A", 1, 1.0, True],
    )
    append_sheet(
        book,
        "结论稳定性汇总",
        ["核心结论", "分析方法", "稳定范围", "是否保持"],
        ["方案A最优", "多算法", "三种算法", True],
    )
    book.save(path)


def capabilities(**overrides):
    values = {
        "has_explicit_constraints": False,
        "requires_feasibility_check": False,
        "requires_equilibrium_residual": False,
        "requires_conservation_residual": False,
        "requires_discretization_check": False,
        "requires_convergence_diagnostic": False,
        "requires_out_of_sample_validation": False,
        "requires_uncertainty_quantification": False,
        "requires_leakage_check": False,
        "requires_calibration_check": False,
        "requires_identifiability_check": False,
    }
    values.update(overrides)
    return values


def framework_text() -> str:
    return (
        "# 模型论文框架\n\n"
        "> 本文件只保留当前有效版本。\n\n"
        "- 最近同步：`design`\n"
        "- 最近同步时间：`x`\n"
        "- 当前状态：`current`\n\n"
        "## 当前有效口径\n\n"
        "## 各问模型与结果\n\n"
        "### Q1\n\n"
        "## 图表证据链\n\n"
        "## 待办与缺口\n"
    )


def write_state(
    root: Path, status="designed", caps=None, phase="model_design", validated=None
):
    state_dir = root / "state"
    state_dir.mkdir(exist_ok=True)
    solved = status in {"solved", "analyzed", "validated", "written", "completed"}
    analyzed = status in {"analyzed", "validated", "written", "completed"}
    payload = {
        "project": {
            "competition": "测试",
            "problem": "A",
            "current_phase": phase,
        },
        "requirements": {"total": 0, "completed": [], "pending": []},
        "decisions": {},
        "subproblems": {
            "Q1": {
                "status": status,
                "selected_model": "测试模型",
                "classification": {"objective": "optimization", "structures": []},
                "capabilities": caps or capabilities(),
                "result_quality_status": "passed" if solved else "pending",
                "result_analysis_status": "passed" if analyzed else "pending",
                "validation_status": (
                    "passed"
                    if status in {"validated", "written", "completed"}
                    else "pending"
                ),
                "result_summary_status": "current" if solved else "pending",
                "result_summary_anchor": "### Q1" if solved else "",
                "framework_section": "### Q1",
                "artifacts_stale": False,
                "stale_layers": [],
                "evidence": (
                    ["evidence"]
                    if status in {"validated", "written", "completed"}
                    else []
                ),
                "analysis_methods": ["算法一致性"] if analyzed else [],
                "artifact_hashes": {},
                "validated_artifact_hashes": validated or {},
                "optimality_claim": "none",
            }
        },
        "variables": {"locked": [], "source": {}},
        "artifacts": {"code": [], "results": [], "figures": [], "papers": []},
        "paper_framework": {
            "path": "模型论文框架.md",
            "version": "1",
            "mode": "compact",
            "sync_status": "current",
            "last_sync_scope": "design",
            "proposition_limit": 4,
            "proposition_count": 0,
            "proposition_status": "not_assessed",
            "propositions": [],
        },
        "risks": [],
        "next_gate": {"module": "solve_validate", "condition": "test"},
    }
    (state_dir / "project_state.yaml").write_text(
        yaml.safe_dump(payload, allow_unicode=True, sort_keys=False), encoding="utf-8"
    )


def setup_project(
    root: Path,
    *,
    status="designed",
    caps=None,
    phase="model_design",
    include_solution=True,
    include_analysis=True,
):
    (root / "问题一求解.py").write_text(
        "def main():\n    return 1\n\nif __name__ == '__main__':\n    main()\n",
        encoding="utf-8",
    )
    result = root / "结果数据表" / "问题一"
    result.mkdir(parents=True)
    if include_solution:
        write_solution(
            result / "问题一求解结果.xlsx",
            constraint=bool((caps or {}).get("has_explicit_constraints")),
            out_of_sample=bool(
                (caps or {}).get("requires_out_of_sample_validation")
            ),
        )
    if include_analysis:
        write_analysis(result / "问题一结果深化分析.xlsx")
    (root / "模型论文框架.md").write_text(framework_text(), encoding="utf-8")
    write_state(root, status=status, caps=caps, phase=phase)
    return result


def write_code_delivery(root: Path):
    code = root / "问题一求解.py"
    config = {
        "execution_owner": "user",
        "execution_profile": "full_fidelity",
        "stage": "primary",
        "problem_name": "问题一",
        "code_path": code.name,
        "code_sha256": hashlib.sha256(code.read_bytes()).hexdigest(),
        "data_paths": ["data.csv"],
        "data_sha256": "a" * 64,
        "solver": "test",
        "solver_version": "1",
        "random_seed": 2026,
        "tolerance": 1e-8,
        "iteration_or_time_limit": "full",
        "expected_workbook": "结果数据表/问题一/问题一求解结果.xlsx",
        "allow_reduced_data": False,
        "allow_coarser_grid": False,
        "allow_shorter_horizon": False,
        "allow_fewer_repetitions": False,
        "allow_relaxed_tolerance": False,
        "allow_silent_solver_fallback": False,
    }
    (root / "问题一完整运行配置.yaml").write_text(
        yaml.safe_dump(config, allow_unicode=True, sort_keys=False), encoding="utf-8"
    )
    (root / "问题一本地运行说明.md").write_text("完整运行", encoding="utf-8")
    (root / "code_delivery_report.yaml").write_text(
        yaml.safe_dump(
            {
                "status": "passed",
                "checked_configs": ["问题一完整运行配置.yaml"],
                "issues": [],
                "task_code_executed": False,
            },
            allow_unicode=True,
            sort_keys=False,
        ),
        encoding="utf-8",
    )


class TestSyncProject(unittest.TestCase):
    def test_discovers_artifacts_writes_report_and_final_framework_hash(self):
        syncer = load_syncer()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            result = setup_project(root)
            (result / "q1_plot.m").write_text(
                'title(gca, "结果");', encoding="utf-8"
            )
            report = syncer.synchronize(root, write=True, delivery_scope="design")
            self.assertIn("Q1", report["questions"])
            self.assertTrue((root / "sync_report.yaml").is_file())
            updated = yaml.safe_load(
                (root / "state/project_state.yaml").read_text(encoding="utf-8")
            )
            actual_hash = syncer.sha256_file(root / "模型论文框架.md")
            self.assertEqual(updated["paper_framework"]["sha256"], actual_hash)
            self.assertEqual(report["framework_hash"], actual_hash)

    def test_designed_project_does_not_require_workbooks(self):
        syncer = load_syncer()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            (root / "模型论文框架.md").write_text(
                framework_text(), encoding="utf-8"
            )
            write_state(root, status="designed", phase="model_design")
            report = syncer.synchronize(root, write=False)
            self.assertFalse(any("缺少标准" in issue for issue in report["issues"]))

    def test_formal_code_scope_passes_without_result_workbooks(self):
        syncer = load_syncer()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            setup_project(
                root,
                status="designed",
                phase="solve_validate",
                include_solution=False,
                include_analysis=False,
            )
            write_code_delivery(root)
            report = syncer.synchronize(root, write=False, delivery_scope="code")
            self.assertEqual(report["delivery_scope"], "code")
            self.assertFalse(any("工作簿" in issue for issue in report["issues"]))
            self.assertFalse(report["issues"])

    def test_formal_code_scope_requires_config_instructions_and_report(self):
        syncer = load_syncer()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            setup_project(
                root,
                status="designed",
                phase="solve_validate",
                include_solution=False,
                include_analysis=False,
            )
            report = syncer.synchronize(root, write=False, delivery_scope="code")
            self.assertTrue(any("完整运行配置" in issue for issue in report["issues"]))
            self.assertTrue(any("本地运行说明" in issue for issue in report["issues"]))
            self.assertTrue(any("code_delivery_report" in issue for issue in report["issues"]))

    def test_solve_phase_defaults_to_code_scope(self):
        syncer = load_syncer()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            setup_project(
                root,
                status="designed",
                phase="solve_validate",
                include_solution=False,
                include_analysis=False,
            )
            report = syncer.synchronize(root, write=False)
            self.assertEqual(report["delivery_scope"], "code")

    def test_solved_project_requires_only_primary_workbook_before_formal_results_delivery(self):
        syncer = load_syncer()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            setup_project(
                root,
                status="solved",
                phase="solve_validate",
                include_analysis=False,
            )
            report = syncer.synchronize(root, write=False)
            self.assertFalse(
                any("结果深化分析工作簿" in issue for issue in report["issues"])
            )

    def test_analyzed_project_requires_result_analysis_workbook(self):
        syncer = load_syncer()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            setup_project(
                root,
                status="analyzed",
                phase="result_analysis",
                include_analysis=False,
            )
            report = syncer.synchronize(root, write=False)
            self.assertTrue(
                any(
                    "缺少标准结果深化分析工作簿" in issue
                    for issue in report["issues"]
                )
            )

    def test_formal_results_scope_requires_both_stages(self):
        syncer = load_syncer()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            setup_project(
                root,
                status="solved",
                phase="solve_validate",
                include_analysis=False,
            )
            report = syncer.synchronize(
                root, write=False, delivery_scope="results"
            )
            self.assertTrue(
                any("结果深化分析" in issue for issue in report["issues"])
            )

    def test_capability_required_sheet_is_enforced(self):
        syncer = load_syncer()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            caps = capabilities(requires_out_of_sample_validation=True)
            result = setup_project(
                root,
                status="solved",
                caps=caps,
                phase="solve_validate",
                include_analysis=False,
            )
            write_solution(result / "问题一求解结果.xlsx", out_of_sample=False)
            report = syncer.synchronize(root, write=False)
            self.assertTrue(report["issues"], report)

    def test_solution_hash_change_invalidates_quality_and_analysis(self):
        syncer = load_syncer()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            result = setup_project(
                root, status="analyzed", phase="result_analysis"
            )
            initial = syncer.synchronize(root, write=False)
            current = initial["questions"]["Q1"]["artifact_hashes"]
            state_path = root / "state/project_state.yaml"
            state = yaml.safe_load(state_path.read_text(encoding="utf-8"))
            state["subproblems"]["Q1"]["artifact_hashes"] = current
            state["subproblems"]["Q1"]["validated_artifact_hashes"] = current
            state_path.write_text(
                yaml.safe_dump(state, allow_unicode=True, sort_keys=False),
                encoding="utf-8",
            )
            book = Workbook()
            append_sheet(book, "核心指标", ["指标", "数值"], ["目标值", 2.0])
            append_sheet(
                book,
                "数据审计",
                ["等级", "检查项", "信息", "处理方式"],
                ["Info", "完整性", "通过", "无需处理"],
            )
            append_sheet(
                book,
                "主结果质量门",
                ["检查项", "是否通过", "证据"],
                ["收敛", True, "通过"],
            )
            append_sheet(book, "推荐方案", ["方案"], ["B"])
            book.save(result / "问题一求解结果.xlsx")
            report = syncer.synchronize(root, write=True)
            updated = yaml.safe_load(state_path.read_text(encoding="utf-8"))
            entry = updated["subproblems"]["Q1"]
            self.assertEqual(report["stale_questions"], ["Q1"])
            self.assertIn("solution_workbook", entry["stale_layers"])
            self.assertEqual(entry["result_quality_status"], "pending")
            self.assertEqual(entry["result_analysis_status"], "pending")

    def test_figure_scope_checks_declared_export(self):
        syncer = load_syncer()
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            result = setup_project(
                root, status="analyzed", phase="figure_evidence"
            )
            (result / "q1_plot.m").write_text(
                'readcell("问题一求解结果.xlsx"); title(gca, "结果"); '
                'exportgraphics(gca, "图表/missing.png");',
                encoding="utf-8",
            )
            report = syncer.synchronize(
                root, write=False, delivery_scope="figures"
            )
            self.assertTrue(
                any("声明导出的图不存在" in issue for issue in report["issues"])
            )


if __name__ == "__main__":
    unittest.main()
