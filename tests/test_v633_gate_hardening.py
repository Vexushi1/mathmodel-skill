import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path

import yaml
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]


def load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


SYNC = load_module("sync_v633", ROOT / "scripts/sync_project.py")
SCHEMA = yaml.safe_load((ROOT / "core/workbook_schema.yaml").read_text(encoding="utf-8"))


def append_sheet(book: Workbook, title: str, headers, row):
    if len(book.sheetnames) == 1 and book.active["A1"].value is None:
        sheet = book.active
        sheet.title = title
        sheet.delete_rows(1, 1)
    else:
        sheet = book.create_sheet(title)
    sheet.append(list(headers))
    sheet.append(list(row))


def write_solution(path: Path):
    book = Workbook()
    append_sheet(book, "核心指标", ["指标", "数值"], ["目标值", 1.0])
    append_sheet(book, "数据审计", ["等级", "检查项", "信息", "处理方式"], ["Info", "字段", "通过", "无"])
    append_sheet(book, "主结果质量门", ["检查项", "是否通过", "证据"], ["收敛", True, "通过"])
    append_sheet(book, "推荐方案", ["方案"], ["A"])
    book.save(path)


def write_analysis(path: Path):
    book = Workbook()
    append_sheet(
        book,
        "分析设计",
        ["风险来源", "分析问题", "方法", "指标", "通过标准"],
        ["算法", "一致性", "多算法", "差异", "小于1%"],
    )
    append_sheet(book, "算法一致性", ["算法", "重复编号", "目标值", "是否可行"], ["A", 1, 1.0, True])
    append_sheet(
        book,
        "结论稳定性汇总",
        ["核心结论", "分析方法", "稳定范围", "是否保持"],
        ["方案A", "多算法", "三种算法", True],
    )
    book.save(path)


def capabilities():
    return {name: False for name in SCHEMA["capability_contract"]["allowed"]}


def framework_text() -> str:
    return (
        "# 模型论文框架\n\n"
        "> 本文件只保留当前有效版本。\n\n"
        "- 最近同步：`design`\n"
        "- 最近同步时间：`x`\n"
        "- 当前状态：`current`\n\n"
        "## 当前有效口径\n\n"
        "## 各问模型与结果\n\n"
        "### Q1\n\n"
        "## 图表证据链\n\n"
        "## 待办与缺口\n"
    )


def write_state(root: Path, *, status="designed", phase="model_design", stale=False):
    (root / "state").mkdir(parents=True, exist_ok=True)
    solved = status in {"solved", "analyzed", "validated", "written", "completed"}
    analyzed = status in {"analyzed", "validated", "written", "completed"}
    payload = {
        "project": {"competition": "测试", "problem": "A", "current_phase": phase},
        "requirements": {"total": 0, "completed": [], "pending": []},
        "decisions": {},
        "subproblems": {
            "Q1": {
                "status": status,
                "selected_model": "测试模型",
                "classification": {"objective": "optimization", "structures": []},
                "capabilities": capabilities(),
                "result_quality_status": "passed" if solved else "pending",
                "result_analysis_status": "passed" if analyzed else "pending",
                "validation_status": "passed" if status in {"validated", "written", "completed"} else "pending",
                "framework_section": "### Q1",
                "result_summary_status": "stale" if stale else "current" if solved else "pending",
                "result_summary_anchor": "### Q1" if solved else "",
                "artifacts_stale": stale,
                "stale_layers": ["model"] if stale else [],
                "analysis_methods": ["算法一致性"] if analyzed else [],
                "artifact_hashes": {"model": "a" * 64} if stale else {},
                "validated_artifact_hashes": {"model": "b" * 64} if stale else {},
                "evidence": ["evidence"] if status in {"validated", "written", "completed"} else [],
                "optimality_claim": "none",
            }
        },
        "variables": {"locked": [], "source": {}},
        "paper_framework": {
            "path": "模型论文框架.md",
            "version": "1",
            "mode": "compact",
            "sync_status": "stale" if stale else "current",
            "last_sync_scope": "design",
            "proposition_limit": 4,
            "proposition_count": 0,
            "proposition_status": "not_assessed",
            "propositions": [],
        },
        "artifacts": {"code": [], "results": [], "figures": [], "papers": []},
        "risks": [],
        "next_gate": {"module": "solve_validate", "condition": "test"},
    }
    path = root / "state/project_state.yaml"
    path.write_text(yaml.safe_dump(payload, allow_unicode=True, sort_keys=False), encoding="utf-8")
    return path


def setup_project(root: Path, *, status="designed", phase="model_design", stale=False):
    (root / "模型论文框架.md").write_text(framework_text(), encoding="utf-8")
    write_state(root, status=status, phase=phase, stale=stale)
    (root / "问题一求解.py").write_text(
        "def main():\n    return 1\n\nif __name__ == '__main__':\n    main()\n",
        encoding="utf-8",
    )
    result = root / "结果数据表/问题一"
    result.mkdir(parents=True, exist_ok=True)
    write_solution(result / "问题一求解结果.xlsx")
    write_analysis(result / "问题一结果深化分析.xlsx")
    return result


class TestV633GateHardening(unittest.TestCase):
    def test_empty_state_and_arbitrary_framework_fail_preflight(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            (root / "state").mkdir()
            (root / "state/project_state.yaml").write_text("{}\n", encoding="utf-8")
            (root / "模型论文框架.md").write_text("任意内容\n", encoding="utf-8")
            report = SYNC.synchronize(root, write=False, delivery_scope="design")
            joined = "\n".join(report["issues"])
            self.assertIn("项目状态校验", joined)
            self.assertIn("模型论文框架校验", joined)

    def test_figures_scope_cannot_be_bypassed_by_designed_status(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            setup_project(root, status="designed", phase="model_design")
            report = SYNC.synchronize(root, write=False, delivery_scope="figures")
            joined = "\n".join(report["issues"])
            self.assertIn("图表交付缺少MATLAB脚本", joined)

    def test_sync_never_clears_existing_stale(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            setup_project(root, status="designed", phase="model_design", stale=True)
            SYNC.synchronize(root, write=True, delivery_scope="design")
            state = yaml.safe_load((root / "state/project_state.yaml").read_text(encoding="utf-8"))
            self.assertTrue(state["subproblems"]["Q1"]["artifacts_stale"])
            self.assertEqual(state["paper_framework"]["sync_status"], "stale")

    def test_manifest_references_output_contract_stage_requirements(self):
        manifest = yaml.safe_load((ROOT / "core/module_manifest.yaml").read_text(encoding="utf-8"))
        output = yaml.safe_load((ROOT / "core/output_contract.yaml").read_text(encoding="utf-8"))
        gate = manifest["utility_gates"]["project_sync"]
        self.assertNotIn("stage_requirements", gate)
        self.assertEqual(
            gate["stage_requirements_source"],
            "core/output_contract.yaml#project_sync.stage_requirements",
        )
        self.assertEqual(output["project_sync"]["stage_requirements_semantics"], "exact_scope")

    def test_first_figure_evidence_generation_updates_state_evidence(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            result = setup_project(root, status="analyzed", phase="figure_evidence")
            (result / "图表").mkdir()
            (result / "图表/q1.png").write_bytes(b"figure")
            (result / "q1_plot.m").write_text(
                'raw = readcell("问题一求解结果.xlsx");\n'
                'title(gca, "结果");\n'
                'exportgraphics(gca, "图表/q1.png");\n',
                encoding="utf-8",
            )
            report = SYNC.synchronize(root, write=True, delivery_scope="figures")
            self.assertFalse(any("图表交付缺少MATLAB脚本" in issue for issue in report["issues"]))
            state = yaml.safe_load((root / "state/project_state.yaml").read_text(encoding="utf-8"))
            evidence_path = "结果数据表/问题一/figure_evidence.yaml"
            self.assertTrue((root / evidence_path).is_file())
            self.assertIn(evidence_path, state["subproblems"]["Q1"]["evidence"])

    def test_latex_scope_uses_exact_contract_not_cumulative_results(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            (root / "模型论文框架.md").write_text(framework_text(), encoding="utf-8")
            state_path = write_state(root, status="designed", phase="writing_latex")
            state = yaml.safe_load(state_path.read_text(encoding="utf-8"))
            figures = root / "figures"
            figures.mkdir()
            approved = figures / "approved.png"
            approved.write_bytes(b"x")
            state["artifacts"]["approved_figures"] = ["figures/approved.png"]
            state_path.write_text(yaml.safe_dump(state, allow_unicode=True, sort_keys=False), encoding="utf-8")
            final = root / "final_latex"
            final.mkdir()
            (final / "main.tex").write_text("\\documentclass{article}\\begin{document}x\\end{document}", encoding="utf-8")
            (final / "main.pdf").write_bytes(b"pdf")
            (final / "compile_report.yaml").write_text(
                yaml.safe_dump({"status": "passed", "unresolved_references": 0}),
                encoding="utf-8",
            )
            report = SYNC.synchronize(root, write=False, delivery_scope="latex")
            joined = "\n".join(report["issues"])
            self.assertNotIn("标准求解结果工作簿", joined)
            self.assertNotIn("结果深化分析工作簿", joined)
            self.assertNotIn("问题求解Python脚本", joined)


if __name__ == "__main__":
    unittest.main()
