from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[1]


def load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


NUMERICAL = load_module(
    "validate_numerical_evidence_v714", ROOT / "scripts" / "validate_numerical_evidence.py"
)


QUALITY_COLUMNS = [
    "Verification ID", "检查项", "是否通过", "证据", "判定关系",
    "阈值或容差", "实际值", "证据工作表", "阈值来源",
]


class V714NumericalVerificationTests(unittest.TestCase):
    def make_book(self, path: Path, quality_row: list, sheet: str, headers: list, row: list) -> Path:
        book = openpyxl.Workbook()
        quality = book.active
        quality.title = "主结果质量门"
        quality.append(QUALITY_COLUMNS)
        quality.append(quality_row)
        evidence = book.create_sheet(sheet)
        evidence.append(headers)
        evidence.append(row)
        book.save(path)
        return path

    def test_contract_draws_hard_primary_analysis_boundary(self):
        contract = yaml.safe_load(
            (ROOT / "core/numerical_verification_contract.yaml").read_text(encoding="utf-8")
        )
        rule = contract["authority_boundary"]["primary_rule"]
        self.assertIn("alternative algorithms", rule)
        self.assertIn("result analysis after the primary workbook is accepted", rule)
        forbidden = set(contract["scope"]["forbidden_as_primary_quality"])
        for token in (
            "parameter_sensitivity",
            "scenario_stress_testing",
            "alternative_algorithm_comparison",
            "structural_robustness",
            "heterogeneity_analysis",
        ):
            self.assertIn(token, forbidden)

    def test_feasibility_contradiction_is_rejected(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "问题一求解结果.xlsx"
            self.make_book(
                path,
                ["PQ-Q1-01", "可行性", True, "rows", "<=", 1e-6, 1e-2, "约束违反检查", "solver_tolerance"],
                "约束违反检查",
                ["约束编号", "约束含义", "违反量", "容差", "是否满足"],
                ["C1", "x<=1", 1e-2, 1e-6, True],
            )
            passed, issues, _ = NUMERICAL.validate_primary_numerical_evidence(
                path, {"requires_feasibility_check": True}
            )
            self.assertFalse(passed)
            self.assertTrue(any("是否满足" in item for item in issues))
            self.assertTrue(any("是否通过" in item for item in issues))

    def test_consistent_failed_residual_row_cannot_be_hidden_by_lax_summary_threshold(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "问题一求解结果.xlsx"
            self.make_book(
                path,
                ["PQ-Q1-01", "守恒", True, "rows", "<=", 1.0, 1e-2, "守恒残差", "locked_model_tolerance"],
                "守恒残差",
                ["守恒量", "残差", "容差", "是否满足"],
                ["质量", 1e-2, 1e-6, False],
            )
            passed, issues, _ = NUMERICAL.validate_primary_numerical_evidence(
                path, {"requires_conservation_residual": True}
            )
            self.assertFalse(passed)
            self.assertTrue(any("未满足容差" in item for item in issues), issues)

    def test_residual_actual_is_recomputed_from_bottom_level_evidence(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "问题一求解结果.xlsx"
            self.make_book(
                path,
                ["PQ-Q1-01", "守恒", True, "rows", "<=", 1e-6, 1e-9, "守恒残差", "locked_model_tolerance"],
                "守恒残差",
                ["守恒量", "残差", "容差", "是否满足"],
                ["质量", 5e-7, 1e-6, True],
            )
            passed, issues, _ = NUMERICAL.validate_primary_numerical_evidence(
                path, {"requires_conservation_residual": True}
            )
            self.assertFalse(passed)
            self.assertTrue(any("重算指标不一致" in item for item in issues))

    def test_discretization_requires_marked_primary_evidence_and_matches_metric(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "问题一求解结果.xlsx"
            self.make_book(
                path,
                ["PQ-Q1-01", "离散精度", True, "rows", "<=", 1e-3, 5e-4, "离散精度", "numerical_method_accuracy_target"],
                "离散精度",
                ["离散参数", "取值", "目标指标", "相对变化", "用于主判定"],
                ["dt", 0.01, "J", 5e-4, True],
            )
            passed, issues, report = NUMERICAL.validate_primary_numerical_evidence(
                path, {"requires_discretization_check": True}
            )
            self.assertTrue(passed, issues)
            self.assertTrue(report["strict"])

    def test_failed_quality_relation_is_blocking_even_when_declared_false(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "问题一求解结果.xlsx"
            self.make_book(
                path,
                ["PQ-Q1-01", "离散精度", False, "rows", "<=", 1e-4, 5e-4, "离散精度", "numerical_method_accuracy_target"],
                "离散精度",
                ["离散参数", "取值", "目标指标", "相对变化", "用于主判定"],
                ["dt", 0.01, "J", 5e-4, True],
            )
            passed, issues, _ = NUMERICAL.validate_primary_numerical_evidence(
                path, {"requires_discretization_check": True}
            )
            self.assertFalse(passed)
            self.assertTrue(any("未达到主质量判据" in item for item in issues), issues)

    def test_convergence_marker_is_not_result_analysis(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "问题一求解结果.xlsx"
            self.make_book(
                path,
                ["PQ-Q1-01", "主迭代收敛", True, "rows", "bool_true", True, True, "收敛诊断", "solver_tolerance"],
                "收敛诊断",
                ["迭代或样本数", "指标", "数值", "判定", "用于主判定"],
                [100, "residual", 1e-9, True, True],
            )
            passed, issues, _ = NUMERICAL.validate_primary_numerical_evidence(
                path, {"requires_convergence_diagnostic": True}
            )
            self.assertTrue(passed, issues)

    def test_strict_quality_gate_cannot_be_empty(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "问题一求解结果.xlsx"
            book = openpyxl.Workbook()
            quality = book.active
            quality.title = "主结果质量门"
            quality.append(QUALITY_COLUMNS)
            book.save(path)
            passed, issues, _ = NUMERICAL.validate_primary_numerical_evidence(
                path, {}, force_strict=True
            )
            self.assertFalse(passed)
            self.assertTrue(any("至少需要一行严格Verification证据" in item for item in issues), issues)

    def test_missing_evidence_sheet_breaks_strict_trace(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "问题一求解结果.xlsx"
            book = openpyxl.Workbook()
            quality = book.active
            quality.title = "主结果质量门"
            quality.append(QUALITY_COLUMNS)
            quality.append([
                "PQ-Q1-01", "可行性", True, "rows", "<=", 1e-6, 0.0,
                "约束违反检查", "solver_tolerance",
            ])
            book.save(path)
            passed, issues, _ = NUMERICAL.validate_primary_numerical_evidence(
                path, {"requires_feasibility_check": True}
            )
            self.assertFalse(passed)
            self.assertTrue(any("不存在" in item or "需要工作表" in item for item in issues))

    def test_invalid_verification_id_is_rejected(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "问题一求解结果.xlsx"
            self.make_book(
                path,
                ["PQ-Q1-X", "离散精度", True, "rows", "<=", 1e-3, 5e-4, "离散精度", "numerical_method_accuracy_target"],
                "离散精度",
                ["离散参数", "取值", "目标指标", "相对变化", "用于主判定"],
                ["dt", 0.01, "J", 5e-4, True],
            )
            passed, issues, _ = NUMERICAL.validate_primary_numerical_evidence(
                path, {"requires_discretization_check": True}
            )
            self.assertFalse(passed)
            self.assertTrue(any("PQS格式" in item for item in issues))

    def test_unregistered_threshold_source_is_rejected(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "问题一求解结果.xlsx"
            self.make_book(
                path,
                ["PQ-Q1-01", "离散精度", True, "rows", "<=", 1e-3, 5e-4, "离散精度", "looks_good_after_run"],
                "离散精度",
                ["离散参数", "取值", "目标指标", "相对变化", "用于主判定"],
                ["dt", 0.01, "J", 5e-4, True],
            )
            passed, issues, _ = NUMERICAL.validate_primary_numerical_evidence(
                path, {"requires_discretization_check": True}
            )
            self.assertFalse(passed)
            self.assertTrue(any("未登记阈值来源" in item for item in issues))

    def test_negative_abs_tolerance_is_rejected(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "问题一求解结果.xlsx"
            self.make_book(
                path,
                ["PQ-Q1-01", "离散精度", True, "rows", "abs<=", -1e-3, 5e-4, "离散精度", "numerical_method_accuracy_target"],
                "离散精度",
                ["离散参数", "取值", "目标指标", "相对变化", "用于主判定"],
                ["dt", 0.01, "J", 5e-4, True],
            )
            passed, issues, _ = NUMERICAL.validate_primary_numerical_evidence(
                path, {"requires_discretization_check": True}
            )
            self.assertFalse(passed)
            self.assertTrue(any("不得为负数" in item for item in issues), issues)

    def test_v713_boolean_gate_remains_legacy_read_compatible(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "legacy.xlsx"
            book = openpyxl.Workbook()
            quality = book.active
            quality.title = "主结果质量门"
            quality.append(["检查项", "是否通过", "证据"])
            quality.append(["完整运行", True, "legacy"])
            book.save(path)
            passed, issues, report = NUMERICAL.validate_primary_numerical_evidence(path, {})
            self.assertTrue(passed, issues)
            self.assertEqual(report["mode"], "legacy_read")

    def test_runtime_and_receipt_integration_are_declared(self):
        bootstrap = yaml.safe_load((ROOT / "core/bootstrap.yaml").read_text(encoding="utf-8"))
        manifest = yaml.safe_load((ROOT / "core/module_manifest.yaml").read_text(encoding="utf-8"))
        assurance = yaml.safe_load(
            (ROOT / "core/runtime_assurance_contract.yaml").read_text(encoding="utf-8")
        )
        receipt = (ROOT / "scripts/validate_user_execution.py").read_text(encoding="utf-8")
        self.assertEqual(
            bootstrap["authoritative_sources"]["numerical_verification"],
            "core/numerical_verification_contract.yaml",
        )
        self.assertEqual(
            manifest["contracts"]["numerical_verification"],
            "core/numerical_verification_contract.yaml",
        )
        self.assertIn(
            "numerical_verification",
            assurance["contract_dependency_closure"]["contract_dependencies"]["modules/03_solve_validate.md"],
        )
        self.assertIn("validate_primary_numerical_evidence", receipt)
        self.assertIn("delivered_primary_protocol", receipt)

    def test_primary_and_analysis_modules_keep_stage_ownership(self):
        solve = (ROOT / "modules/03_solve_validate.md").read_text(encoding="utf-8")
        analysis = (ROOT / "modules/03_result_analysis.md").read_text(encoding="utf-8")
        self.assertIn("主求解质量检查", solve)
        self.assertIn("不得把参数敏感性", solve)
        self.assertIn("主工作簿 accepted 后", analysis)
        self.assertIn("不得被主求解质量门提前吸收", analysis)


if __name__ == "__main__":
    unittest.main()
