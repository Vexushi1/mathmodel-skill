#!/usr/bin/env python3
"""Check HSK v6.2.2 project structure, workbook contract and software ownership."""
from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from typing import Any, Mapping, Sequence

import yaml
from openpyxl import load_workbook

CN_NUM = "一二三四五六七八九十百"
FIG_EXT = {".png", ".pdf", ".svg", ".jpg", ".jpeg", ".tif", ".tiff"}
PY_PLOT_TOKENS = ("matplotlib", "seaborn", "savefig(", "plt.show(")
ROOT = Path(__file__).resolve().parent.parent
SCHEMA_PATH = ROOT / "core" / "workbook_schema.yaml"


def load_schema() -> dict[str, Any]:
    return yaml.safe_load(SCHEMA_PATH.read_text(encoding="utf-8")) or {}


def check_code(root: Path) -> list[str]:
    issues: list[str] = []
    py_dir, matlab_dir = root / "Python求解", root / "MATLAB绘图"
    if not py_dir.exists():
        issues.append("missing: Python求解/")
    if not matlab_dir.exists():
        issues.append("missing: MATLAB绘图/")
    for path in py_dir.rglob("*.py") if py_dir.exists() else []:
        text = path.read_text(encoding="utf-8", errors="ignore")
        if any(token in text for token in PY_PLOT_TOKENS):
            issues.append(f"Python formal-plot ownership violation: {path.relative_to(root)}")
        if "if __name__" not in text and any(key in path.stem for key in ("求解", "敏感性", "鲁棒性")):
            issues.append(f"Python main script lacks entry point: {path.relative_to(root)}")
    return issues


def worksheet_has_data(worksheet) -> bool:
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_index == 1:
            continue
        if any(value not in (None, "") for value in row):
            return True
    return False


def worksheet_headers(worksheet) -> list[str]:
    row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [str(value).strip() if value is not None else "" for value in row]


def required_solution_sheets(schema: Mapping[str, Any], problem_types: Sequence[str]) -> set[str]:
    section = schema["solution_workbook"]
    required = set(section.get("common_required_sheets", {}))
    selected = set(problem_types)
    for config in section.get("conditional_requirements", {}).values():
        if selected.intersection(config.get("problem_types", [])):
            required.update(config.get("required_sheets", []))
    return required


def sheet_schema_map(schema: Mapping[str, Any], kind: str) -> dict[str, Mapping[str, Any]]:
    if kind == "solution":
        section = schema["solution_workbook"]
        return {
            **dict(section.get("common_required_sheets", {})),
            **dict(section.get("common_recommended_sheets", {})),
        }
    return dict(schema["sensitivity_robustness_workbook"].get("sheet_schemas", {}))


def check_sheet_values(path: Path, worksheet) -> list[str]:
    issues: list[str] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        for value in row:
            if isinstance(value, float) and not math.isfinite(value):
                issues.append(f"non-finite numeric value: {path} -> {worksheet.title}")
                return issues
    return issues


def inspect_workbook(path: Path, kind: str, problem_types: Sequence[str] = ()) -> list[str]:
    issues: list[str] = []
    schema = load_schema()
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return [f"cannot open workbook {path}: {exc}"]

    try:
        names = set(workbook.sheetnames)
        if kind == "solution":
            missing = required_solution_sheets(schema, problem_types) - names
            if missing:
                issues.append(f"solution workbook missing sheets {sorted(missing)}: {path}")
            profiles = schema.get("solution_workbook", {}).get("task_profiles", {})
            for problem_type in problem_types:
                required_any = set(profiles.get(problem_type, {}).get("required_any", []))
                if required_any and not names.intersection(required_any):
                    issues.append(
                        f"solution workbook lacks task-specific sheet for {problem_type} "
                        f"{sorted(required_any)}: {path}"
                    )
        else:
            allowed = set(schema["sensitivity_robustness_workbook"].get("required_any_sheets", []))
            if not names.intersection(allowed):
                issues.append(f"robustness workbook lacks analysis or applicability sheet: {path}")

        schemas = sheet_schema_map(schema, kind)
        for worksheet in workbook.worksheets:
            if len(worksheet.title) > 31:
                issues.append(f"worksheet name exceeds 31 characters: {path} -> {worksheet.title}")
            if not worksheet_has_data(worksheet):
                issues.append(f"empty worksheet is forbidden: {path} -> {worksheet.title}")
                continue

            headers = worksheet_headers(worksheet)
            if len(headers) != len(set(headers)):
                issues.append(f"duplicate worksheet columns: {path} -> {worksheet.title}")
            spec = schemas.get(worksheet.title, {})
            required_columns = [str(item) for item in spec.get("required_columns", [])]
            missing_columns = [column for column in required_columns if column not in headers]
            if missing_columns:
                issues.append(
                    f"worksheet missing required columns {missing_columns}: {path} -> {worksheet.title}"
                )
            issues.extend(check_sheet_values(path, worksheet))
    finally:
        workbook.close()
    return issues


def resolve_problem_types(root: Path, explicit: Sequence[str]) -> tuple[str, ...]:
    if explicit:
        return tuple(dict.fromkeys(explicit))
    state_path = root / "state" / "project_state.yaml"
    if not state_path.is_file():
        return ()
    payload = yaml.safe_load(state_path.read_text(encoding="utf-8")) or {}
    values = payload.get("project", {}).get("problem_types", [])
    return tuple(str(item) for item in values)


def check_results(root: Path, problem_types: Sequence[str]) -> list[str]:
    issues: list[str] = []
    base = root / "结果数据表"
    if not base.exists():
        return ["missing: 结果数据表/"]
    questions = [
        path for path in base.iterdir()
        if path.is_dir() and re.fullmatch(rf"问题[{CN_NUM}]+", path.name)
    ]
    if not questions:
        return ["missing: no 结果数据表/问题X/ directories"]
    for question in questions:
        data_dir = question / f"{question.name}结果数据"
        if not data_dir.exists():
            issues.append(f"missing: {data_dir.relative_to(root)}")
            continue
        solution = data_dir / f"{question.name}求解结果.xlsx"
        robustness = data_dir / f"{question.name}敏感性与鲁棒性结果.xlsx"
        for path, kind in ((solution, "solution"), (robustness, "robustness")):
            if not path.is_file():
                issues.append(f"missing: {path.relative_to(root)}")
            else:
                issues.extend(inspect_workbook(path, kind, problem_types))
    return issues


def check_figures(root: Path) -> list[str]:
    issues: list[str] = []
    for dirname in ("figures", "figures_editable"):
        directory = root / dirname
        if not directory.exists():
            continue
        for path in directory.rglob("*"):
            if path.is_file() and path.suffix.lower() in FIG_EXT and any(ord(char) >= 128 for char in path.name):
                issues.append(f"figure filename should be ASCII: {path.relative_to(root)}")
    return issues


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("project", nargs="?", default=".")
    parser.add_argument("--mode", choices=["full", "code", "data", "figures"], default="full")
    parser.add_argument(
        "--problem-types",
        nargs="*",
        default=[],
        help="题型标签；未提供时尝试从 state/project_state.yaml 读取",
    )
    args = parser.parse_args()
    root = Path(args.project).resolve()
    problem_types = resolve_problem_types(root, args.problem_types)

    issues: list[str] = []
    if args.mode in {"full", "code"}:
        issues += check_code(root)
    if args.mode in {"full", "data"}:
        issues += check_results(root, problem_types)
    if args.mode in {"full", "figures"}:
        issues += check_figures(root)
    if issues:
        print("HSK artifact check: ISSUES FOUND")
        for item in issues:
            print("-", item)
        return 1
    print("HSK artifact check: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
