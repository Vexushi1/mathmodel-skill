#!/usr/bin/env python3
"""Validate user-produced preprocessing/solve/analysis workbooks and advance state only after evidence passes."""
from __future__ import annotations

import argparse
import hashlib
from pathlib import Path
from typing import Any

import openpyxl
import yaml

FALSE_FLAGS = (
    "allow_reduced_data", "allow_coarser_grid", "allow_shorter_horizon",
    "allow_fewer_repetitions", "allow_relaxed_tolerance",
    "allow_silent_solver_fallback",
)
VALID_DECISIONS = {"not_needed", "question_local", "project_level"}
PREPROCESSING_EVIDENCE_SHEETS = {
    "运行配置": ("项目", "值"),
    "数据审计": ("数据源", "检查项", "结论", "处理方式"),
    "预处理参数": ("步骤", "参数", "取值", "单位", "依据"),
    "预处理方法证据": ("步骤", "问题证据", "数学方法", "参数依据", "验证方法", "验证结论"),
    "处理前后对比": ("对象", "指标", "处理前", "处理后"),
    "绘图数据索引": ("图组", "图作用", "源工作表"),
    "预处理质量门": ("检查项", "是否通过", "证据"),
}


def load_yaml(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def question_key(problem: str) -> str:
    order = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十"]
    suffix = problem.removeprefix("问题")
    return f"Q{order.index(suffix) + 1}" if suffix in order else problem


def as_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "是", "通过", "满足"}:
        return True
    if text in {"false", "0", "no", "否", "未通过", "不满足"}:
        return False
    return None


def is_sha256(value: Any) -> bool:
    text = str(value).strip().lower()
    return len(text) == 64 and all(character in "0123456789abcdef" for character in text)


def configuration_map(workbook: Path) -> tuple[dict[str, Any], list[str]]:
    book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    if "运行配置" not in book.sheetnames:
        return {}, ["缺少运行配置工作表"]
    rows = list(book["运行配置"].iter_rows(values_only=True))
    if not rows or tuple(rows[0][:2]) != ("项目", "值"):
        return {}, ["运行配置表头必须为项目|值"]
    mapping = {
        str(row[0]).strip(): row[1]
        for row in rows[1:]
        if row and row[0] not in (None, "")
    }
    required = {
        "execution_owner", "execution_profile", "stage", "problem_name", "code_sha256",
        "data_sha256", "solver", "solver_version", "tolerance", "iteration_or_time_limit",
        "actual_stop_reason", "random_seed", "repetitions_or_scenarios", "grid_or_time_range",
        "fallback_used", "platform", *FALSE_FLAGS,
    }
    issues = [f"运行配置缺少项目: {item}" for item in sorted(required - set(mapping))]
    if "code_sha256" in mapping and not is_sha256(mapping["code_sha256"]):
        issues.append("运行配置code_sha256必须为64位十六进制SHA-256")
    if "data_sha256" in mapping and not is_sha256(mapping["data_sha256"]):
        issues.append("运行配置data_sha256必须为64位十六进制SHA-256")
    return mapping, issues


def workbook_identity(root: Path, workbook: Path) -> tuple[str, str, list[str]]:
    root = root.resolve()
    workbook = workbook.resolve()
    try:
        workbook.relative_to(root)
    except ValueError:
        return "", "", ["工作簿路径越出项目根目录"]

    if workbook == (root / "数据预处理" / "数据预处理结果.xlsx").resolve():
        return "数据预处理", "preprocessing", []

    parent = workbook.parent
    problem = ""
    if parent.parent == root and parent.name.endswith("求解"):
        problem = parent.name.removesuffix("求解")
    elif parent.parent.name == "结果数据表" and parent.parent.parent == root:
        problem = parent.name
    else:
        return "", "", ["工作簿必须位于数据预处理/、问题X求解/或旧版结果数据表/问题X/目录"]

    if not problem.startswith("问题") or len(problem) <= len("问题"):
        return "", "", ["工作簿目录无法解析有效问题编号"]

    if workbook.name == f"{problem}求解结果.xlsx":
        stage = "primary"
    elif workbook.name == f"{problem}结果深化分析.xlsx":
        stage = "analysis"
    else:
        return problem, "", [
            f"工作簿名必须为{problem}求解结果.xlsx或{problem}结果深化分析.xlsx"
        ]
    return problem, stage, []


def _boolean_gate(workbook: Path, sheet: str, column: str) -> tuple[bool, list[str]]:
    book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    if sheet not in book.sheetnames:
        return False, [f"缺少{sheet}工作表"]
    rows = list(book[sheet].iter_rows(values_only=True))
    if not rows:
        return False, [f"{sheet}为空"]
    headers = [str(item) if item is not None else "" for item in rows[0]]
    if column not in headers:
        return False, [f"{sheet}缺少{column}列"]
    index = headers.index(column)
    failures = [
        str(row[0]) for row in rows[1:]
        if row and as_bool(row[index] if len(row) > index else None) is not True
    ]
    return not failures, [f"{sheet}未通过: {item}" for item in failures]


def _evidence_sheet_issues(book: openpyxl.Workbook, sheet: str, required: tuple[str, ...]) -> list[str]:
    if sheet not in book.sheetnames:
        return [f"预处理工作簿缺少工作表: {sheet}"]
    rows = list(book[sheet].iter_rows(values_only=True))
    if len(rows) < 2:
        return [f"{sheet}必须包含表头和至少一行实质证据"]
    headers = [str(item).strip() if item is not None else "" for item in rows[0]]
    missing = [column for column in required if column not in headers]
    issues = [f"{sheet}缺少列: {column}" for column in missing]
    if not any(any(value not in (None, "") for value in row) for row in rows[1:]):
        issues.append(f"{sheet}没有实质数据")
    return issues


def preprocessing_passed(workbook: Path) -> tuple[bool, list[str]]:
    book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    issues: list[str] = []
    for sheet, required_columns in PREPROCESSING_EVIDENCE_SHEETS.items():
        issues.extend(_evidence_sheet_issues(book, sheet, required_columns))
    if issues:
        return False, list(dict.fromkeys(issues))

    # 绘图数据索引必须至少指向一个真实、非空的底层数据工作表，避免MATLAB从摘要数字反推。
    index_rows = list(book["绘图数据索引"].iter_rows(values_only=True))
    headers = [str(item).strip() if item is not None else "" for item in index_rows[0]]
    source_idx = headers.index("源工作表")
    referenced = {
        str(row[source_idx]).strip()
        for row in index_rows[1:]
        if len(row) > source_idx and row[source_idx] not in (None, "")
    }
    if not referenced:
        issues.append("绘图数据索引至少必须登记一个源工作表")
    else:
        for sheet in sorted(referenced):
            if sheet not in book.sheetnames:
                issues.append(f"绘图数据索引引用不存在的工作表: {sheet}")
                continue
            rows = list(book[sheet].iter_rows(values_only=True))
            if len(rows) < 2:
                issues.append(f"绘图数据索引引用的工作表无底层数据: {sheet}")

    gate_passed, gate_issues = _boolean_gate(workbook, "预处理质量门", "是否通过")
    issues.extend(gate_issues)
    return gate_passed and not issues, list(dict.fromkeys(issues))


def quality_passed(workbook: Path) -> tuple[bool, list[str]]:
    return _boolean_gate(workbook, "主结果质量门", "是否通过")


def analysis_passed(workbook: Path) -> tuple[bool, str, list[str]]:
    book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    required = {"分析设计", "结论稳定性汇总"}
    missing = sorted(required - set(book.sheetnames))
    if missing:
        return False, "failed", [f"缺少工作表: {missing}"]
    rows = list(book["结论稳定性汇总"].iter_rows(values_only=True))
    if len(rows) < 2:
        return False, "failed", ["结论稳定性汇总无实质数据"]
    headers = [str(item) if item is not None else "" for item in rows[0]]
    if "是否保持" not in headers:
        return False, "failed", ["结论稳定性汇总缺少是否保持列"]
    index = headers.index("是否保持")
    unstable = [
        row for row in rows[1:]
        if row and as_bool(row[index] if len(row) > index else None) is not True
    ]
    return (
        not unstable,
        "passed" if not unstable else "redo_required",
        ["存在核心结论未保持"] if unstable else [],
    )


def validate_execution_evidence(
    config: dict[str, Any], state: dict[str, Any], entry: dict[str, Any], stage: str
) -> list[str]:
    issues: list[str] = []
    if config.get("execution_owner") != "user":
        issues.append("execution_owner必须为user")
    if config.get("execution_profile") != "full_fidelity":
        issues.append("execution_profile必须为full_fidelity")
    if as_bool(config.get("fallback_used")) is not False:
        issues.append("fallback_used必须为false")
    for flag in FALSE_FLAGS:
        if as_bool(config.get(flag)) is not False:
            issues.append(f"{flag}必须为false")

    if stage == "preprocessing":
        preprocessing = state.get("preprocessing") or {}
        if preprocessing.get("decision") != "project_level":
            issues.append("只有preprocessing.decision=project_level时才允许验收数据预处理结果.xlsx")
        expected_code_hash = preprocessing.get("code_sha256")
        expected_data_hash = ((state.get("data") or {}).get("version_hashes") or {}).get("preprocessing_input")
    else:
        expected_code_hash = (
            entry.get("analysis_code_sha256") if stage == "analysis"
            else entry.get("primary_code_sha256")
        )
        expected_data_hash = entry.get("data_hash")
        decision = str(((state.get("preprocessing") or {}).get("decision", "")))
        if decision not in VALID_DECISIONS:
            issues.append("项目状态缺少有效preprocessing.decision")
        if stage == "primary" and decision == "project_level":
            preprocessing = state.get("preprocessing") or {}
            if preprocessing.get("status") != "accepted" or preprocessing.get("quality_status") != "passed":
                issues.append("project_level项目的预处理工作簿尚未accepted/passed")

    if not expected_code_hash:
        issues.append("项目状态缺少已交付代码哈希")
    elif str(config.get("code_sha256", "")).lower() != str(expected_code_hash).lower():
        issues.append("工作簿code_sha256与已交付代码不一致")
    if not expected_data_hash:
        issues.append("项目状态缺少代码交付时锁定的数据哈希")
    elif str(config.get("data_sha256", "")).lower() != str(expected_data_hash).lower():
        issues.append("工作簿data_sha256与代码交付时锁定的数据哈希不一致")
    return issues


def validate_one(root: Path, workbook: Path, state: dict[str, Any], write: bool) -> list[str]:
    config, issues = configuration_map(workbook)
    problem, stage, identity_issues = workbook_identity(root, workbook)
    issues.extend(identity_issues)
    if identity_issues:
        return list(dict.fromkeys(issues))

    configured_stage = str(config.get("stage", ""))
    configured_problem = str(config.get("problem_name", ""))
    if configured_stage != stage:
        issues.append(f"运行配置stage={configured_stage or '<missing>'}与工作簿文件名对应{stage}阶段不一致")
    if configured_problem != problem:
        issues.append("运行配置problem_name与工作簿目录/文件名不一致")
    if configured_stage != stage or configured_problem != problem:
        return list(dict.fromkeys(issues))

    key = question_key(problem)
    entry = {} if stage == "preprocessing" else (state.get("subproblems") or {}).get(key, {})
    issues.extend(validate_execution_evidence(config, state, entry, stage))

    if stage == "preprocessing":
        passed, quality_issues = preprocessing_passed(workbook)
        issues.extend(quality_issues)
        if write:
            preprocessing = state.setdefault("preprocessing", {})
            accepted = not issues and passed
            preprocessing["status"] = "accepted" if accepted else "rejected"
            preprocessing["quality_status"] = "passed" if accepted else "failed"
            preprocessing["workbook"] = workbook.relative_to(root).as_posix()
            preprocessing["workbook_sha256"] = file_hash(workbook)
            if accepted:
                state.setdefault("data", {})["active_source_mode"] = "preprocessed"
                state.setdefault("project", {})["current_phase"] = "solve_validate"
        return list(dict.fromkeys(issues))

    if stage == "primary":
        passed, quality_issues = quality_passed(workbook)
        issues.extend(quality_issues)
        if write:
            entry["primary_execution_status"] = "accepted" if not issues and passed else "rejected"
            entry["result_quality_status"] = "passed" if not issues and passed else "failed"
            entry["solution_workbook"] = workbook.relative_to(root).as_posix()
            entry.setdefault("artifact_hashes", {})["solution_workbook"] = file_hash(workbook)
            if not issues and passed:
                entry.setdefault("validated_artifact_hashes", {})["solution_workbook"] = file_hash(workbook)
                entry["status"] = "solved"
    else:
        passed, result_status, analysis_issues = analysis_passed(workbook)
        issues.extend(analysis_issues)
        if write:
            entry["analysis_execution_status"] = (
                "accepted" if not issues and passed
                else "redo_required" if result_status == "redo_required"
                else "rejected"
            )
            entry["result_analysis_status"] = result_status
            entry["result_analysis_workbook"] = workbook.relative_to(root).as_posix()
            entry.setdefault("artifact_hashes", {})["result_analysis_workbook"] = file_hash(workbook)
            if not issues and passed:
                entry.setdefault("validated_artifact_hashes", {})["result_analysis_workbook"] = file_hash(workbook)
                entry["status"] = "analyzed"
            elif result_status == "redo_required":
                entry["artifacts_stale"] = True
                entry["stale_layers"] = [
                    "result_analysis_workbook", "matlab_script", "figure_bundle", "framework"
                ]
                entry["result_summary_status"] = "stale"
                state.setdefault("project", {})["current_phase"] = "solve_validate"
    return list(dict.fromkeys(issues))


def discover(root: Path) -> list[Path]:
    current_patterns = (
        "数据预处理/数据预处理结果.xlsx",
        "问题*求解/问题*求解结果.xlsx",
        "问题*求解/问题*结果深化分析.xlsx",
    )
    legacy_patterns = (
        "结果数据表/问题*/问题*求解结果.xlsx",
        "结果数据表/问题*/问题*结果深化分析.xlsx",
    )
    return sorted({
        path.resolve()
        for pattern in (*current_patterns, *legacy_patterns)
        for path in root.glob(pattern)
    })


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("project_root", type=Path)
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--scope", default="results")
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--strict", action="store_true")
    args = parser.parse_args()
    root = args.project_root.resolve()
    state_path = root / "state" / "project_state.yaml"
    if not state_path.is_file():
        raise SystemExit("缺少state/project_state.yaml")
    state = load_yaml(state_path)
    workbooks = (
        [args.workbook if args.workbook.is_absolute() else root / args.workbook]
        if args.workbook else discover(root)
    )
    all_issues: list[str] = []
    checked: list[str] = []
    for workbook in workbooks:
        issues = validate_one(root, workbook, state, args.write)
        all_issues.extend(f"{workbook.name}: {item}" for item in issues)
        checked.append(workbook.relative_to(root).as_posix())
    if args.write:
        state_path.write_text(
            yaml.safe_dump(state, allow_unicode=True, sort_keys=False), encoding="utf-8"
        )
    report = {
        "status": "passed" if not all_issues else "failed",
        "checked_workbooks": checked,
        "issues": all_issues,
        "task_code_executed": False,
        "report_persisted": False,
    }
    print(yaml.safe_dump(report, allow_unicode=True, sort_keys=False).rstrip())
    if all_issues:
        print("\n".join(all_issues))
        return 1 if args.strict else 0
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
