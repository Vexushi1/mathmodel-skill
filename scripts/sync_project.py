#!/usr/bin/env python3
"""Synchronize project artifacts and enforce v6.3.2 delivery-gate contracts.

The synchronizer is conservative: it discovers and validates existing artifacts,
computes provenance hashes and propagates stale state. It never invents model
semantics, numerical results, figure approval or validation success.
"""
from __future__ import annotations

import argparse
import hashlib
import importlib.util
import json
import re
import sys
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping

import yaml
from openpyxl import load_workbook

SKILL_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SCHEMA_PATH = SKILL_ROOT / "core" / "workbook_schema.yaml"
QUESTION_RE = re.compile(r"问题([一二三四五六七八九十百]+)")
MATLAB_TITLE_RE = re.compile(r"\b(?:title|sgtitle)\s*\(", re.IGNORECASE)
EXPORT_RE = re.compile(r"(?:exportgraphics|print)\s*\([^\n]*?[\"']([^\"']+\.(?:png|pdf|svg|tif|tiff))[\"']", re.IGNORECASE)
WORKBOOK_REF_RE = re.compile(r"[\"']([^\"']+\.xlsx)[\"']", re.IGNORECASE)
LATEX_GRAPHICS_RE = re.compile(r"\\includegraphics(?:\[[^\]]*\])?\{([^}]+)\}")
FIGURE_SUFFIXES = {".png", ".pdf", ".svg", ".tif", ".tiff", ".jpg", ".jpeg"}
DATA_SUFFIXES = {".csv", ".xlsx", ".xls", ".json", ".yaml", ".yml", ".txt"}
IGNORED_ROOT_NAMES = {".git", ".idea", ".vscode", "__pycache__", "结果数据表", "draft_docx", "final_latex", "submission", "figures", "figures_editable", "state"}
STATUS_RANK = {"pending": 0, "audited": 1, "designed": 2, "solved": 3, "validated": 4, "written": 5, "completed": 6}
SCOPE_RANK = {"design": 0, "results": 1, "figures": 2, "docx": 3, "latex": 4, "submission": 5}
PHASE_SCOPE = {
    "problem_audit": "design", "model_design": "design", "solve_validate": "results",
    "figure_evidence": "figures", "writing_docx": "docx", "writing_latex": "latex",
    "ai_cleanup": "latex", "latex_compile_quality": "latex", "review_delivery": "submission",
    "completed": "submission",
}
HASH_KEYS = ("data", "model", "solution_workbook", "robustness_workbook", "matlab_script", "figure_bundle", "framework")


def _load_shared_validator():
    path = SKILL_ROOT / "templates" / "code" / "hsk_pipeline" / "workbook_validation.py"
    spec = importlib.util.spec_from_file_location("hsk_workbook_validation", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


WORKBOOK_VALIDATION = _load_shared_validator()


@dataclass
class SheetInfo:
    headers: list[str] = field(default_factory=list)
    data_rows: int = 0
    max_column: int = 0


@dataclass
class WorkbookInfo:
    path: str
    sha256: str
    sheets: dict[str, SheetInfo] = field(default_factory=dict)
    issues: list[str] = field(default_factory=list)


@dataclass
class QuestionSnapshot:
    key: str
    chinese_name: str
    code_files: list[str] = field(default_factory=list)
    model_hash: str | None = None
    solution_workbook: WorkbookInfo | None = None
    robustness_workbook: WorkbookInfo | None = None
    matlab_script: str | None = None
    matlab_hash: str | None = None
    matlab_has_title: bool = False
    workbook_references: list[str] = field(default_factory=list)
    handoff_path: str | None = None
    handoff_payload: dict[str, Any] = field(default_factory=dict)
    exported_figures: list[str] = field(default_factory=list)
    discovered_figures: list[str] = field(default_factory=list)
    figure_hash: str | None = None
    figure_evidence_path: str | None = None
    figure_evidence: dict[str, Any] = field(default_factory=dict)
    framework_hash: str | None = None
    issues: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def load_yaml(path: Path) -> dict[str, Any]:
    if not path.is_file():
        return {}
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def load_json_or_yaml(path: Path) -> dict[str, Any]:
    if not path.is_file():
        return {}
    try:
        if path.suffix.lower() == ".json":
            return json.loads(path.read_text(encoding="utf-8")) or {}
        return load_yaml(path)
    except Exception:  # noqa: BLE001
        return {}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_text(text: str) -> str:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def combined_hash(paths: Iterable[Path], root: Path) -> str | None:
    files = sorted({path.resolve() for path in paths if path.is_file()}, key=lambda item: item.as_posix())
    if not files:
        return None
    digest = hashlib.sha256()
    for path in files:
        try:
            relative = path.relative_to(root).as_posix()
        except ValueError:
            relative = path.as_posix()
        digest.update(relative.encode("utf-8")); digest.update(b"\0")
        digest.update(bytes.fromhex(sha256_file(path)))
    return digest.hexdigest()


def question_key(chinese_name: str) -> str:
    order = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十"]
    suffix = chinese_name.removeprefix("问题")
    try:
        return f"Q{order.index(suffix) + 1}"
    except ValueError:
        return chinese_name


def chinese_question_name(key: str) -> str:
    order = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十"]
    match = re.fullmatch(r"Q(\d+)", key)
    if match and 1 <= int(match.group(1)) <= len(order):
        return f"问题{order[int(match.group(1)) - 1]}"
    return key


def data_source_files(root: Path, state: Mapping[str, Any]) -> tuple[list[Path], str, list[str], list[str]]:
    issues: list[str] = []
    warnings: list[str] = []
    entries = ((state.get("data") or {}).get("sources") or []) if state else []
    files: list[Path] = []
    if entries:
        for entry in entries:
            relative = str((entry or {}).get("path", "")).strip()
            if not relative:
                issues.append("data.sources 存在空路径")
                continue
            path = (root / relative).resolve()
            try:
                path.relative_to(root)
            except ValueError:
                issues.append(f"data.sources 路径越出项目根目录: {relative}")
                continue
            if path.is_file():
                files.append(path)
            elif path.is_dir():
                files.extend(item for item in path.rglob("*") if item.is_file())
            else:
                issues.append(f"data.sources 文件不存在: {relative}")
        return files, "declared_sources", issues, warnings
    for path in root.iterdir():
        if not path.is_file() or path.name.startswith("."):
            continue
        if path.name in {"模型论文框架.md", "sync_report.yaml"}:
            continue
        if path.suffix.lower() in DATA_SUFFIXES:
            files.append(path)
    warnings.append("项目状态未声明data.sources；data hash使用受限根目录数据文件回退扫描")
    return files, "fallback_scan", issues, warnings


def inspect_workbook(path: Path, root: Path) -> WorkbookInfo:
    info = WorkbookInfo(path=path.relative_to(root).as_posix(), sha256=sha256_file(path))
    try:
        book = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        info.issues.append(f"无法读取工作簿: {exc}")
        return info
    try:
        for sheet in book.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            first = next(iterator, None)
            headers = [str(value).strip() if value is not None else "" for value in (first or [])]
            while headers and not headers[-1]:
                headers.pop()
            rows = sum(1 for row in iterator if any(value is not None for value in row[: len(headers)]))
            info.sheets[sheet.title] = SheetInfo(headers, rows, sheet.max_column)
    finally:
        book.close()
    return info


def framework_section_text(path: Path, anchor: str) -> str | None:
    if not path.is_file() or not anchor.strip():
        return None
    lines = path.read_text(encoding="utf-8").replace("\r\n", "\n").replace("\r", "\n").splitlines()
    target = anchor.strip()
    start = next((index for index, line in enumerate(lines) if line.strip() == target), None)
    if start is None:
        start = next((index for index, line in enumerate(lines) if line.lstrip().startswith("#") and target in line.strip()), None)
    if start is None:
        return None
    heading = lines[start].lstrip()
    level = len(heading) - len(heading.lstrip("#"))
    end = len(lines)
    for index in range(start + 1, len(lines)):
        stripped = lines[index].lstrip()
        if not stripped.startswith("#"):
            continue
        next_level = len(stripped) - len(stripped.lstrip("#"))
        if next_level <= level:
            end = index
            break
    return "\n".join(lines[start:end]).strip() + "\n"


def framework_section_hash(path: Path, anchor: str) -> str | None:
    text = framework_section_text(path, anchor)
    return sha256_text(text) if text else None


def _collect_strings(value: Any) -> list[str]:
    values: list[str] = []
    if isinstance(value, str):
        values.append(value)
    elif isinstance(value, Mapping):
        for item in value.values():
            values.extend(_collect_strings(item))
    elif isinstance(value, list):
        for item in value:
            values.extend(_collect_strings(item))
    return values


def discover_questions(root: Path, state: Mapping[str, Any]) -> dict[str, QuestionSnapshot]:
    snapshots: dict[str, QuestionSnapshot] = {}
    for key in (state.get("subproblems", {}) or {}):
        snapshots[key] = QuestionSnapshot(key, chinese_question_name(key))
    result_root = root / "结果数据表"
    if result_root.is_dir():
        for directory in sorted(path for path in result_root.iterdir() if path.is_dir()):
            if QUESTION_RE.fullmatch(directory.name):
                key = question_key(directory.name)
                snapshots.setdefault(key, QuestionSnapshot(key, directory.name))
    for script in sorted(root.glob("问题*.py")):
        match = QUESTION_RE.match(script.stem)
        if match:
            chinese_name = match.group(0); key = question_key(chinese_name)
            snapshots.setdefault(key, QuestionSnapshot(key, chinese_name)).code_files.append(script.relative_to(root).as_posix())
    framework_path = root / "模型论文框架.md"
    subproblems = state.get("subproblems", {}) or {}
    for key, snapshot in snapshots.items():
        sub = subproblems.get(key, {}) or {}
        snapshot.model_hash = combined_hash([root / item for item in snapshot.code_files], root)
        result_dir = result_root / snapshot.chinese_name
        solution = result_dir / f"{snapshot.chinese_name}求解结果.xlsx"
        robustness = result_dir / f"{snapshot.chinese_name}敏感性与鲁棒性结果.xlsx"
        if solution.is_file():
            snapshot.solution_workbook = inspect_workbook(solution, root)
        if robustness.is_file():
            snapshot.robustness_workbook = inspect_workbook(robustness, root)
        number = re.sub(r"\D", "", snapshot.key)
        matlab = result_dir / f"q{number}_plot.m" if number else None
        if matlab and matlab.is_file():
            snapshot.matlab_script = matlab.relative_to(root).as_posix()
            snapshot.matlab_hash = sha256_file(matlab)
            text = matlab.read_text(encoding="utf-8", errors="replace")
            snapshot.matlab_has_title = bool(MATLAB_TITLE_RE.search(text))
            snapshot.workbook_references = sorted(set(WORKBOOK_REF_RE.findall(text)))
            snapshot.exported_figures = sorted(set(EXPORT_RE.findall(text)))
        for name in ("matlab_figure_handoff.json", "matlab_figure_handoff.yaml", "matlab_figure_handoff.yml"):
            candidate = result_dir / name
            if candidate.is_file():
                snapshot.handoff_path = candidate.relative_to(root).as_posix()
                snapshot.handoff_payload = load_json_or_yaml(candidate)
                handoff_refs = [item for item in _collect_strings(snapshot.handoff_payload) if item.lower().endswith(".xlsx")]
                snapshot.workbook_references = sorted(set([*snapshot.workbook_references, *handoff_refs]))
                break
        evidence = result_dir / "figure_evidence.yaml"
        if evidence.is_file():
            snapshot.figure_evidence_path = evidence.relative_to(root).as_posix()
            snapshot.figure_evidence = load_yaml(evidence)
        figure_dir = result_dir / "图表"
        if figure_dir.is_dir():
            figure_paths = [path for path in figure_dir.rglob("*") if path.is_file() and path.suffix.lower() in FIGURE_SUFFIXES]
            snapshot.discovered_figures = sorted(path.relative_to(root).as_posix() for path in figure_paths)
            snapshot.figure_hash = combined_hash(figure_paths, root)
        anchor = str(sub.get("framework_section", "")).strip()
        snapshot.framework_hash = framework_section_hash(framework_path, anchor)
        if anchor and snapshot.framework_hash is None:
            snapshot.issues.append(f"模型论文框架中未找到小问章节: {anchor}")
    return snapshots


def active_contract(subproblem: Mapping[str, Any]) -> tuple[str | None, tuple[str, ...], tuple[str, ...], Mapping[str, bool] | None]:
    classification = subproblem.get("classification") or {}
    objective = classification.get("objective")
    structures = tuple(classification.get("structures", []) or [])
    legacy = subproblem.get("problem_types") or {}
    labels = [legacy.get("primary"), *(legacy.get("secondary", []) or [])]
    problem_types = tuple(dict.fromkeys(str(item) for item in labels if item))
    capabilities = subproblem.get("capabilities")
    return objective, structures, problem_types, capabilities if isinstance(capabilities, Mapping) else None


def validate_workbooks(
    snapshot: QuestionSnapshot,
    subproblem: Mapping[str, Any],
    schema: Mapping[str, Any],
    *,
    require_solution: bool,
    require_robustness: bool,
    root: Path,
) -> list[str]:
    objective, structures, problem_types, capabilities = active_contract(subproblem)
    solution = root / snapshot.solution_workbook.path if snapshot.solution_workbook else None
    robustness = root / snapshot.robustness_workbook.path if snapshot.robustness_workbook else None
    return WORKBOOK_VALIDATION.validate_pair(
        solution, robustness, schema=schema,
        require_solution=require_solution, require_robustness=require_robustness,
        objective=objective, structures=structures, problem_types=problem_types,
        capabilities=capabilities,
    )


def infer_scope(state: Mapping[str, Any], explicit: str | None) -> str:
    if explicit:
        if explicit not in SCOPE_RANK:
            raise ValueError(f"未知交付范围: {explicit}")
        return explicit
    phase = str((state.get("project", {}) or {}).get("current_phase", "model_design"))
    return PHASE_SCOPE.get(phase, "design")


def artifact_hashes(snapshot: QuestionSnapshot, data_hash: str | None) -> dict[str, str]:
    values = {
        "data": data_hash,
        "model": snapshot.model_hash,
        "solution_workbook": snapshot.solution_workbook.sha256 if snapshot.solution_workbook else None,
        "robustness_workbook": snapshot.robustness_workbook.sha256 if snapshot.robustness_workbook else None,
        "matlab_script": snapshot.matlab_hash,
        "figure_bundle": snapshot.figure_hash,
        "framework": snapshot.framework_hash,
    }
    return {key: value for key, value in values.items() if value}


def stale_layers(current: Mapping[str, str], validated: Mapping[str, str]) -> list[str]:
    return [key for key in HASH_KEYS if validated.get(key) and current.get(key) != validated.get(key)]


def _figure_hash_map(snapshot: QuestionSnapshot, root: Path) -> dict[str, str]:
    return {relative: sha256_file(root / relative) for relative in snapshot.discovered_figures if (root / relative).is_file()}


def _figure_evidence_payload(snapshot: QuestionSnapshot, root: Path, timestamp: str) -> dict[str, Any]:
    return {
        "schema_version": "1.0.0",
        "generated_at": timestamp,
        "question": snapshot.key,
        "source_workbook_hashes": {
            "solution": snapshot.solution_workbook.sha256 if snapshot.solution_workbook else None,
            "robustness": snapshot.robustness_workbook.sha256 if snapshot.robustness_workbook else None,
        },
        "matlab_script": snapshot.matlab_script,
        "matlab_script_hash": snapshot.matlab_hash,
        "figure_bundle_hash": snapshot.figure_hash,
        "figures": _figure_hash_map(snapshot, root),
    }


def _figure_evidence_issues(snapshot: QuestionSnapshot, root: Path) -> list[str]:
    if not snapshot.figure_evidence:
        return []
    payload = snapshot.figure_evidence
    source = payload.get("source_workbook_hashes", {}) or {}
    expected_solution = snapshot.solution_workbook.sha256 if snapshot.solution_workbook else None
    expected_robustness = snapshot.robustness_workbook.sha256 if snapshot.robustness_workbook else None
    issues: list[str] = []
    if source.get("solution") != expected_solution:
        issues.append("figure_evidence记录的求解工作簿哈希与当前文件不一致")
    if source.get("robustness") != expected_robustness:
        issues.append("figure_evidence记录的鲁棒性工作簿哈希与当前文件不一致")
    if payload.get("matlab_script_hash") != snapshot.matlab_hash:
        issues.append("figure_evidence记录的MATLAB脚本哈希与当前文件不一致")
    if payload.get("figure_bundle_hash") != snapshot.figure_hash:
        issues.append("figure_evidence记录的图表包哈希与当前文件不一致")
    declared = payload.get("figures", {}) or {}
    actual = _figure_hash_map(snapshot, root)
    if declared != actual:
        issues.append("figure_evidence记录的逐图哈希与当前正式图不一致")
    return issues


def figure_chain_issues(snapshot: QuestionSnapshot, root: Path) -> tuple[list[str], list[str]]:
    issues: list[str] = []
    warnings: list[str] = []
    if not snapshot.matlab_script:
        return ["图表交付缺少MATLAB脚本"], warnings
    if not snapshot.matlab_has_title:
        issues.append("MATLAB脚本缺少title或sgtitle")
    expected_books = {
        Path(snapshot.solution_workbook.path).name if snapshot.solution_workbook else "",
        Path(snapshot.robustness_workbook.path).name if snapshot.robustness_workbook else "",
    } - {""}
    referenced = {Path(item).name for item in snapshot.workbook_references}
    if not referenced:
        issues.append("MATLAB脚本与handoff均未发现标准工作簿引用")
    elif not referenced.issubset(expected_books):
        issues.append(f"MATLAB引用了非本问标准工作簿: {sorted(referenced - expected_books)}")
    result_dir = root / "结果数据表" / snapshot.chinese_name
    discovered_names = {Path(item).name for item in snapshot.discovered_figures}
    for declared in snapshot.exported_figures:
        name = Path(declared).name
        if name not in discovered_names and not (result_dir / declared).is_file():
            issues.append(f"MATLAB声明导出的图不存在: {declared}")
    if not snapshot.discovered_figures:
        issues.append("图表交付未发现正式图文件")
    evidence_issues = _figure_evidence_issues(snapshot, root)
    issues.extend(evidence_issues)
    if not snapshot.figure_evidence:
        warnings.append("缺少figure_evidence.yaml；--write将在图表链无错误时生成哈希证据")
        source_paths = [root / snapshot.matlab_script]
        for book in (snapshot.solution_workbook, snapshot.robustness_workbook):
            if book:
                source_paths.append(root / book.path)
        newest_source = max((path.stat().st_mtime for path in source_paths if path.is_file()), default=0)
        for relative in snapshot.discovered_figures:
            if (root / relative).stat().st_mtime < newest_source:
                warnings.append(f"mtime辅助检查：正式图早于工作簿或MATLAB脚本: {relative}")
    return issues, warnings


def update_framework_header(path: Path, *, stale: bool, scope: str, timestamp: str) -> bool:
    if not path.is_file():
        return False
    text = path.read_text(encoding="utf-8")
    replacements = {
        r"(?m)^- 最近同步：.*$": f"- 最近同步：`{scope}`",
        r"(?m)^- 最近同步时间：.*$": f"- 最近同步时间：`{timestamp}`",
        r"(?m)^- 当前状态：.*$": f"- 当前状态：`{'stale' if stale else 'current'}`",
    }
    changed = False
    for pattern, replacement in replacements.items():
        text, count = re.subn(pattern, replacement, text)
        changed = changed or bool(count)
    if changed:
        path.write_text(text, encoding="utf-8")
    return changed


def _existing_files(root: Path, values: Iterable[str]) -> list[str]:
    return [value for value in values if isinstance(value, str) and (root / value).is_file()]


def _approved_figure_issues(root: Path, state: Mapping[str, Any]) -> list[str]:
    artifacts = state.get("artifacts", {}) or {}
    approved = artifacts.get("approved_figures", []) or []
    if not approved:
        return ["正式写作交付缺少artifacts.approved_figures人工批准清单"]
    missing = [value for value in approved if not (root / str(value)).is_file()]
    return [f"批准图文件不存在: {value}" for value in missing]


def _docx_artifacts(root: Path) -> list[Path]:
    directory = root / "draft_docx"
    return sorted(path for path in directory.rglob("*.docx") if path.is_file()) if directory.is_dir() else []


def _latex_graphics_issues(main_tex: Path) -> list[str]:
    if not main_tex.is_file():
        return []
    issues: list[str] = []
    text = main_tex.read_text(encoding="utf-8", errors="replace")
    for raw in LATEX_GRAPHICS_RE.findall(text):
        candidate = (main_tex.parent / raw).resolve()
        choices = [candidate] if candidate.suffix else [candidate.with_suffix(ext) for ext in (".pdf", ".png", ".jpg", ".jpeg", ".svg")]
        if not any(path.is_file() for path in choices):
            issues.append(f"LaTeX引用图片不存在: {raw}")
    return issues


def _compile_report_issues(path: Path) -> list[str]:
    if not path.is_file():
        return ["LaTeX交付缺少final_latex/compile_report.yaml"]
    payload = load_yaml(path)
    issues: list[str] = []
    status = str(payload.get("status", payload.get("conclusion", ""))).lower()
    if status not in {"passed", "pass", "success", "completed"}:
        issues.append(f"compile_report状态不是通过: {status or '<empty>'}")
    unresolved = payload.get("unresolved_references")
    if unresolved not in (None, 0, "0", [], {}):
        issues.append("compile_report仍包含未解决引用")
    return issues


def _submission_candidates(root: Path) -> list[Path]:
    candidates: list[Path] = []
    directory = root / "submission"
    if directory.is_dir():
        candidates.extend(path for path in directory.rglob("*.zip") if path.is_file())
    candidates.extend(path for path in root.glob("*.zip") if any(token in path.stem.lower() for token in ("submission", "submit", "提交")))
    return sorted(set(candidates))


def _submission_zip_issues(path: Path, *, require_matlab: bool) -> list[str]:
    try:
        with zipfile.ZipFile(path) as archive:
            names = [name for name in archive.namelist() if not name.endswith("/")]
    except Exception as exc:  # noqa: BLE001
        return [f"无法读取提交ZIP: {exc}"]
    lowered = [name.lower() for name in names]
    requirements = {
        "最终论文PDF": any(name.endswith(".pdf") for name in lowered),
        "Python代码": any(name.endswith(".py") for name in lowered),
        "结果工作簿": any(name.endswith(".xlsx") for name in lowered),
        "复现说明": any(("readme" in name or "复现" in name or "运行说明" in name) and name.endswith((".md", ".txt")) for name in lowered),
    }
    if require_matlab:
        requirements["MATLAB绘图代码"] = any(name.endswith(".m") for name in lowered)
    return [f"提交ZIP缺少{label}" for label, present in requirements.items() if not present]


def delivery_artifact_issues(root: Path, state: Mapping[str, Any], scope: str, snapshots: Mapping[str, QuestionSnapshot]) -> tuple[list[str], dict[str, Any]]:
    issues: list[str] = []
    discovered: dict[str, Any] = {}
    framework = root / "模型论文框架.md"
    state_path = root / "state" / "project_state.yaml"
    if not state_path.is_file():
        issues.append("交付缺少state/project_state.yaml")
    if not framework.is_file():
        issues.append("交付缺少模型论文框架.md")
    if scope == "docx":
        issues.extend(_approved_figure_issues(root, state))
        docx_files = _docx_artifacts(root)
        discovered["docx"] = [path.relative_to(root).as_posix() for path in docx_files]
        if not docx_files:
            issues.append("DOCX交付缺少draft_docx/*.docx")
    if SCOPE_RANK[scope] >= SCOPE_RANK["latex"]:
        issues.extend(_approved_figure_issues(root, state))
        main_tex = root / "final_latex" / "main.tex"
        main_pdf = root / "final_latex" / "main.pdf"
        compile_report = root / "final_latex" / "compile_report.yaml"
        discovered.update({
            "latex_source": main_tex.relative_to(root).as_posix() if main_tex.is_file() else None,
            "compiled_pdf": main_pdf.relative_to(root).as_posix() if main_pdf.is_file() else None,
            "compile_report": compile_report.relative_to(root).as_posix() if compile_report.is_file() else None,
        })
        if not main_tex.is_file():
            issues.append("LaTeX交付缺少final_latex/main.tex")
        if not main_pdf.is_file():
            issues.append("LaTeX交付缺少final_latex/main.pdf")
        issues.extend(_compile_report_issues(compile_report))
        issues.extend(_latex_graphics_issues(main_tex))
    if SCOPE_RANK[scope] >= SCOPE_RANK["submission"]:
        candidates = _submission_candidates(root)
        if not candidates:
            issues.append("提交交付缺少submission目录或命名明确的提交ZIP")
        else:
            package = candidates[0]
            discovered["submission_package"] = package.relative_to(root).as_posix()
            require_matlab = any(snapshot.matlab_script for snapshot in snapshots.values())
            issues.extend(_submission_zip_issues(package, require_matlab=require_matlab))
    return issues, discovered


def snapshot_to_dict(snapshot: QuestionSnapshot, current_hashes: Mapping[str, str]) -> dict[str, Any]:
    def workbook_payload(info: WorkbookInfo | None) -> dict[str, Any] | None:
        if info is None:
            return None
        return {
            "path": info.path, "sha256": info.sha256,
            "sheets": {name: {"headers": sheet.headers, "data_rows": sheet.data_rows, "max_column": sheet.max_column} for name, sheet in info.sheets.items()},
            "issues": info.issues,
        }
    return {
        "question": snapshot.chinese_name,
        "code_files": snapshot.code_files,
        "artifact_hashes": dict(current_hashes),
        "solution_workbook": workbook_payload(snapshot.solution_workbook),
        "robustness_workbook": workbook_payload(snapshot.robustness_workbook),
        "matlab_script": snapshot.matlab_script,
        "matlab_has_title": snapshot.matlab_has_title,
        "workbook_references": snapshot.workbook_references,
        "handoff_path": snapshot.handoff_path,
        "exported_figures": snapshot.exported_figures,
        "discovered_figures": snapshot.discovered_figures,
        "figure_evidence": snapshot.figure_evidence_path,
        "issues": snapshot.issues,
        "warnings": snapshot.warnings,
    }


def synchronize(
    root: Path,
    *,
    write: bool,
    question: str | None = None,
    delivery_scope: str | None = None,
    schema_path: Path = DEFAULT_SCHEMA_PATH,
) -> dict[str, Any]:
    root = root.resolve()
    state_path = root / "state" / "project_state.yaml"
    state = load_yaml(state_path)
    schema = load_yaml(schema_path)
    scope = infer_scope(state, delivery_scope)
    timestamp = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    data_files, data_hash_mode, data_issues, data_warnings = data_source_files(root, state)
    data_hash = combined_hash(data_files, root)
    snapshots = discover_questions(root, state)
    if question:
        normalized = question if question.startswith("Q") else question_key(question)
        snapshots = {key: value for key, value in snapshots.items() if key == normalized}
        if not snapshots:
            raise ValueError(f"未发现小问: {question}")
    issues: list[str] = list(data_issues)
    warnings: list[str] = list(data_warnings)
    stale_questions: list[str] = []
    subproblems = state.get("subproblems", {}) if state else {}
    snapshots_payload: dict[str, Any] = {}

    global_issues, discovered_delivery = delivery_artifact_issues(root, state, scope, snapshots)
    issues.extend(global_issues)

    for key, snapshot in snapshots.items():
        sub = subproblems.get(key)
        if not isinstance(sub, dict):
            issues.append(f"{key}: 项目状态中缺少对应小问")
            continue
        status_rank = STATUS_RANK.get(str(sub.get("status", "pending")), 0)
        require_results = status_rank >= STATUS_RANK["solved"] or SCOPE_RANK[scope] >= SCOPE_RANK["results"]
        require_figures = SCOPE_RANK[scope] >= SCOPE_RANK["figures"] and status_rank >= STATUS_RANK["solved"]
        contract_issues = validate_workbooks(
            snapshot, sub, schema, require_solution=require_results,
            require_robustness=require_results, root=root,
        )
        if require_results and not snapshot.code_files:
            contract_issues.append("缺少问题求解Python脚本")
        if require_figures:
            figure_issues, figure_warnings = figure_chain_issues(snapshot, root)
            contract_issues.extend(figure_issues)
            snapshot.warnings.extend(figure_warnings)
        snapshot.issues.extend(contract_issues)
        issues.extend(f"{key}: {item}" for item in snapshot.issues)
        warnings.extend(f"{key}: {item}" for item in snapshot.warnings)

        current = artifact_hashes(snapshot, data_hash)
        validated = dict(sub.get("validated_artifact_hashes", {}) or {})
        if not validated:
            if sub.get("validated_data_hash"):
                validated["data"] = sub["validated_data_hash"]
            if sub.get("validated_model_hash"):
                validated["model"] = sub["validated_model_hash"]
        changed = stale_layers(current, validated)
        sub["artifact_hashes"] = current
        if changed:
            stale_questions.append(key)
            sub["artifacts_stale"] = True
            sub["stale_layers"] = changed
            sub["result_summary_status"] = "stale"
            if sub.get("status") in {"validated", "written", "completed"}:
                sub["status"] = "solved"
            if sub.get("validation_status") == "passed":
                sub["validation_status"] = "pending"
        elif sub.get("artifacts_stale") is True and not snapshot.issues:
            sub["artifacts_stale"] = False
            sub["stale_layers"] = []
        if data_hash:
            sub["data_hash"] = data_hash
        if snapshot.model_hash:
            sub["model_hash"] = snapshot.model_hash
        if snapshot.code_files:
            sub["code"] = snapshot.code_files[0]
        if snapshot.solution_workbook:
            sub["solution_workbook"] = snapshot.solution_workbook.path
        if snapshot.robustness_workbook:
            sub["robustness_workbook"] = snapshot.robustness_workbook.path
        if snapshot.matlab_script:
            sub["matlab_script"] = snapshot.matlab_script
        evidence = set(sub.get("evidence", []))
        for value in [
            *snapshot.code_files,
            snapshot.solution_workbook.path if snapshot.solution_workbook else None,
            snapshot.robustness_workbook.path if snapshot.robustness_workbook else None,
            snapshot.matlab_script,
            snapshot.handoff_path,
            snapshot.figure_evidence_path,
            *snapshot.discovered_figures,
        ]:
            if value:
                evidence.add(value)
        sub["evidence"] = sorted(evidence)
        snapshots_payload[key] = snapshot_to_dict(snapshot, current)

    stale = bool(stale_questions)
    if write:
        update_framework_header(root / "模型论文框架.md", stale=stale, scope=question or f"{scope}交付同步", timestamp=timestamp)
        if SCOPE_RANK[scope] >= SCOPE_RANK["figures"]:
            for snapshot in snapshots.values():
                if snapshot.matlab_script and snapshot.discovered_figures and not snapshot.issues:
                    result_dir = root / "结果数据表" / snapshot.chinese_name
                    evidence_path = result_dir / "figure_evidence.yaml"
                    evidence_path.write_text(
                        yaml.safe_dump(_figure_evidence_payload(snapshot, root, timestamp), allow_unicode=True, sort_keys=False),
                        encoding="utf-8",
                    )
                    snapshot.figure_evidence_path = evidence_path.relative_to(root).as_posix()
                    snapshot.figure_evidence = load_yaml(evidence_path)
                    snapshots_payload[snapshot.key] = snapshot_to_dict(
                        snapshot, artifact_hashes(snapshot, data_hash)
                    )
    framework_hash = sha256_file(root / "模型论文框架.md") if (root / "模型论文框架.md").is_file() else None

    if state:
        execution = state.setdefault("execution", {})
        execution["last_run"] = timestamp
        execution["command"] = f"python scripts/sync_project.py --write --delivery-scope {scope}"
        execution["last_sync_report"] = "sync_report.yaml"
        artifacts = state.setdefault("artifacts", {})
        artifacts["code"] = sorted({path for item in snapshots.values() for path in item.code_files})
        artifacts["results"] = sorted({path for item in snapshots.values() for path in [item.solution_workbook.path if item.solution_workbook else None, item.robustness_workbook.path if item.robustness_workbook else None] if path})
        artifacts["figures"] = sorted({path for item in snapshots.values() for path in item.discovered_figures})
        artifacts["sync_report"] = "sync_report.yaml"
        if discovered_delivery.get("docx"):
            artifacts["docx"] = discovered_delivery["docx"]
        for key in ("latex_source", "compiled_pdf", "compile_report", "submission_package"):
            if discovered_delivery.get(key):
                artifacts[key] = discovered_delivery[key]
        framework = state.setdefault("paper_framework", {})
        framework["path"] = "模型论文框架.md"
        framework["sync_status"] = "stale" if stale else "current"
        framework["last_synced_at"] = timestamp
        framework["last_sync_scope"] = question or scope
        if framework_hash:
            framework["sha256"] = framework_hash

    report = {
        "sync_version": "1.2.0",
        "skill_version": "6.3.2",
        "generated_at": timestamp,
        "project_root": root.as_posix(),
        "delivery_scope": scope,
        "write_requested": write,
        "data_hash": data_hash,
        "data_hash_mode": data_hash_mode,
        "data_sources": [path.relative_to(root).as_posix() for path in data_files],
        "framework_hash": framework_hash,
        "stale_questions": stale_questions,
        "questions": snapshots_payload,
        "delivery_artifacts": discovered_delivery,
        "issues": issues,
        "warnings": warnings,
        "policy": {
            "promotes_validation": False,
            "approves_figures": False,
            "rewrites_model_semantics": False,
            "stale_propagation": True,
        },
    }
    if write:
        if state:
            state_path.parent.mkdir(parents=True, exist_ok=True)
            state_path.write_text(yaml.safe_dump(state, allow_unicode=True, sort_keys=False), encoding="utf-8")
        (root / "sync_report.yaml").write_text(yaml.safe_dump(report, allow_unicode=True, sort_keys=False), encoding="utf-8")
        if state and framework_hash:
            persisted = load_yaml(state_path)
            if (persisted.get("paper_framework", {}) or {}).get("sha256") != framework_hash:
                raise RuntimeError("同步后框架哈希写入不一致")
    return report


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("project_root", nargs="?", default=".")
    parser.add_argument("--question", help="Q1或问题一")
    parser.add_argument("--delivery-scope", choices=sorted(SCOPE_RANK, key=SCOPE_RANK.get))
    parser.add_argument("--schema", default=str(DEFAULT_SCHEMA_PATH))
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--strict", action="store_true")
    args = parser.parse_args()
    try:
        report = synchronize(
            Path(args.project_root), write=args.write, question=args.question,
            delivery_scope=args.delivery_scope, schema_path=Path(args.schema),
        )
    except (ValueError, OSError, RuntimeError) as exc:
        raise SystemExit(str(exc)) from exc
    if not args.write:
        print(yaml.safe_dump(report, allow_unicode=True, sort_keys=False))
    if args.strict and (report["issues"] or report["stale_questions"]):
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
