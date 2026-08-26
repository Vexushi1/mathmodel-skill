#!/usr/bin/env python3
"""Independently recheck intrinsic primary-solve numerical evidence.

This validator never runs task-specific model code. It reads a returned primary
workbook, applies the capability-driven contract in
``core/numerical_verification_contract.yaml``, and reports contradictions between
bottom-level evidence and the main quality summary.
"""
from __future__ import annotations

import argparse
import math
from pathlib import Path
from typing import Any, Mapping

import openpyxl
import yaml


def _as_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "是", "通过", "满足"}:
        return True
    if text in {"false", "0", "no", "否", "未通过", "不满足"}:
        return False
    return None


def _as_number(value: Any) -> float | None:
    if isinstance(value, bool) or value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _contract_candidates(explicit: Path | None) -> list[Path]:
    candidates: list[Path] = []
    if explicit is not None:
        candidates.append(Path(explicit))
    here = Path(__file__).resolve()
    candidates.extend(parent / "core" / "numerical_verification_contract.yaml" for parent in here.parents)
    return candidates


def load_contract(path: Path | None = None) -> dict[str, Any]:
    for candidate in _contract_candidates(path):
        if candidate.is_file():
            return yaml.safe_load(candidate.read_text(encoding="utf-8")) or {}
    raise FileNotFoundError("找不到core/numerical_verification_contract.yaml")


def _sheet_records(book: openpyxl.Workbook, sheet: str) -> tuple[list[str], list[dict[str, Any]]]:
    if sheet not in book.sheetnames:
        return [], []
    rows = list(book[sheet].iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = ["" if item is None else str(item).strip() for item in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(value not in (None, "") for value in row[: len(headers)]):
            continue
        records.append({headers[index]: row[index] if index < len(row) else None for index in range(len(headers))})
    return headers, records


def _numeric_equal(left: float, right: float, contract: Mapping[str, Any]) -> bool:
    tolerance = (contract.get("numeric_consistency") or {}).get("comparison_tolerance") or {}
    absolute = float(tolerance.get("absolute", 1.0e-12))
    relative = float(tolerance.get("relative", 1.0e-9))
    return abs(left - right) <= max(absolute, relative * max(abs(left), abs(right), 1.0))


def _relation_result(relation: str, actual: Any, threshold: Any) -> tuple[bool | None, str | None]:
    if relation == "bool_true":
        parsed = _as_bool(actual)
        return parsed is True, None if parsed is not None else "实际值无法解析为布尔值"

    actual_number = _as_number(actual)
    threshold_number = _as_number(threshold)
    if actual_number is None or threshold_number is None:
        return None, "实际值/阈值必须为有限数值"
    if relation == "<=":
        return actual_number <= threshold_number, None
    if relation == ">=":
        return actual_number >= threshold_number, None
    if relation == "abs<=":
        return abs(actual_number) <= abs(threshold_number), None
    if relation == "==":
        return actual_number == threshold_number, None
    return None, f"未知判定关系: {relation}"


def _require_columns(sheet: str, headers: list[str], columns: list[str]) -> list[str]:
    return [f"{sheet}缺少字段: {column}" for column in columns if column not in headers]


def _row_residual_metric(
    sheet: str,
    headers: list[str],
    records: list[dict[str, Any]],
    value_column: str,
) -> tuple[float | None, list[str]]:
    issues = _require_columns(sheet, headers, [value_column, "容差", "是否满足"])
    if issues:
        return None, issues
    if not records:
        return None, [f"{sheet}没有实质数据"]
    values: list[float] = []
    for index, record in enumerate(records, start=2):
        value = _as_number(record.get(value_column))
        tolerance = _as_number(record.get("容差"))
        actual = _as_bool(record.get("是否满足"))
        if value is None or tolerance is None:
            issues.append(f"{sheet}第{index}行{value_column}/容差必须为有限数值")
            continue
        expected = abs(value) <= tolerance
        if actual is None or actual != expected:
            issues.append(f"{sheet}第{index}行是否满足与{value_column}/容差不一致")
        values.append(abs(value))
    return (max(values) if values else None), issues


def _marked_relative_metric(
    sheet: str,
    headers: list[str],
    records: list[dict[str, Any]],
) -> tuple[float | None, list[str]]:
    issues = _require_columns(sheet, headers, ["相对变化", "用于主判定"])
    if issues:
        return None, issues
    marked = [record for record in records if _as_bool(record.get("用于主判定")) is True]
    if not marked:
        return None, [f"{sheet}至少需要一行用于主判定=true的离散精度证据"]
    values: list[float] = []
    for record in marked:
        value = _as_number(record.get("相对变化"))
        if value is None:
            issues.append(f"{sheet}用于主判定行的相对变化必须为有限数值")
        else:
            values.append(abs(value))
    return (max(values) if values else None), issues


def _marked_boolean_metric(
    sheet: str,
    headers: list[str],
    records: list[dict[str, Any]],
) -> tuple[bool | None, list[str]]:
    issues = _require_columns(sheet, headers, ["判定", "用于主判定"])
    if issues:
        return None, issues
    marked = [record for record in records if _as_bool(record.get("用于主判定")) is True]
    if not marked:
        return None, [f"{sheet}至少需要一行用于主判定=true的收敛证据"]
    parsed = [_as_bool(record.get("判定")) for record in marked]
    if any(value is None for value in parsed):
        issues.append(f"{sheet}用于主判定行的判定必须为可解析布尔值")
        return None, issues
    return all(value is True for value in parsed), issues


def _all_boolean_metric(
    sheet: str,
    headers: list[str],
    records: list[dict[str, Any]],
) -> tuple[bool | None, list[str]]:
    issues = _require_columns(sheet, headers, ["是否通过"])
    if issues:
        return None, issues
    if not records:
        return None, [f"{sheet}没有实质数据"]
    parsed = [_as_bool(record.get("是否通过")) for record in records]
    if any(value is None for value in parsed):
        issues.append(f"{sheet}的是否通过必须为可解析布尔值")
        return None, issues
    return all(value is True for value in parsed), issues


def _recheck_evidence(
    mode: str,
    sheet: str,
    headers: list[str],
    records: list[dict[str, Any]],
) -> tuple[Any, list[str]]:
    if mode == "max_violation":
        return _row_residual_metric(sheet, headers, records, "违反量")
    if mode == "max_abs_residual":
        return _row_residual_metric(sheet, headers, records, "残差")
    if mode == "marked_relative_change":
        return _marked_relative_metric(sheet, headers, records)
    if mode == "marked_boolean":
        return _marked_boolean_metric(sheet, headers, records)
    if mode == "all_boolean":
        return _all_boolean_metric(sheet, headers, records)
    if mode == "summary_criterion":
        return None, []
    return None, [f"未实现的numerical verification recheck_mode: {mode}"]


def validate_primary_numerical_evidence(
    workbook: Path,
    capabilities: Mapping[str, bool] | None = None,
    *,
    contract_path: Path | None = None,
    force_strict: bool | None = None,
) -> tuple[bool, list[str], dict[str, Any]]:
    """Validate one primary workbook without executing task-specific code."""
    contract = load_contract(contract_path)
    workbook = Path(workbook)
    book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    issues: list[str] = []
    try:
        quality_headers, quality_rows = _sheet_records(book, "主结果质量门")
        if not quality_headers:
            return False, ["缺少主结果质量门工作表"], {"mode": "invalid", "strict": False}

        marker_present = "Verification ID" in quality_headers and any(
            str(row.get("Verification ID", "")).strip() for row in quality_rows
        )
        strict = marker_present if force_strict is None else bool(force_strict)

        protocol_map = contract.get("capability_protocols") or {}
        active_capabilities = sorted(
            capability for capability, enabled in (capabilities or {}).items()
            if enabled and capability in protocol_map
        )

        # Legacy v7.13 read compatibility: preserve old Boolean-gate admission while
        # still detecting contradictions in residual sheets that already exist.
        if not strict:
            for sheet, value_column in (
                ("约束违反检查", "违反量"),
                ("均衡残差", "残差"),
                ("守恒残差", "残差"),
            ):
                headers, records = _sheet_records(book, sheet)
                if headers:
                    _, row_issues = _row_residual_metric(sheet, headers, records, value_column)
                    issues.extend(row_issues)
            report = {
                "mode": "legacy_read",
                "strict": False,
                "active_capabilities": active_capabilities,
                "checked_verification_ids": [],
            }
            return not issues, list(dict.fromkeys(issues)), report

        strict_contract = contract.get("strict_v714_trace") or {}
        required_columns = [str(item) for item in strict_contract.get("required_columns_when_active", [])]
        issues.extend(_require_columns("主结果质量门", quality_headers, required_columns))
        if issues:
            return False, list(dict.fromkeys(issues)), {
                "mode": "v7.14_strict", "strict": True, "active_capabilities": active_capabilities,
            }

        allowed_relations = {str(item) for item in strict_contract.get("allowed_relations", [])}
        seen_ids: set[str] = set()
        referenced_sheets: set[str] = set()
        computed_by_sheet: dict[str, Any] = {}
        modes_by_sheet: dict[str, set[str]] = {}

        for capability in active_capabilities:
            spec = protocol_map[capability] or {}
            sheet = str(spec.get("evidence_sheet", "")).strip()
            mode = str(spec.get("recheck_mode", "")).strip()
            if sheet:
                modes_by_sheet.setdefault(sheet, set()).add(mode)

        # Recompute bottom-level evidence once per sheet/mode.
        for sheet, modes in modes_by_sheet.items():
            headers, records = _sheet_records(book, sheet)
            if not headers:
                issues.append(f"启用的primary capability需要工作表但不存在: {sheet}")
                continue
            if not records:
                issues.append(f"启用的primary capability证据工作表无实质数据: {sheet}")
                continue
            if len(modes) > 1:
                issues.append(f"同一证据工作表映射到多个不兼容recheck_mode: {sheet} -> {sorted(modes)}")
                continue
            mode = next(iter(modes))
            computed, evidence_issues = _recheck_evidence(mode, sheet, headers, records)
            computed_by_sheet[sheet] = computed
            issues.extend(evidence_issues)

        checked_ids: list[str] = []
        for row_number, row in enumerate(quality_rows, start=2):
            verification_id = str(row.get("Verification ID", "")).strip()
            if not verification_id:
                issues.append(f"主结果质量门第{row_number}行缺少Verification ID")
                continue
            if not verification_id.startswith("PQ-"):
                issues.append(f"主结果质量门第{row_number}行Verification ID必须以PQ-开头")
            if verification_id in seen_ids:
                issues.append(f"主结果质量门Verification ID重复: {verification_id}")
            seen_ids.add(verification_id)
            checked_ids.append(verification_id)

            evidence_sheet = str(row.get("证据工作表", "")).strip()
            if not evidence_sheet:
                issues.append(f"{verification_id}缺少证据工作表")
            elif evidence_sheet not in book.sheetnames:
                issues.append(f"{verification_id}引用不存在的证据工作表: {evidence_sheet}")
            else:
                _, evidence_records = _sheet_records(book, evidence_sheet)
                if not evidence_records:
                    issues.append(f"{verification_id}引用的证据工作表无实质数据: {evidence_sheet}")
                referenced_sheets.add(evidence_sheet)

            if not str(row.get("阈值来源", "")).strip():
                issues.append(f"{verification_id}缺少阈值来源")

            relation = str(row.get("判定关系", "")).strip()
            if relation not in allowed_relations:
                issues.append(f"{verification_id}使用未登记判定关系: {relation}")
                continue
            expected_pass, relation_issue = _relation_result(
                relation, row.get("实际值"), row.get("阈值或容差")
            )
            if relation_issue:
                issues.append(f"{verification_id}: {relation_issue}")
                continue
            declared_pass = _as_bool(row.get("是否通过"))
            if declared_pass is None or declared_pass != expected_pass:
                issues.append(f"{verification_id}的是否通过与实际值/阈值/判定关系不一致")

            if evidence_sheet in computed_by_sheet and computed_by_sheet[evidence_sheet] is not None:
                computed = computed_by_sheet[evidence_sheet]
                if isinstance(computed, bool):
                    actual_bool = _as_bool(row.get("实际值"))
                    if actual_bool is None or actual_bool != computed:
                        issues.append(f"{verification_id}实际值与{evidence_sheet}重算布尔结果不一致")
                else:
                    actual_number = _as_number(row.get("实际值"))
                    if actual_number is None or not _numeric_equal(actual_number, float(computed), contract):
                        issues.append(f"{verification_id}实际值与{evidence_sheet}重算指标不一致")

        required_sheets = {
            str((protocol_map[capability] or {}).get("evidence_sheet", "")).strip()
            for capability in active_capabilities
            if str((protocol_map[capability] or {}).get("evidence_sheet", "")).strip()
        }
        for sheet in sorted(required_sheets - referenced_sheets):
            issues.append(f"主结果质量门未追溯启用capability的证据工作表: {sheet}")

        report = {
            "mode": "v7.14_strict",
            "strict": True,
            "active_capabilities": active_capabilities,
            "required_evidence_sheets": sorted(required_sheets),
            "referenced_evidence_sheets": sorted(referenced_sheets),
            "checked_verification_ids": checked_ids,
            "task_code_executed": False,
        }
        return not issues, list(dict.fromkeys(issues)), report
    finally:
        book.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--contract", type=Path)
    parser.add_argument("--capability", action="append", default=[])
    parser.add_argument("--strict", action="store_true")
    args = parser.parse_args()
    capabilities = {name: True for name in args.capability}
    passed, issues, report = validate_primary_numerical_evidence(
        args.workbook,
        capabilities,
        contract_path=args.contract,
        force_strict=True if args.strict else None,
    )
    report["status"] = "passed" if passed else "failed"
    report["issues"] = issues
    print(yaml.safe_dump(report, allow_unicode=True, sort_keys=False).rstrip())
    return 0 if passed else 1


if __name__ == "__main__":
    raise SystemExit(main())
