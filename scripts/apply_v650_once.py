#!/usr/bin/env python3
"""Apply the coordinated v6.5.0 user-executed full-fidelity workflow migration once."""
from __future__ import annotations

import re
import textwrap
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OLD = "6.4.1"
NEW = "6.5.0"
WORKFLOW = ROOT / ".github" / "workflows" / "refresh-generated.yml"
ORIGINAL_WORKFLOW = '''name: Refresh generated repository metadata

on:
  push:
    branches:
      - main
      - "codex/**"
      - "fix/**"
      - "refactor/**"
      - "upgrade/**"
    paths-ignore:
      - SKILL_FILE_INDEX.md
      - TEMPLATE_INDEX.md
      - HSK_SKILL_FILE_INDEX_V622.md
      - HSK_TEMPLATE_INDEX_V622.md
      - MANIFEST.sha256
  workflow_dispatch:

permissions:
  contents: write

concurrency:
  group: refresh-generated-${{ github.ref }}
  cancel-in-progress: true

jobs:
  refresh:
    if: github.actor != 'github-actions[bot]'
    runs-on: ubuntu-latest
    steps:
      - uses: actions/checkout@v4
        with:
          ref: ${{ github.ref_name }}
      - uses: actions/setup-python@v5
        with:
          python-version: "3.12"
      - name: Rebuild active indexes and manifest
        run: python scripts/generate_indexes.py
      - name: Commit generated metadata when changed
        run: |
          if git diff --quiet -- SKILL_FILE_INDEX.md TEMPLATE_INDEX.md HSK_SKILL_FILE_INDEX_V622.md HSK_TEMPLATE_INDEX_V622.md MANIFEST.sha256; then
            echo "Generated metadata is current."
            exit 0
          fi
          git config user.name "github-actions[bot]"
          git config user.email "41898282+github-actions[bot]@users.noreply.github.com"
          git add SKILL_FILE_INDEX.md TEMPLATE_INDEX.md HSK_SKILL_FILE_INDEX_V622.md HSK_TEMPLATE_INDEX_V622.md MANIFEST.sha256
          git commit -m "chore: refresh generated repository metadata"
          git push
'''


def dedent(text: str) -> str:
    return textwrap.dedent(text).lstrip("\n")


def write(relative: str, content: str) -> None:
    path = ROOT / relative
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(dedent(content), encoding="utf-8", newline="\n")


def replace_once(relative: str, old: str, new: str) -> None:
    path = ROOT / relative
    text = path.read_text(encoding="utf-8")
    if old not in text:
        raise RuntimeError(f"missing replacement anchor in {relative}: {old[:100]!r}")
    path.write_text(text.replace(old, new, 1), encoding="utf-8", newline="\n")


def regex_replace(relative: str, pattern: str, replacement: str) -> None:
    path = ROOT / relative
    text = path.read_text(encoding="utf-8")
    updated, count = re.subn(pattern, replacement, text, count=1, flags=re.S)
    if count != 1:
        raise RuntimeError(f"regex replacement failed in {relative}: {pattern}")
    path.write_text(updated, encoding="utf-8", newline="\n")


def create_contracts_and_tools() -> None:
    write("core/user_execution_contract.yaml", r'''
        version: 1.0.0
        skill_version: 6.5.0
        purpose: Define the user-executed full-fidelity numerical workflow without allowing the assistant to run task-specific solve or result-analysis programs.
        default_mode:
          execution_owner: user
          execution_profile: full_fidelity
          assistant_runs_task_specific_code: false
          local_pipeline_remains_runnable_by_user: true
        assistant_policy:
          prohibited:
          - 运行、导入或通过subprocess/runpy/notebook执行问题X求解.py
          - 运行、导入或通过subprocess/runpy/notebook执行问题X结果深化分析.py
          - 为节省时间自动缩减数据、网格、时域、场景、重复次数、随机种子、迭代次数或容差
          - 在完整版失败后静默切换轻量模型、替代求解器、粗粒度近似或演示结果
          - 用户未返回工作簿前声称已求解、已分析或已得到正式数值
          allowed:
          - 审题、模型设计、公式闭环和完整版代码生成
          - 不导入求解脚本的静态代码检查、依赖检查和语法检查
          - 读取用户返回的工作簿并执行Schema、哈希、质量门和证据一致性校验
          - 运行本Skill仓库自身的lint、单元测试、索引生成和LaTeX模板CI
        full_fidelity_flags:
          execution_profile: full_fidelity
          allow_reduced_data: false
          allow_coarser_grid: false
          allow_shorter_horizon: false
          allow_fewer_repetitions: false
          allow_relaxed_tolerance: false
          allow_silent_solver_fallback: false
        code_delivery:
          required_artifacts:
          - task_specific_python_code
          - full_run_config
          - execution_instructions
          required_config_fields:
          - execution_owner
          - execution_profile
          - stage
          - problem_name
          - code_path
          - code_sha256
          - data_paths
          - solver
          - solver_version
          - random_seed
          - tolerance
          - iteration_or_time_limit
          - expected_workbook
          - allow_reduced_data
          - allow_coarser_grid
          - allow_shorter_horizon
          - allow_fewer_repetitions
          - allow_relaxed_tolerance
          - allow_silent_solver_fallback
          no_placeholder_markers: [TODO, FIXME, __QUESTION_NAME__, NotImplementedError]
        execution_states:
          primary: [pending, code_delivered, awaiting_user_execution, workbook_received, accepted, rejected]
          analysis: [pending, code_delivered, awaiting_user_execution, workbook_received, accepted, rejected, redo_required]
          rules:
          - 代码交付不得把小问status提升为solved或analyzed。
          - primary_execution_status=accepted且主结果质量门通过后，status才可进入solved。
          - analysis_execution_status=accepted且结论稳定性校验通过后，status才可进入analyzed。
          - 主工作簿未验收前不得生成最终结果深化分析代码；只能形成候选分析方向。
        returned_workbook:
          required_sheet: 运行配置
          required_columns: [项目, 值]
          required_items:
          - execution_owner
          - execution_profile
          - stage
          - problem_name
          - code_sha256
          - data_sha256
          - solver
          - solver_version
          - tolerance
          - iteration_or_time_limit
          - actual_stop_reason
          - random_seed
          - repetitions_or_scenarios
          - grid_or_time_range
          - fallback_used
          - platform
          acceptance_rules:
          - execution_owner必须为user。
          - execution_profile必须为full_fidelity。
          - 所有缩减、放宽和静默回退标志必须为false。
          - fallback_used必须为false；需要替代求解器时必须形成新版代码与新版配置后重跑。
          - code_sha256必须与已交付代码完全一致。
          - 主结果工作簿必须通过主结果质量门，结果深化分析工作簿必须包含分析设计和结论稳定性汇总。
        filenames:
          primary_code: 问题X求解.py
          primary_config: 问题X完整运行配置.yaml
          primary_instructions: 问题X本地运行说明.md
          primary_workbook: 结果数据表/问题X/问题X求解结果.xlsx
          analysis_code: 问题X结果深化分析.py
          analysis_config: 问题X结果深化完整运行配置.yaml
          analysis_instructions: 问题X结果深化本地运行说明.md
          analysis_workbook: 结果数据表/问题X/问题X结果深化分析.xlsx
          code_delivery_report: code_delivery_report.yaml
          receipt_report: user_execution_validation_report.yaml
    ''')

    write("templates/code/full_fidelity_config.yaml", r'''
        execution_owner: user
        execution_profile: full_fidelity
        stage: primary
        problem_name: 问题一
        code_path: 问题一求解.py
        code_sha256: REPLACE_WITH_ACTUAL_SHA256
        data_paths: []
        data_sha256: REPLACE_WITH_ACTUAL_SHA256
        solver: REPLACE_WITH_SOLVER
        solver_version: REPLACE_WITH_VERSION
        random_seed: 2026
        tolerance: REPLACE_WITH_MODEL_SPECIFIC_TOLERANCE
        iteration_or_time_limit: REPLACE_WITH_FULL_RUN_LIMIT
        repetitions_or_scenarios: REPLACE_WITH_FULL_RUN_COUNT
        grid_or_time_range: REPLACE_WITH_FULL_RESOLUTION
        expected_workbook: 结果数据表/问题一/问题一求解结果.xlsx
        allow_reduced_data: false
        allow_coarser_grid: false
        allow_shorter_horizon: false
        allow_fewer_repetitions: false
        allow_relaxed_tolerance: false
        allow_silent_solver_fallback: false
    ''')

    write("templates/code/user_execution_instructions.md", r'''
        # 问题X本地完整运行说明

        1. 确认代码、数据和 `完整运行配置.yaml` 位于约定路径，且代码哈希与配置一致。
        2. 使用配置指定的完整数据、完整网格/时域、完整重复次数和模型专属容差运行代码。
        3. 不得为缩短时间改成抽样数据、粗网格、短时域、少场景、少随机种子或宽松容差。
        4. 求解器不可用或运行失败时停止并反馈错误；不得静默切换替代算法。
        5. 将生成的标准工作簿原样返回。工作簿必须含 `运行配置` 工作表及实际停止原因、平台、求解器版本和代码/数据哈希。
        6. 在工作簿验收通过前，不使用其中数值绘图、写论文或生成下一阶段最终分析代码。
    ''')

    write("scripts/validate_code_delivery.py", r'''
        #!/usr/bin/env python3
        """Validate a full-fidelity code handoff without importing or executing task code."""
        from __future__ import annotations

        import argparse
        import hashlib
        import re
        from pathlib import Path
        from typing import Any

        import yaml

        ROOT = Path(__file__).resolve().parents[1]
        CONTRACT = ROOT / "core" / "user_execution_contract.yaml"
        FALSE_FLAGS = (
            "allow_reduced_data", "allow_coarser_grid", "allow_shorter_horizon",
            "allow_fewer_repetitions", "allow_relaxed_tolerance",
            "allow_silent_solver_fallback",
        )
        PLACEHOLDERS = ("TODO", "FIXME", "__QUESTION_NAME__", "NotImplementedError")


        def load_yaml(path: Path) -> dict[str, Any]:
            return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


        def sha256(path: Path) -> str:
            return hashlib.sha256(path.read_bytes()).hexdigest()


        def infer_stage(config: dict[str, Any]) -> str:
            stage = str(config.get("stage", "")).strip()
            if stage not in {"primary", "analysis"}:
                raise ValueError("stage必须为primary或analysis")
            return stage


        def validate_config(project_root: Path, config_path: Path) -> tuple[list[str], dict[str, Any], Path]:
            issues: list[str] = []
            config = load_yaml(config_path)
            contract = load_yaml(CONTRACT)
            required = contract["code_delivery"]["required_config_fields"]
            for field in required:
                if field not in config or config[field] in (None, "", []):
                    issues.append(f"完整运行配置缺少字段: {field}")
            try:
                stage = infer_stage(config)
            except ValueError as exc:
                issues.append(str(exc))
                stage = "primary"
            if config.get("execution_owner") != "user":
                issues.append("execution_owner必须为user")
            if config.get("execution_profile") != "full_fidelity":
                issues.append("execution_profile必须为full_fidelity")
            for flag in FALSE_FLAGS:
                if config.get(flag) is not False:
                    issues.append(f"{flag}必须显式为false")
            code_path = project_root / str(config.get("code_path", ""))
            if not code_path.is_file():
                issues.append(f"代码文件不存在: {code_path}")
                return issues, config, code_path
            text = code_path.read_text(encoding="utf-8", errors="ignore")
            for marker in PLACEHOLDERS:
                if marker in text:
                    issues.append(f"正式代码仍含占位标记: {marker}")
            if "if __name__ == \"__main__\":" not in text and "if __name__ == '__main__':" not in text:
                issues.append("正式代码缺少main入口")
            declared = str(config.get("code_sha256", "")).lower()
            actual = sha256(code_path)
            if declared != actual:
                issues.append(f"code_sha256不匹配: declared={declared}, actual={actual}")
            expected_suffix = "结果深化分析.py" if stage == "analysis" else "求解.py"
            if not code_path.name.endswith(expected_suffix):
                issues.append(f"{stage}阶段代码文件名应以{expected_suffix}结尾")
            instructions = config_path.with_name(
                config_path.name.replace("完整运行配置.yaml", "本地运行说明.md")
            )
            if not instructions.is_file():
                issues.append(f"缺少本地运行说明: {instructions.name}")
            return issues, config, code_path


        def update_state(project_root: Path, config: dict[str, Any], code_path: Path) -> None:
            state_path = project_root / "state" / "project_state.yaml"
            if not state_path.is_file():
                return
            state = load_yaml(state_path)
            problem = str(config["problem_name"])
            order = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十"]
            suffix = problem.removeprefix("问题")
            key = f"Q{order.index(suffix) + 1}" if suffix in order else problem
            entry = state.setdefault("subproblems", {}).setdefault(key, {})
            relative = code_path.relative_to(project_root).as_posix()
            stage = infer_stage(config)
            if stage == "primary":
                entry["code"] = relative
                entry["primary_code_sha256"] = sha256(code_path)
                entry["primary_execution_status"] = "awaiting_user_execution"
                entry.setdefault("analysis_execution_status", "pending")
            else:
                if entry.get("primary_execution_status") != "accepted":
                    raise ValueError("主工作簿未accepted，禁止交付最终结果深化分析代码")
                entry["result_analysis_code"] = relative
                entry["analysis_code_sha256"] = sha256(code_path)
                entry["analysis_execution_status"] = "awaiting_user_execution"
            execution = state.setdefault("execution", {})
            execution.update({
                "owner": "user",
                "profile": "full_fidelity",
                "assistant_task_execution_allowed": False,
                **{flag: False for flag in FALSE_FLAGS},
            })
            state_path.write_text(yaml.safe_dump(state, allow_unicode=True, sort_keys=False), encoding="utf-8")


        def main() -> int:
            parser = argparse.ArgumentParser()
            parser.add_argument("project_root", type=Path)
            parser.add_argument("--config", type=Path)
            parser.add_argument("--write", action="store_true")
            parser.add_argument("--strict", action="store_true")
            args = parser.parse_args()
            root = args.project_root.resolve()
            configs = [args.config] if args.config else sorted(root.glob("问题*完整运行配置.yaml")) + sorted(root.glob("问题*结果深化完整运行配置.yaml"))
            issues: list[str] = []
            checked: list[str] = []
            for raw in configs:
                path = raw if raw.is_absolute() else root / raw
                item_issues, config, code_path = validate_config(root, path)
                issues.extend(f"{path.name}: {item}" for item in item_issues)
                checked.append(path.relative_to(root).as_posix())
                if args.write and not item_issues:
                    update_state(root, config, code_path)
            report = {"status": "passed" if not issues else "failed", "checked_configs": checked, "issues": issues, "task_code_executed": False}
            (root / "code_delivery_report.yaml").write_text(yaml.safe_dump(report, allow_unicode=True, sort_keys=False), encoding="utf-8")
            if issues:
                print("\n".join(issues))
                return 1 if args.strict else 0
            print("full-fidelity code delivery validated without executing task code")
            return 0


        if __name__ == "__main__":
            raise SystemExit(main())
    ''')

    write("scripts/validate_user_execution.py", r'''
        #!/usr/bin/env python3
        """Validate user-produced workbooks and advance state only after evidence passes."""
        from __future__ import annotations

        import argparse
        import hashlib
        from pathlib import Path
        from typing import Any

        import openpyxl
        import yaml

        FALSE_FLAGS = (
            "allow_reduced_data", "allow_coarser_grid", "allow_shorter_horizon",
            "allow_fewer_repetitions", "allow_relaxed_tolerance",
            "allow_silent_solver_fallback",
        )
        TRUE_VALUES = {True, 1, "1", "true", "yes", "是", "通过", "满足"}


        def load_yaml(path: Path) -> dict[str, Any]:
            return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


        def file_hash(path: Path) -> str:
            return hashlib.sha256(path.read_bytes()).hexdigest()


        def question_key(problem: str) -> str:
            order = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十"]
            suffix = problem.removeprefix("问题")
            return f"Q{order.index(suffix) + 1}" if suffix in order else problem


        def as_bool(value: Any) -> bool | None:
            if isinstance(value, bool):
                return value
            text = str(value).strip().lower()
            if text in {"true", "1", "yes", "是", "通过", "满足"}:
                return True
            if text in {"false", "0", "no", "否", "未通过", "不满足"}:
                return False
            return None


        def configuration_map(workbook: Path) -> tuple[dict[str, Any], list[str]]:
            issues: list[str] = []
            book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
            if "运行配置" not in book.sheetnames:
                return {}, ["缺少运行配置工作表"]
            sheet = book["运行配置"]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows or tuple(rows[0][:2]) != ("项目", "值"):
                return {}, ["运行配置表头必须为项目|值"]
            mapping = {str(row[0]).strip(): row[1] for row in rows[1:] if row and row[0] not in (None, "")}
            required = {
                "execution_owner", "execution_profile", "stage", "problem_name", "code_sha256",
                "data_sha256", "solver", "solver_version", "tolerance", "iteration_or_time_limit",
                "actual_stop_reason", "random_seed", "repetitions_or_scenarios", "grid_or_time_range",
                "fallback_used", "platform",
            }
            missing = sorted(required - set(mapping))
            issues.extend(f"运行配置缺少项目: {item}" for item in missing)
            return mapping, issues


        def quality_passed(workbook: Path) -> tuple[bool, list[str]]:
            book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
            if "主结果质量门" not in book.sheetnames:
                return False, ["缺少主结果质量门工作表"]
            rows = list(book["主结果质量门"].iter_rows(values_only=True))
            if not rows:
                return False, ["主结果质量门为空"]
            headers = [str(item) if item is not None else "" for item in rows[0]]
            if "是否通过" not in headers:
                return False, ["主结果质量门缺少是否通过列"]
            index = headers.index("是否通过")
            failures = [str(row[0]) for row in rows[1:] if row and as_bool(row[index] if len(row) > index else None) is not True]
            return not failures, [f"质量门未通过: {item}" for item in failures]


        def analysis_passed(workbook: Path) -> tuple[bool, str, list[str]]:
            book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
            required = {"分析设计", "结论稳定性汇总"}
            missing = sorted(required - set(book.sheetnames))
            if missing:
                return False, "failed", [f"缺少工作表: {missing}"]
            rows = list(book["结论稳定性汇总"].iter_rows(values_only=True))
            if len(rows) < 2:
                return False, "failed", ["结论稳定性汇总无实质数据"]
            headers = [str(item) if item is not None else "" for item in rows[0]]
            if "是否保持" not in headers:
                return False, "failed", ["结论稳定性汇总缺少是否保持列"]
            index = headers.index("是否保持")
            unstable = [row for row in rows[1:] if row and as_bool(row[index] if len(row) > index else None) is not True]
            return (not unstable, "passed" if not unstable else "redo_required", ["存在核心结论未保持"] if unstable else [])


        def validate_one(root: Path, workbook: Path, state: dict[str, Any], write: bool) -> list[str]:
            issues: list[str] = []
            config, config_issues = configuration_map(workbook)
            issues.extend(config_issues)
            stage = str(config.get("stage", ""))
            problem = str(config.get("problem_name", ""))
            key = question_key(problem)
            entry = (state.get("subproblems") or {}).get(key, {})
            if config.get("execution_owner") != "user":
                issues.append("execution_owner必须为user")
            if config.get("execution_profile") != "full_fidelity":
                issues.append("execution_profile必须为full_fidelity")
            if as_bool(config.get("fallback_used")) is not False:
                issues.append("fallback_used必须为false")
            for flag in FALSE_FLAGS:
                if flag in config and as_bool(config.get(flag)) is not False:
                    issues.append(f"{flag}必须为false")
            expected_hash = entry.get("analysis_code_sha256") if stage == "analysis" else entry.get("primary_code_sha256")
            if not expected_hash:
                issues.append("项目状态缺少已交付代码哈希")
            elif str(config.get("code_sha256", "")).lower() != str(expected_hash).lower():
                issues.append("工作簿code_sha256与已交付代码不一致")
            if stage == "primary":
                passed, quality_issues = quality_passed(workbook)
                issues.extend(quality_issues)
                if write:
                    entry["primary_execution_status"] = "accepted" if not issues and passed else "rejected"
                    entry["result_quality_status"] = "passed" if not issues and passed else "failed"
                    entry["solution_workbook"] = workbook.relative_to(root).as_posix()
                    entry.setdefault("artifact_hashes", {})["solution_workbook"] = file_hash(workbook)
                    if not issues and passed:
                        entry.setdefault("validated_artifact_hashes", {})["solution_workbook"] = file_hash(workbook)
                        entry["status"] = "solved"
            elif stage == "analysis":
                passed, result_status, analysis_issues = analysis_passed(workbook)
                issues.extend(analysis_issues)
                if write:
                    entry["analysis_execution_status"] = "accepted" if not issues and passed else ("redo_required" if result_status == "redo_required" else "rejected")
                    entry["result_analysis_status"] = result_status
                    entry["result_analysis_workbook"] = workbook.relative_to(root).as_posix()
                    entry.setdefault("artifact_hashes", {})["result_analysis_workbook"] = file_hash(workbook)
                    if not issues and passed:
                        entry.setdefault("validated_artifact_hashes", {})["result_analysis_workbook"] = file_hash(workbook)
                        entry["status"] = "analyzed"
                    elif result_status == "redo_required":
                        entry["artifacts_stale"] = True
                        entry["stale_layers"] = ["result_analysis_workbook", "matlab_script", "figure_bundle", "framework"]
                        entry["result_summary_status"] = "stale"
                        state.setdefault("project", {})["current_phase"] = "solve_validate"
            else:
                issues.append("运行配置stage必须为primary或analysis")
            return issues


        def discover(root: Path) -> list[Path]:
            return sorted((root / "结果数据表").glob("问题*/问题*求解结果.xlsx")) + sorted((root / "结果数据表").glob("问题*/问题*结果深化分析.xlsx"))


        def main() -> int:
            parser = argparse.ArgumentParser()
            parser.add_argument("project_root", type=Path)
            parser.add_argument("--workbook", type=Path)
            parser.add_argument("--scope", default="results")
            parser.add_argument("--write", action="store_true")
            parser.add_argument("--strict", action="store_true")
            args = parser.parse_args()
            root = args.project_root.resolve()
            state_path = root / "state" / "project_state.yaml"
            if not state_path.is_file():
                raise SystemExit("缺少state/project_state.yaml")
            state = load_yaml(state_path)
            workbooks = [args.workbook if args.workbook.is_absolute() else root / args.workbook] if args.workbook else discover(root)
            all_issues: list[str] = []
            checked: list[str] = []
            for workbook in workbooks:
                issues = validate_one(root, workbook, state, args.write)
                all_issues.extend(f"{workbook.name}: {item}" for item in issues)
                checked.append(workbook.relative_to(root).as_posix())
            if args.write:
                state_path.write_text(yaml.safe_dump(state, allow_unicode=True, sort_keys=False), encoding="utf-8")
            report = {"status": "passed" if not all_issues else "failed", "checked_workbooks": checked, "issues": all_issues, "task_code_executed": False}
            (root / "user_execution_validation_report.yaml").write_text(yaml.safe_dump(report, allow_unicode=True, sort_keys=False), encoding="utf-8")
            if all_issues:
                print("\n".join(all_issues))
                return 1 if args.strict else 0
            print("user-produced workbooks accepted without executing task code")
            return 0


        if __name__ == "__main__":
            raise SystemExit(main())
    ''')

    write("tests/test_user_execution_contract.py", r'''
        from __future__ import annotations

        import hashlib
        import importlib.util
        tempfile
        import unittest
        from pathlib import Path

        import openpyxl
        import yaml

        ROOT = Path(__file__).resolve().parents[1]


        def load_module(name: str, path: Path):
            spec = importlib.util.spec_from_file_location(name, path)
            module = importlib.util.module_from_spec(spec)
            assert spec.loader is not None
            spec.loader.exec_module(module)
            return module


        CODE = load_module("validate_code_delivery", ROOT / "scripts" / "validate_code_delivery.py")
        RECEIPT = load_module("validate_user_execution", ROOT / "scripts" / "validate_user_execution.py")


        class UserExecutionContractTests(unittest.TestCase):
            def make_project(self, root: Path) -> tuple[Path, Path]:
                (root / "state").mkdir()
                code = root / "问题一求解.py"
                code.write_text('def main():\n    return 0\n\nif __name__ == "__main__":\n    raise SystemExit(main())\n', encoding="utf-8")
                digest = hashlib.sha256(code.read_bytes()).hexdigest()
                config = root / "问题一完整运行配置.yaml"
                payload = {
                    "execution_owner": "user", "execution_profile": "full_fidelity", "stage": "primary",
                    "problem_name": "问题一", "code_path": code.name, "code_sha256": digest,
                    "data_paths": ["data.csv"], "data_sha256": "a" * 64, "solver": "test",
                    "solver_version": "1", "random_seed": 2026, "tolerance": 1e-8,
                    "iteration_or_time_limit": "full", "repetitions_or_scenarios": 100,
                    "grid_or_time_range": "full", "expected_workbook": "结果数据表/问题一/问题一求解结果.xlsx",
                    "allow_reduced_data": False, "allow_coarser_grid": False,
                    "allow_shorter_horizon": False, "allow_fewer_repetitions": False,
                    "allow_relaxed_tolerance": False, "allow_silent_solver_fallback": False,
                }
                config.write_text(yaml.safe_dump(payload, allow_unicode=True, sort_keys=False), encoding="utf-8")
                (root / "问题一本地运行说明.md").write_text("full run", encoding="utf-8")
                state = {
                    "project": {"competition": "test", "problem": "A", "current_phase": "solve_validate"},
                    "requirements": {"total": 0, "completed": [], "pending": []},
                    "decisions": {},
                    "subproblems": {"Q1": {"status": "designed", "selected_model": "m", "capabilities": {}, "result_quality_status": "pending", "result_analysis_status": "pending", "framework_section": "Q1", "result_summary_status": "pending"}},
                    "variables": {"locked": [], "source": {}},
                    "paper_framework": {"path": "模型论文框架.md", "version": "1", "sync_status": "stale", "last_sync_scope": "design", "proposition_limit": 4, "proposition_count": 0, "proposition_status": "not_assessed", "propositions": []},
                    "artifacts": {"code": [], "results": [], "figures": [], "papers": []},
                    "risks": [], "next_gate": {"module": "solve_validate", "condition": "code"},
                }
                (root / "state" / "project_state.yaml").write_text(yaml.safe_dump(state, allow_unicode=True, sort_keys=False), encoding="utf-8")
                return config, code

            def test_code_delivery_does_not_mark_solved(self):
                with tempfile.TemporaryDirectory() as temp:
                    root = Path(temp)
                    config, code = self.make_project(root)
                    issues, payload, path = CODE.validate_config(root, config)
                    self.assertEqual(issues, [])
                    CODE.update_state(root, payload, path)
                    state = yaml.safe_load((root / "state" / "project_state.yaml").read_text(encoding="utf-8"))
                    self.assertEqual(state["subproblems"]["Q1"]["status"], "designed")
                    self.assertEqual(state["subproblems"]["Q1"]["primary_execution_status"], "awaiting_user_execution")

            def test_reduced_flag_is_rejected(self):
                with tempfile.TemporaryDirectory() as temp:
                    root = Path(temp)
                    config, _ = self.make_project(root)
                    payload = yaml.safe_load(config.read_text(encoding="utf-8"))
                    payload["allow_reduced_data"] = True
                    config.write_text(yaml.safe_dump(payload, allow_unicode=True, sort_keys=False), encoding="utf-8")
                    issues, _, _ = CODE.validate_config(root, config)
                    self.assertTrue(any("allow_reduced_data" in item for item in issues))

            def test_primary_workbook_acceptance_marks_solved(self):
                with tempfile.TemporaryDirectory() as temp:
                    root = Path(temp)
                    config, code = self.make_project(root)
                    _, payload, path = CODE.validate_config(root, config)
                    CODE.update_state(root, payload, path)
                    result_dir = root / "结果数据表" / "问题一"
                    result_dir.mkdir(parents=True)
                    workbook = result_dir / "问题一求解结果.xlsx"
                    book = openpyxl.Workbook()
                    sheet = book.active
                    sheet.title = "运行配置"
                    sheet.append(["项目", "值"])
                    items = {
                        "execution_owner": "user", "execution_profile": "full_fidelity", "stage": "primary",
                        "problem_name": "问题一", "code_sha256": hashlib.sha256(code.read_bytes()).hexdigest(),
                        "data_sha256": "a" * 64, "solver": "test", "solver_version": "1",
                        "tolerance": 1e-8, "iteration_or_time_limit": "full", "actual_stop_reason": "optimal",
                        "random_seed": 2026, "repetitions_or_scenarios": 100, "grid_or_time_range": "full",
                        "fallback_used": False, "platform": "test",
                    }
                    for key, value in items.items():
                        sheet.append([key, value])
                    quality = book.create_sheet("主结果质量门")
                    quality.append(["检查项", "是否通过", "证据"])
                    quality.append(["完整运行", True, "ok"])
                    book.save(workbook)
                    state_path = root / "state" / "project_state.yaml"
                    state = yaml.safe_load(state_path.read_text(encoding="utf-8"))
                    issues = RECEIPT.validate_one(root, workbook, state, True)
                    self.assertEqual(issues, [])
                    self.assertEqual(state["subproblems"]["Q1"]["status"], "solved")
                    self.assertEqual(state["subproblems"]["Q1"]["primary_execution_status"], "accepted")


        if __name__ == "__main__":
            unittest.main()
    ''')


def rewrite_policy_modules_and_pack() -> None:
    write("core/hsk_core_policy.md", r'''
        # HSK Core Policy v6.5.0

        本文件只保存全局硬规则。用户执行合同以 `core/user_execution_contract.yaml` 为准；路由、产物图、输出、工作簿和项目状态分别以对应 `core/` 文件为准。

        ## 1. 总目标与优先级

        数学建模任务必须形成题意正确、机制闭合、数值可信、可复现和可审查的成果链。优先级为：

        $$
        \text{题意正确}>\text{机制与变量闭合}>\text{数据可信}>\text{完整版数值求解}>\text{结果证据}>\text{图表}>\text{论文表达}>\text{形式创新}.
        $$

        不能落地、不能解释、不能检验或不能复现的模型必须否决、降级或重构。

        ## 2. 默认执行所有权

        新项目默认采用 `execution_owner=user`、`execution_profile=full_fidelity`。助手负责审题、模型设计、公式闭环、生成可以直接本地运行的完整版 Python 代码、静态审查和返回工作簿验收；用户负责实际运行赛题主求解与结果深化分析代码。

        助手不得运行、导入或间接执行 `问题X求解.py`、`问题X结果深化分析.py`，不得为了计算时间缩减数据、网格、时域、场景、重复次数、随机种子、迭代次数或放宽容差，也不得静默切换求解器或轻量近似。仓库自身 lint、单元测试、索引生成和 LaTeX CI 不属于赛题数值执行，可以运行。

        ## 3. 默认工作顺序

        ```text
        审题与数据协议
        → 模型路线比较与模型锁定
        → 更新模型论文框架.md
        → 输出问题X求解.py、完整运行配置和本地运行说明
        → 状态停在awaiting_user_execution
        → 用户本地完整运行并返回问题X求解结果.xlsx
        → 验收运行配置、代码/数据哈希和主结果质量门
        → 仅在主工作簿accepted后设计并输出问题X结果深化分析.py
        → 用户本地完整运行并返回问题X结果深化分析.xlsx
        → 验收分析设计、稳定性结论和运行配置
        → MATLAB读取真实工作簿绘图
        → 直接编写并持续修改LaTeX
        → 编译与终审
        ```

        一个聊天不能伪造越过用户执行门。完整工作流允许在两个执行门处暂停，用户返回工作簿后从当前状态继续。

        ## 4. 代码交付门

        每次正式代码交付必须同时包含：题目专属 Python 代码、完整版运行配置、本地运行说明和代码交付报告。配置必须显式记录求解器、版本、随机种子、模型专属容差、完整迭代/时间限制、完整场景或重复次数、完整网格/时域、预期工作簿和代码/数据哈希。

        所有 `allow_reduced_*`、`allow_coarser_grid`、`allow_shorter_horizon`、`allow_fewer_repetitions`、`allow_relaxed_tolerance` 和 `allow_silent_solver_fallback` 必须为 `false`。正式代码不得含 TODO、FIXME、`__QUESTION_NAME__` 或 `NotImplementedError` 占位。

        代码交付只把执行状态改为 `awaiting_user_execution`，不得把小问提升为 `solved` 或 `analyzed`。

        ## 5. 用户返回工作簿验收

        两类工作簿都必须包含 `运行配置` 工作表，记录实际求解器版本、停止原因、平台、容差、随机种子、重复/场景、网格/时域、fallback 状态以及代码/数据哈希。工作簿中的 `code_sha256` 必须与已交付代码一致，`fallback_used` 必须为 `false`。

        主工作簿还必须通过 `主结果质量门`，之后 `primary_execution_status` 才可变为 `accepted`、小问状态才可进入 `solved`。主工作簿未验收前，只能列出候选深化方向，不得生成依赖实际结果的最终深化分析代码。

        分析工作簿必须包含 `分析设计`、至少一个实质分析表和 `结论稳定性汇总`。通过后 `analysis_execution_status=accepted` 且状态进入 `analyzed`；若核心结论在合理变化下失效，必须标记 `redo_required` 和下游 stale，回退模型设计或主求解。

        ## 6. 事实源与软件职责

        - 模型语义和论文结构：`模型论文框架.md`；
        - 机器状态、执行所有权、哈希和 stale：`state/project_state.yaml`；
        - 主数值事实：用户返回并验收的 `问题X求解结果.xlsx`；
        - 稳定范围与失效边界：用户返回并验收的 `问题X结果深化分析.xlsx`。

        Python 代码负责完整数据处理、主求解、质量门、结果深化分析和工作簿输出，但由用户本地执行。MATLAB 只读取验收后的真实工作簿绘制正式结果图；LaTeX 负责终稿；DOCX 仅显式按需。

        ## 7. 正式交付同步

        - 代码交付使用 `scripts/validate_code_delivery.py`，不得执行赛题代码；
        - 用户返回工作簿使用 `scripts/validate_user_execution.py`，只读工作簿并在通过后更新执行状态；
        - 图表、论文和提交包继续使用 `scripts/sync_project.py` 检查完整证据链。

        `results`、`figures`、`docx`、`latex` 和 `submission` 交付仍要求两类工作簿通过且下游非 stale。缺失用户返回结果时必须暂停，而不是生成示意数值。

        ## 8. 图表、写作与终审

        MATLAB 字段定位采用精确表头唯一匹配，不得模糊匹配或根据摘要反推数据。LaTeX 只能引用验收后的数值和已批准图表。终稿必须核对题意覆盖、模型闭合、执行配置、主结果质量门、结果深化选择理由、代码—工作簿—MATLAB—图表—正文证据链和编译状态。
    ''')

    write("modules/03_solve_validate.md", r'''
        # Module 03A：完整版主求解代码交付与用户执行门

        本模块生成当前题意、数据和模型口径下可直接本地运行的完整版主求解代码，但助手不得运行、导入或间接执行该代码。

        ## 主链

        ```text
        锁定数据协议、变量、公式、目标和约束
        → 生成问题X求解.py
        → 生成问题X完整运行配置.yaml
        → 生成问题X本地运行说明.md
        → 静态检查代码、依赖、哈希、占位符和完整精度标志
        → primary_execution_status = awaiting_user_execution
        → 用户本地运行
        → 返回问题X求解结果.xlsx
        → 验收运行配置与主结果质量门
        ```

        正式代码必须包含完整数据检查、模型、求解器状态、容差、停止条件、约束/残差、收敛或外样本检查、随机种子、底层结果和标准工作簿输出。禁止演示数据、抽样运行、粗网格、短时域、少场景、少重复、宽容差和静默 fallback。

        ## 代码交付

        固定交付：

        - `问题X求解.py`；
        - `问题X完整运行配置.yaml`；
        - `问题X本地运行说明.md`；
        - `code_delivery_report.yaml`；
        - 当前版 `模型论文框架.md`。

        使用 `scripts/validate_code_delivery.py` 做静态交付检查。代码交付不得生成工作簿，不得把状态标记为 solved。

        ## 工作簿验收

        用户返回的 `问题X求解结果.xlsx` 必须含 `运行配置`、`核心指标`、`数据审计`、`主结果质量门` 和题型专项底层表。`运行配置` 中的代码哈希必须匹配已交付代码，所有缩减和回退标志必须为 false。

        只有 `scripts/validate_user_execution.py` 验收通过后，`primary_execution_status=accepted`、`result_quality_status=passed`、状态进入 solved，并允许生成最终结果深化分析代码。
    ''')

    write("modules/03_result_analysis.md", r'''
        # Module 03B：结果深化分析代码交付与用户执行门

        本模块只接受已经由用户返回且通过主结果验收的工作簿。主工作簿未 accepted 时，只能形成候选分析方向，不得生成依赖实际结果的最终分析代码。

        ## 风险驱动设计

        读取真实主结果后，按题目、模型、数据、数值表现和评委风险选择参数敏感性、场景压力、多算法/多初值、结构稳健性、阈值边界、异质性、误差分解或外样本稳定性。禁止所有题统一 ±5%、±10% 扰动。

        ## 代码交付主链

        ```text
        主工作簿accepted
        → 建立result_analysis_plan
        → 生成问题X结果深化分析.py
        → 生成问题X结果深化完整运行配置.yaml
        → 生成问题X结果深化本地运行说明.md
        → 静态交付检查
        → analysis_execution_status = awaiting_user_execution
        → 用户本地完整运行并返回分析工作簿
        → 验收后进入analyzed或redo_required
        ```

        助手不得运行结果深化代码，不得为了获得快速图表而减少场景、重复、算法、参数范围或网格。

        用户返回的工作簿必须包含 `运行配置`、`分析设计`、至少一个实质分析表和 `结论稳定性汇总`。若核心结论未保持，必须进入 `redo_required`，标记下游 stale 并回退重算；不得包装旧结果继续绘图或写作。
    ''')

    write("packs/artifact/code.md", r'''
        # Artifact Pack：用户执行的完整版代码

        ## 进入条件

        用户要求 Python 求解代码、完整求解、结果深化代码或可复现代码附件时加载。代码必须对应已锁定模型，助手只生成和静态检查，不运行赛题代码。

        ## 完整精度硬规则

        每份正式代码必须使用 `execution_owner=user`、`execution_profile=full_fidelity`，并显式禁止缩减数据、粗化网格、缩短时域、减少重复/场景、放宽容差和静默求解器 fallback。运行失败时应明确失败，不得自动切换轻量近似。

        ## 主求解交付

        ```text
        问题X求解.py
        问题X完整运行配置.yaml
        问题X本地运行说明.md
        code_delivery_report.yaml
        ```

        主代码完整保留数据审计、目标函数、约束、求解器、停止条件、容差、约束/残差、收敛、外样本或可识别性检查，并输出包含 `运行配置` 的标准主工作簿。静态检查通过后状态停在 `awaiting_user_execution`。

        ## 结果深化交付

        只有用户返回的主工作簿通过验收后，才依据真实结果生成：

        ```text
        问题X结果深化分析.py
        问题X结果深化完整运行配置.yaml
        问题X结果深化本地运行说明.md
        ```

        分析方法由具体风险驱动，代码输出包含 `运行配置`、`分析设计`、实质分析底层表和 `结论稳定性汇总`。助手仍不运行该代码。

        ## 验收

        - 代码无 TODO、FIXME、占位问题名或 NotImplementedError；
        - 代码哈希、数据哈希和配置闭合；
        - 代码交付不产生 solved/analyzed 状态；
        - 用户工作簿的运行配置与已交付代码一致；
        - fallback_used=false；
        - 主质量门通过后才生成最终分析代码；
        - 两类工作簿验收后才进入 MATLAB 和 LaTeX。
    ''')


def patch_authoritative_files() -> None:
    replace_once("core/bootstrap.yaml", "  output: core/output_contract.yaml\n", "  output: core/output_contract.yaml\n  user_execution: core/user_execution_contract.yaml\n")
    replace_once("core/bootstrap.yaml", "- Python负责数据处理、主求解、质量门、结果深化分析与工作簿；MATLAB只读取工作簿绘制正式结果图。", "- 新项目默认由用户本地执行完整版主求解与结果深化代码；助手只生成代码、静态校验和验收返回工作簿，不得运行赛题代码。\n- 禁止自动缩减数据、网格、时域、场景、重复次数、迭代或容差，禁止静默求解器fallback和轻量近似。\n- Python代码负责完整数据处理、主求解、质量门、结果深化分析与工作簿；MATLAB只读取验收后的工作簿绘图。")
    replace_once("core/bootstrap.yaml", "  review: python scripts/score_submission.py\n", "  review: python scripts/score_submission.py\n  validate_code_delivery: python scripts/validate_code_delivery.py\n  validate_user_execution: python scripts/validate_user_execution.py\n")

    replace_once("core/output_contract.yaml", "schema: core/workbook_schema.yaml\n", "schema: core/workbook_schema.yaml\nuser_execution_contract: core/user_execution_contract.yaml\n")
    replace_once("core/output_contract.yaml", "result_policy:\n", dedent(r'''
        execution_policy:
          default_owner: user
          default_profile: full_fidelity
          assistant_task_code_execution_allowed: false
          code_delivery_gate: scripts/validate_code_delivery.py
          returned_workbook_gate: scripts/validate_user_execution.py
          pause_states: [awaiting_user_execution]
          code_delivery_does_not_promote_result_status: true
          analysis_code_requires_primary_workbook_accepted: true
          full_fidelity_flags_must_be_false: [allow_reduced_data, allow_coarser_grid, allow_shorter_horizon, allow_fewer_repetitions, allow_relaxed_tolerance, allow_silent_solver_fallback]
        result_policy:
    '''))
    replace_once("core/output_contract.yaml", "  delivery_scopes: [design, results, figures, docx, latex, submission]", "  delivery_scopes: [design, code, results, figures, docx, latex, submission]")
    replace_once("core/output_contract.yaml", "    solve_validate: results\n    result_analysis: results", "    solve_validate: code\n    result_analysis: code")
    replace_once("core/output_contract.yaml", "    design: [project_state, model_paper_framework]\n    results:", "    design: [project_state, model_paper_framework]\n    code: [project_state, model_paper_framework, python_code, full_run_config, execution_instructions, code_delivery_report]\n    results:")
    replace_once("core/output_contract.yaml", "  Python: 主求解阶段完成高精度求解与质量门；随后按题设计结果深化分析并输出对应工作簿", "  Python: 生成可由用户本地执行的完整版主求解与结果深化代码；助手不得运行赛题代码，工作簿由用户执行后返回验收")
    replace_once("core/output_contract.yaml", "- 每次正式模型、代码、工作簿、图表、DOCX、LaTeX或提交包交付前必须成功执行project_sync gate并生成sync_report。", "- 正式代码交付必须先通过validate_code_delivery且不得执行赛题代码；图表、DOCX、LaTeX和提交包继续通过project_sync。\n- 用户返回工作簿必须先通过validate_user_execution；代码交付本身不得把状态提升为solved或analyzed。")

    schema = ROOT / "core/project_state.schema.yaml"
    text = schema.read_text(encoding="utf-8")
    text = text.replace("        code: {type: string}\n", dedent('''
                code: {type: string}
                result_analysis_code: {type: string}
                primary_code_sha256: {type: string, pattern: '^[0-9a-fA-F]{64}$'}
                analysis_code_sha256: {type: string, pattern: '^[0-9a-fA-F]{64}$'}
                primary_execution_status:
                  type: string
                  enum: [pending, code_delivered, awaiting_user_execution, workbook_received, accepted, rejected]
                analysis_execution_status:
                  type: string
                  enum: [pending, code_delivered, awaiting_user_execution, workbook_received, accepted, rejected, redo_required]
        '''), 1)
    text = text.replace("      python_version: {type: string}\n", dedent('''
              owner: {type: string, enum: [user, assistant, external]}
              profile: {type: string, enum: [full_fidelity]}
              assistant_task_execution_allowed: {type: boolean}
              allow_reduced_data: {type: boolean}
              allow_coarser_grid: {type: boolean}
              allow_shorter_horizon: {type: boolean}
              allow_fewer_repetitions: {type: boolean}
              allow_relaxed_tolerance: {type: boolean}
              allow_silent_solver_fallback: {type: boolean}
              python_version: {type: string}
        '''), 1)
    schema.write_text(text, encoding="utf-8", newline="\n")

    replace_once("core/workflow_router.yaml", "  - Primary solving must pass the result-quality gate before result analysis starts.", "  - Generate full-fidelity task code and stop at awaiting_user_execution; never execute task-specific solve or analysis code.\n  - Primary user-produced workbook must be accepted before final result-analysis code is generated.")
    regex_replace("core/workflow_router.yaml", r"  full_solution:\n.*?(?=  full_workflow:)", dedent(r'''
          full_solution:
            triggers: [完整求解, 全部计算, 求出各问, 完成求解]
            infer_keywords: [完整求解, 全部计算, 求出各问, 完成求解]
            load: [modules/01_problem_audit.md, packs/task/classifier.md]
            then: [modules/02_model_design.md, modules/03_solve_validate.md, packs/artifact/code.md]
            load_classified_task_packs: true
            load_competition_pack: true
            terminal_outputs: [python_code, full_run_config, execution_instructions, code_delivery_report, awaiting_user_execution, proposition_plan, model_paper_framework]
            formal_delivery: true
            delivery_scope: code
            pause_for_user_execution: true
    '''))
    regex_replace("core/workflow_router.yaml", r"  full_workflow:\n.*?(?=  problem_analysis:)", dedent(r'''
          full_workflow:
            triggers: [全流程, 完整论文, 全套成果, 完整交付]
            infer_keywords: [全流程, 完整论文, 全套成果, 完整交付]
            load: [modules/01_problem_audit.md, packs/task/classifier.md]
            then: [modules/02_model_design.md, modules/03_solve_validate.md, packs/artifact/code.md]
            load_classified_task_packs: true
            load_competition_pack: true
            terminal_outputs: [python_code, full_run_config, execution_instructions, code_delivery_report, awaiting_user_execution, model_paper_framework]
            formal_delivery: true
            delivery_scope: code
            pause_for_user_execution: true
    '''))
    regex_replace("core/workflow_router.yaml", r"  code_and_solution:\n.*?(?=  result_analysis:)", dedent(r'''
          code_and_solution:
            triggers: [代码, Python, 求解, 算法, 最优解]
            infer_keywords: [Python代码, 求解代码, 开始求解, 求解, 继续求解, 重新求解, 重算, 算法实现, 最优解]
            load: [modules/03_solve_validate.md, packs/artifact/code.md, packs/task/classifier.md]
            load_classified_task_packs: true
            terminal_outputs: [python_code, full_run_config, execution_instructions, code_delivery_report, awaiting_user_execution, model_paper_framework]
            formal_delivery: true
            delivery_scope: code
            pause_for_user_execution: true
    '''))
    regex_replace("core/workflow_router.yaml", r"  result_analysis:\n.*?(?=  validation:)", dedent(r'''
          result_analysis:
            triggers: [结果分析, 深化分析, 敏感性, 鲁棒性, 多算法, 稳健性, 阈值分析, 结构稳健性, 异质性]
            infer_keywords: [结果分析, 敏感性分析, 鲁棒性分析, 多算法验证, 稳健性分析, 阈值分析, 结构稳健性, 异质性分析]
            load: [modules/03_result_analysis.md, packs/artifact/code.md, packs/task/classifier.md]
            load_classified_task_packs: true
            terminal_outputs: [result_analysis_plan, result_analysis_code, full_run_config, execution_instructions, code_delivery_report, awaiting_user_execution, model_paper_framework]
            formal_delivery: true
            delivery_scope: code
            pause_for_user_execution: true
          returned_workbook_validation:
            triggers: [返回工作簿, 验收工作簿, 校验求解结果, 校验结果深化分析, 本地运行完成]
            infer_keywords: [返回工作簿, 验收工作簿, 本地运行完成, 校验求解结果]
            load: [scripts/validate_user_execution.py]
            pre_delivery_gates: [user_execution_receipt]
            terminal_outputs: [user_execution_validation_report, project_state]
            formal_delivery: false
    '''))
    replace_once("core/workflow_router.yaml", "- Primary solving precedes result analysis; result analysis precedes figures and writing.", "- Primary code delivery pauses for user execution; accepted primary workbooks precede result-analysis code delivery.\n- Accepted result-analysis workbooks precede figures and writing.")

    replace_once("core/module_manifest.yaml", "- discovered_artifacts\n", "- discovered_artifacts\n- accepted_solution_workbook\n- accepted_result_analysis_workbook\n")
    replace_once("core/module_manifest.yaml", "  python_code: Python主求解与结果分析代码\n", "  python_code: 用户本地执行的完整版Python主求解代码\n  result_analysis_code: 用户本地执行的完整版结果深化分析代码\n  full_run_config: 完整精度运行参数、代码与数据哈希和禁止降级标志\n  execution_instructions: 用户本地运行与返回工作簿说明\n  code_delivery_report: 不执行赛题代码的静态交付检查报告\n  awaiting_user_execution: 等待用户本地运行并返回工作簿的暂停状态\n  user_execution_validation_report: 用户返回工作簿的运行配置、哈希和质量门验收报告\n")
    regex_replace("core/module_manifest.yaml", r"  full_solution:\n.*?(?=  full_workflow:)", dedent(r'''
          full_solution:
            modules: [problem_audit, model_design, solve_validate]
            module_terminal_outputs: [python_code, full_run_config, execution_instructions, code_delivery_report, awaiting_user_execution, proposition_plan, model_paper_framework]
            pre_delivery_gates: [code_delivery]
            terminal_outputs: [python_code, full_run_config, execution_instructions, code_delivery_report, awaiting_user_execution, proposition_plan, model_paper_framework, project_state]
    '''))
    regex_replace("core/module_manifest.yaml", r"  full_workflow:\n.*?(?=modules:)", dedent(r'''
          full_workflow:
            modules: [problem_audit, model_design, solve_validate]
            module_terminal_outputs: [python_code, full_run_config, execution_instructions, code_delivery_report, awaiting_user_execution, model_paper_framework]
            pre_delivery_gates: [code_delivery]
            terminal_outputs: [python_code, full_run_config, execution_instructions, code_delivery_report, awaiting_user_execution, model_paper_framework, project_state]
        modules:
    '''))
    regex_replace("core/module_manifest.yaml", r"  solve_validate:\n.*?(?=  result_analysis:)", dedent(r'''
          solve_validate:
            path: modules/03_solve_validate.md
            inputs: [locked_model_spec, formula_closure, proposition_plan, model_paper_framework, data_schema, data, workbook_schema]
            outputs: [python_code, full_run_config, execution_instructions, code_delivery_report, awaiting_user_execution, model_paper_framework]
    '''))
    regex_replace("core/module_manifest.yaml", r"  result_analysis:\n.*?(?=  figure_evidence:)", dedent(r'''
          result_analysis:
            path: modules/03_result_analysis.md
            inputs: [accepted_solution_workbook, result_quality_report, validation_plan, model_paper_framework, workbook_schema]
            outputs: [result_analysis_plan, result_analysis_code, full_run_config, execution_instructions, code_delivery_report, awaiting_user_execution, model_paper_framework]
    '''))
    replace_once("core/module_manifest.yaml", "utility_gates:\n  project_sync:", dedent(r'''
        utility_gates:
          code_delivery:
            path: scripts/validate_code_delivery.py
            command: python scripts/validate_code_delivery.py <project_root> --write --strict
            inputs: [existing_project_state, existing_model_paper_framework, python_code, full_run_config, execution_instructions]
            outputs: [code_delivery_report, project_state, awaiting_user_execution]
          user_execution_receipt:
            path: scripts/validate_user_execution.py
            command: python scripts/validate_user_execution.py <project_root> --write --strict
            inputs: [existing_project_state, discovered_artifacts]
            outputs: [user_execution_validation_report, project_state, solution_workbook, result_analysis_workbook]
          project_sync:
    '''))
    replace_once("core/module_manifest.yaml", "    delivery_scopes: [design, results, figures, docx, latex, submission]", "    delivery_scopes: [design, code, results, figures, docx, latex, submission]")

    replace_once("scripts/resolve_workflow.py", "SCOPE_RANK = {\"design\": 0, \"results\": 1, \"figures\": 2, \"docx\": 3, \"latex\": 4, \"submission\": 5}", "SCOPE_RANK = {\"design\": 0, \"code\": 1, \"results\": 2, \"figures\": 3, \"docx\": 4, \"latex\": 5, \"submission\": 6}")
    replace_once("scripts/resolve_workflow.py", "    explicit_gates: list[str] = []\n", "    explicit_gates: list[str] = []\n    pause_for_user_execution = False\n")
    replace_once("scripts/resolve_workflow.py", "        explicit_gates.extend(route.get(\"pre_delivery_gates\", []))\n", "        explicit_gates.extend(route.get(\"pre_delivery_gates\", []))\n        pause_for_user_execution = pause_for_user_execution or bool(route.get(\"pause_for_user_execution\"))\n")
    replace_once("scripts/resolve_workflow.py", "        \"sync_required_before_delivery\": any(gate[\"name\"] == \"project_sync\" for gate in gates),\n", "        \"sync_required_before_delivery\": any(gate[\"name\"] == \"project_sync\" for gate in gates),\n        \"pause_for_user_execution\": pause_for_user_execution,\n        \"task_code_execution_allowed\": False,\n")

    replace_once("core/project_state.schema.yaml", "        result_quality_status: {type: string, enum: [pending, passed, failed]}", "        result_quality_status: {type: string, enum: [pending, passed, failed]}\n        execution_note: {type: string, description: 代码交付与用户本地运行状态说明}")


def patch_templates_docs_and_changelog() -> None:
    replace_once("templates/code/hsk_pipeline/main_pipeline.py", "    random_seed: int = 2026\n", dedent('''
            random_seed: int = 2026
            execution_owner: Literal["user"] = "user"
            execution_profile: Literal["full_fidelity"] = "full_fidelity"
            allow_reduced_data: bool = False
            allow_coarser_grid: bool = False
            allow_shorter_horizon: bool = False
            allow_fewer_repetitions: bool = False
            allow_relaxed_tolerance: bool = False
            allow_silent_solver_fallback: bool = False
    '''))
    replace_once("templates/code/hsk_pipeline/main_pipeline.py", "        if not self.framework_path.is_file():\n", dedent('''
            if self.execution_owner != "user" or self.execution_profile != "full_fidelity":
                raise ValueError("v6.5.0正式代码必须由用户以full_fidelity模式执行")
            forbidden_flags = {
                "allow_reduced_data": self.allow_reduced_data,
                "allow_coarser_grid": self.allow_coarser_grid,
                "allow_shorter_horizon": self.allow_shorter_horizon,
                "allow_fewer_repetitions": self.allow_fewer_repetitions,
                "allow_relaxed_tolerance": self.allow_relaxed_tolerance,
                "allow_silent_solver_fallback": self.allow_silent_solver_fallback,
            }
            enabled = sorted(name for name, value in forbidden_flags.items() if value)
            if enabled:
                raise ValueError(f"完整版运行禁止启用降级标志: {enabled}")
            if not self.framework_path.is_file():
    '''))

    for relative in ("SKILL.md", "skills/mathmodel-skill/SKILL.md"):
        path = ROOT / relative
        text = path.read_text(encoding="utf-8")
        marker = "## 启动\n"
        insertion = dedent('''
            ## v6.5.0 默认执行方式

            赛题数值代码默认由用户本地以 `full_fidelity` 模式运行。助手输出完整版代码、运行配置和说明后停在 `awaiting_user_execution`；用户返回主工作簿并验收后，才输出最终结果深化分析代码。助手不得运行赛题代码或自动采用轻量近似。

        ''')
        if insertion not in text:
            text = text.replace(marker, insertion + marker, 1)
        path.write_text(text, encoding="utf-8", newline="\n")

    for relative in ("README.md", "PROJECT_INSTRUCTIONS.md", "RUNTIME_ROUTER.md", "AGENTS.md"):
        path = ROOT / relative
        text = path.read_text(encoding="utf-8")
        heading = "\n## v6.5.0 用户执行完整版代码\n"
        block = dedent('''

            ## v6.5.0 用户执行完整版代码

            默认不由助手运行赛题主求解或结果深化分析程序。助手交付题目专属完整版代码、完整运行配置和本地说明，用户运行后返回标准工作簿；工作簿通过运行配置、代码/数据哈希和质量门验收后，工作流才继续。禁止自动降采样、粗网格、短时域、少重复、宽容差、静默求解器 fallback 或用轻量结果代替正式结果。
        ''')
        if heading not in text:
            text += block
        path.write_text(text, encoding="utf-8", newline="\n")

    path = ROOT / "agents/openai.yaml"
    text = path.read_text(encoding="utf-8")
    text = text.replace("Python solves and writes two standard workbooks", "Generate full-fidelity Python code for user execution; accept the two standard workbooks only after the user returns them")
    if "task_code_execution_allowed: false" not in text:
        text += "\nuser_execution:\n  default_owner: user\n  profile: full_fidelity\n  task_code_execution_allowed: false\n"
    path.write_text(text, encoding="utf-8", newline="\n")

    changelog = ROOT / "CHANGELOG.md"
    text = changelog.read_text(encoding="utf-8")
    entry = dedent('''
        ## Current release: 6.5.0

        - Default execution ownership is now user-managed full-fidelity: the assistant generates task-specific code but never runs solve or result-analysis programs.
        - Added formal code-delivery and returned-workbook gates, execution states, full-run configuration, code/data hash checks, and the mandatory `运行配置` workbook evidence contract.
        - Primary code delivery pauses at `awaiting_user_execution`; final result-analysis code is generated only after the returned primary workbook is accepted.
        - Existing local pipelines remain runnable by the user; legacy projects without the new optional execution fields remain readable.

        ## Previous release: 6.4.1
    ''')
    text = text.replace("## Current release: 6.4.1", entry, 1)
    changelog.write_text(text, encoding="utf-8", newline="\n")


def bump_versions() -> None:
    generated = {"SKILL_FILE_INDEX.md", "TEMPLATE_INDEX.md", "HSK_SKILL_FILE_INDEX_V622.md", "HSK_TEMPLATE_INDEX_V622.md", "MANIFEST.sha256"}
    for path in ROOT.rglob("*"):
        if not path.is_file():
            continue
        relative = path.relative_to(ROOT)
        if relative.as_posix() in generated or relative.parts[0] == "legacy":
            continue
        if relative.name.startswith("CHANGELOG_V") or relative.name in {"CHANGELOG.md", "SKILL_CHANGE_GOVERNANCE.md", "apply_v650_once.py"}:
            continue
        if path.suffix.lower() not in {".md", ".yaml", ".yml", ".json", ".py"}:
            continue
        try:
            text = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            continue
        if OLD in text:
            path.write_text(text.replace(OLD, NEW), encoding="utf-8", newline="\n")


def main() -> int:
    create_contracts_and_tools()
    rewrite_policy_modules_and_pack()
    patch_authoritative_files()
    patch_templates_docs_and_changelog()
    bump_versions()
    WORKFLOW.write_text(ORIGINAL_WORKFLOW, encoding="utf-8", newline="\n")
    Path(__file__).unlink()
    print("v6.5.0 user-executed full-fidelity migration applied")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
